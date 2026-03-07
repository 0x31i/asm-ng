#!/usr/bin/env python3
"""
ASM-NG Final Consolidation
Merges all 7 group findings into a single consolidated output.
Produces: CONSOLIDATED_FINDINGS.csv, CONSOLIDATED_FINDINGS.xlsx, EXECUTIVE_SUMMARY.md
"""

import csv
import os
from collections import defaultdict, Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed - Excel output will be skipped")

OUTPUT_DIR = './output'
FINAL_DIR = OUTPUT_DIR  # Put consolidated files in same output dir

# ============================================================================
# GRADE CONFIGURATION (from grade_config.py)
# ============================================================================
GRADE_CATEGORIES = {
    'Network Security': 1.0,
    'Web App Security': 1.0,
    'Information Leakage': 0.8,
    'General Health': 0.8,
    'External Account Exposure': 0.7,
    'DNS Health': 0.7,
    'IP Reputation': 0.6,
}

SEVERITY_POINTS = {
    'Critical': -20,
    'High': -10,
    'Medium': -5,
    'Low': -1,
    'Info': 0,
}

def calculate_grade(score):
    if score >= 90: return 'A'
    if score >= 80: return 'B'
    if score >= 70: return 'C'
    if score >= 60: return 'D'
    return 'F'

# ============================================================================
# LOAD ALL GROUP FINDINGS
# ============================================================================
print("=== Loading Group Findings ===")

all_findings = []
group_stats = {}
fieldnames = ['Category', 'Tab', 'Priority', 'Item', 'Description', 'Recommendation', 'Tracking_Status', 'Avg_Risk']

for g in range(1, 8):
    fpath = os.path.join(OUTPUT_DIR, f'findings_GROUP{g}_FULL_CONSOLIDATED.csv')
    if not os.path.exists(fpath):
        print(f"  GROUP {g}: MISSING - {fpath}")
        continue
    with open(fpath, 'r') as f:
        reader = csv.DictReader(f)
        group_findings = list(reader)
        all_findings.extend(group_findings)

        # Count by severity
        sev_counts = Counter(f.get('Priority', 'Unknown') for f in group_findings)
        group_stats[g] = {
            'total': len(group_findings),
            'critical': sev_counts.get('Critical', 0),
            'high': sev_counts.get('High', 0),
            'medium': sev_counts.get('Medium', 0),
            'low': sev_counts.get('Low', 0),
        }
        print(f"  GROUP {g}: {len(group_findings)} findings loaded")

print(f"\nTotal findings: {len(all_findings)}")

# ============================================================================
# CALCULATE CATEGORY SCORES AND GRADES
# ============================================================================
print("\n=== Calculating Grades ===")

category_scores = {}
for cat_name, weight in GRADE_CATEGORIES.items():
    cat_findings = [f for f in all_findings if f.get('Category', '') == cat_name]

    # Calculate raw score from severity points
    raw_score = 0
    for f in cat_findings:
        priority = f.get('Priority', 'Medium')
        # For consolidated findings (multi-instance), extract count from Item
        item = f.get('Item', '')
        instance_count = 1
        import re
        count_match = re.search(r'(\d+) instances', item)
        if count_match:
            instance_count = int(count_match.group(1))

        points = SEVERITY_POINTS.get(priority, -5)
        raw_score += points * min(instance_count, 10)  # Cap at 10 to prevent extreme scores

    # Score formula: max(0, 100 + (raw_score * weight))
    score = max(0, 100 + (raw_score * weight))
    score = min(100, score)  # Cap at 100
    grade = calculate_grade(score)

    category_scores[cat_name] = {
        'weight': weight,
        'findings': len(cat_findings),
        'raw_score': raw_score,
        'score': round(score, 1),
        'grade': grade,
    }
    print(f"  {cat_name}: Score={score:.1f}, Grade={grade}, Findings={len(cat_findings)}, Raw={raw_score}")

# Overall weighted grade
total_weight = sum(GRADE_CATEGORIES.values())
weighted_sum = sum(category_scores[cat]['score'] * category_scores[cat]['weight'] for cat in category_scores)
overall_score = round(weighted_sum / total_weight, 1)
overall_grade = calculate_grade(overall_score)
print(f"\n  OVERALL: Score={overall_score}, Grade={overall_grade}")

# ============================================================================
# SEVERITY BREAKDOWN
# ============================================================================
severity_counts = Counter()
for f in all_findings:
    severity_counts[f.get('Priority', 'Unknown')] += 1

# ============================================================================
# QUALITY FAILSAFE — Check descriptions and recommendations are specific
# ============================================================================
print("\n=== Quality Failsafe Check ===")

import sys as _sys
_sys.path.insert(0, './scripts')
try:
    from finding_enrichment import enrich_finding as _enrich_finding
    _HAS_ENRICH = True
    print("  finding_enrichment library loaded OK")
except ImportError:
    _HAS_ENRICH = False
    print("  WARNING: finding_enrichment.py not found — auto-fix unavailable")

# Phrases that indicate a description/recommendation is still generic boilerplate
_GENERIC_DESC = [
    "This represents a systematic security issue affecting multiple systems requiring coordinated remediation",
    "The widespread nature indicates common misconfiguration",
    "consistent security control gaps",
    "Exploitation of any affected system could lead to unauthorized access",
    "implement appropriate remediation",
    "follow security best practices",
]
_GENERIC_REC = [
    "Implement systematic remediation across all",
    "Deploy consistent security controls and monitoring across all affected assets",
    "Deploy defense-in-depth measures and monitoring controls consistently",
    "Prioritize remediation based on asset criticality and verify each of the",
    "Implement automated controls to detect and prevent similar vulnerabilities",
    "implement appropriate remediation",
    "follow security best practices",
]
_MIN_DESC_LEN = 150
_MIN_REC_LEN  = 100

def _is_generic(text, patterns, min_len):
    if not text or len(text.strip()) < min_len:
        return True
    return any(p in text for p in patterns)

def _extract_lead(description):
    """Pull the first sentence from the description to use as the lead."""
    if not description:
        return ''
    idx = description.find('. ')
    if idx > 0:
        return description[:idx + 1].strip()
    return description.strip()

_quality_issues  = []   # (idx, tab, item, issue_type, fixed)
_auto_fixed      = 0
_needs_review    = 0

for _i, _f in enumerate(all_findings):
    _tab  = _f.get('Tab', '')
    _item = _f.get('Item', '')
    _desc = _f.get('Description', '')
    _rec  = _f.get('Recommendation', '')

    _bad_desc = _is_generic(_desc, _GENERIC_DESC, _MIN_DESC_LEN)
    _bad_rec  = _is_generic(_rec,  _GENERIC_REC,  _MIN_REC_LEN)

    if not (_bad_desc or _bad_rec):
        continue  # Both look fine

    _issue = []
    if _bad_desc: _issue.append('desc')
    if _bad_rec:  _issue.append('rec')

    _fixed = False
    if _HAS_ENRICH:
        _lead = _extract_lead(_desc)
        _new_desc, _new_rec = _enrich_finding(_tab, _lead)
        if _new_rec:  # Library returned something specific
            if _bad_desc:
                all_findings[_i]['Description'] = _new_desc
            if _bad_rec:
                all_findings[_i]['Recommendation'] = _new_rec
            _fixed = True
            _auto_fixed += 1

    if not _fixed:
        _needs_review += 1

    _quality_issues.append((_i, _tab, _item[:60], '+'.join(_issue), _fixed))

# Print report
if not _quality_issues:
    print("  All descriptions and recommendations passed quality check.")
else:
    print(f"  Found {len(_quality_issues)} quality issues:")
    print(f"    Auto-fixed:    {_auto_fixed}")
    print(f"    Needs review:  {_needs_review}")
    print()
    _W = 32
    print(f"  {'Tab':<30} {'Issue':<10} {'Fixed':<6}  Item")
    print(f"  {'-'*30} {'-'*10} {'-'*6}  {'-'*50}")
    for _, _tab, _item, _issue, _fixed in _quality_issues:
        _status = 'YES' if _fixed else 'NO *'
        print(f"  {_tab:<30} {_issue:<10} {_status:<6}  {_item}")

# Save quality report
_qr_path = os.path.join(FINAL_DIR, 'QUALITY_REPORT.md')
with open(_qr_path, 'w') as _qf:
    _qf.write("# Quality Failsafe Report\n\n")
    if not _quality_issues:
        _qf.write("All descriptions and recommendations passed quality check. No issues found.\n")
    else:
        _qf.write(f"**Total issues found:** {len(_quality_issues)}  \n")
        _qf.write(f"**Auto-fixed:** {_auto_fixed}  \n")
        _qf.write(f"**Needs manual review:** {_needs_review}\n\n")
        if _auto_fixed:
            _qf.write("## Auto-Fixed\n\n")
            _qf.write("| Tab | Issue | Item |\n")
            _qf.write("|-----|-------|------|\n")
            for _, _tab, _item, _issue, _fixed in _quality_issues:
                if _fixed:
                    _qf.write(f"| `{_tab}` | {_issue} | {_item} |\n")
        if _needs_review:
            _qf.write("\n## Needs Manual Review\n\n")
            _qf.write("These findings have generic/weak content and are NOT in the enrichment library.\n")
            _qf.write("Open `finding_enrichment.py` and add a `FINDING_SPECS` entry for each.\n\n")
            _qf.write("| Tab | Issue | Item |\n")
            _qf.write("|-----|-------|------|\n")
            for _, _tab, _item, _issue, _fixed in _quality_issues:
                if not _fixed:
                    _qf.write(f"| `{_tab}` | {_issue} | {_item} |\n")
print(f"  Quality report saved: QUALITY_REPORT.md")

# ============================================================================
# SAVE CONSOLIDATED CSV
# ============================================================================
print("\n=== Saving Consolidated CSV ===")

csv_path = os.path.join(FINAL_DIR, 'CONSOLIDATED_FINDINGS.csv')
with open(csv_path, 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
    writer.writeheader()
    # Sort by severity then category
    severity_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Info': 4}
    sorted_findings = sorted(all_findings, key=lambda x: (severity_order.get(x.get('Priority', 'Medium'), 2), x.get('Category', '')))
    writer.writerows(sorted_findings)
print(f"  Saved: {csv_path} ({len(all_findings)} findings)")

# ============================================================================
# SAVE EXECUTIVE SUMMARY
# ============================================================================
print("\n=== Saving Executive Summary ===")

summary_path = os.path.join(FINAL_DIR, 'EXECUTIVE_SUMMARY.md')
with open(summary_path, 'w') as f:
    f.write("# ASM-NG Security Analysis - Executive Summary\n\n")
    f.write(f"**Analysis Date:** 2026-02-20\n")
    f.write(f"**Overall Score:** {overall_score}/100\n")
    f.write(f"**Overall Grade:** {overall_grade}\n\n")

    f.write("---\n\n")
    f.write("## Grade Summary\n\n")
    f.write("| Category | Weight | Score | Grade | Findings |\n")
    f.write("|----------|--------|-------|-------|----------|\n")
    for cat_name in GRADE_CATEGORIES:
        cs = category_scores[cat_name]
        f.write(f"| {cat_name} | {cs['weight']} | {cs['score']} | {cs['grade']} | {cs['findings']} |\n")
    f.write(f"| **Overall** | **weighted** | **{overall_score}** | **{overall_grade}** | **{len(all_findings)}** |\n\n")

    f.write("*Grade thresholds: A=90+, B=80+, C=70+, D=60+, F=<60*\n")
    f.write("*Score = max(0, 100 + (raw_score * weight))*\n\n")

    f.write("---\n\n")
    f.write("## Severity Breakdown\n\n")
    f.write("| Severity | Count |\n")
    f.write("|----------|-------|\n")
    for sev in ['Critical', 'High', 'Medium', 'Low']:
        f.write(f"| {sev} | {severity_counts.get(sev, 0)} |\n")
    f.write(f"| **Total** | **{len(all_findings)}** |\n\n")

    f.write("---\n\n")
    f.write("## Key Findings by Category\n\n")
    for g in range(1, 8):
        cat_names = ['', 'Network Security', 'Web App Security', 'Information Leakage',
                     'General Health', 'External Account Exposure', 'DNS Health', 'IP Reputation']
        cat_name = cat_names[g]
        cs = category_scores.get(cat_name, {})
        gs = group_stats.get(g, {})
        f.write(f"### GROUP {g}: {cat_name} (Grade: {cs.get('grade', '--')})\n\n")
        f.write(f"- **Score:** {cs.get('score', '--')}/100\n")
        f.write(f"- **Findings:** {gs.get('total', 0)} (Critical: {gs.get('critical', 0)}, High: {gs.get('high', 0)}, Medium: {gs.get('medium', 0)}, Low: {gs.get('low', 0)})\n\n")

    f.write("---\n\n")
    f.write("## Recommendations Priority\n\n")
    f.write("1. **Immediate (Critical):** Address all Critical findings within 24-48 hours\n")
    f.write("2. **Urgent (High):** Remediate High findings within 7 days\n")
    f.write("3. **Standard (Medium):** Address Medium findings within 30 days\n")
    f.write("4. **Routine (Low):** Include Low findings in next maintenance cycle\n\n")
    f.write("---\n\n")
    f.write("*Generated by ASM-NG Analysis Pipeline on 2026-02-20*\n")

print(f"  Saved: {summary_path}")

# ============================================================================
# SAVE EXCEL WORKBOOK (if openpyxl available)
# ============================================================================
if HAS_OPENPYXL:
    print("\n=== Saving Excel Workbook ===")

    wb = openpyxl.Workbook()

    # --- Findings Sheet ---
    ws = wb.active
    ws.title = "Findings"

    # Header styling - black background, white bold uppercase text
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    severity_fills = {
        'Critical': PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),  # Purple
        'High': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),       # Red
        'Medium': PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),     # Yellow/Orange
        'Low': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),        # Blue
    }
    severity_fonts = {
        'Critical': Font(color="FFFFFF", bold=True),
        'High': Font(color="FFFFFF", bold=True),
        'Medium': Font(color="FFFFFF", bold=True),
        'Low': Font(color="FFFFFF", bold=True),
    }

    # Headers - uppercase
    headers = fieldnames
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    sorted_findings = sorted(all_findings, key=lambda x: (severity_order.get(x.get('Priority', 'Medium'), 2), x.get('Category', '')))
    for row_idx, finding in enumerate(sorted_findings, 2):
        for col_idx, field in enumerate(headers, 1):
            val = finding.get(field, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Color the Priority column
            if field == 'Priority' and val in severity_fills:
                cell.fill = severity_fills[val]
                cell.font = severity_fonts[val]

    # Column widths
    col_widths = [20, 30, 10, 50, 80, 80, 12, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Auto-filter
    ws.auto_filter.ref = f"A1:H{len(sorted_findings)+1}"

    # --- Executive Summary Sheet ---
    ws2 = wb.create_sheet("Executive Summary")

    ws2.cell(row=1, column=1, value="ASM-NG Security Analysis").font = Font(bold=True, size=16)
    ws2.cell(row=2, column=1, value=f"Analysis Date: 2026-02-20")
    ws2.cell(row=3, column=1, value=f"Overall Score: {overall_score}/100")
    ws2.cell(row=3, column=1).font = Font(bold=True, size=14)
    ws2.cell(row=4, column=1, value=f"Overall Grade: {overall_grade}")
    ws2.cell(row=4, column=1).font = Font(bold=True, size=14)

    # Grade summary table - uppercase headers
    grade_headers = ['Category', 'Weight', 'Score', 'Grade', 'Findings']
    for col, h in enumerate(grade_headers, 1):
        cell = ws2.cell(row=6, column=col, value=h.upper())
        cell.fill = header_fill
        cell.font = header_font

    row = 7
    for cat_name in GRADE_CATEGORIES:
        cs = category_scores[cat_name]
        ws2.cell(row=row, column=1, value=cat_name)
        ws2.cell(row=row, column=2, value=cs['weight'])
        ws2.cell(row=row, column=3, value=cs['score'])
        grade_cell = ws2.cell(row=row, column=4, value=cs['grade'])
        ws2.cell(row=row, column=5, value=cs['findings'])

        # Color grade
        grade_colors = {'A': '00B050', 'B': '92D050', 'C': 'FFD700', 'D': 'FF6600', 'F': 'FF0000'}
        if cs['grade'] in grade_colors:
            grade_cell.fill = PatternFill(start_color=grade_colors[cs['grade']], end_color=grade_colors[cs['grade']], fill_type="solid")
            if cs['grade'] in ('D', 'F'):
                grade_cell.font = Font(color="FFFFFF", bold=True)
            else:
                grade_cell.font = Font(bold=True)
        row += 1

    # Overall row
    ws2.cell(row=row, column=1, value="OVERALL").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="weighted").font = Font(bold=True)
    ws2.cell(row=row, column=3, value=overall_score).font = Font(bold=True)
    overall_cell = ws2.cell(row=row, column=4, value=overall_grade)
    overall_cell.font = Font(bold=True, size=14)
    if overall_grade in grade_colors:
        overall_cell.fill = PatternFill(start_color=grade_colors[overall_grade], end_color=grade_colors[overall_grade], fill_type="solid")
    ws2.cell(row=row, column=5, value=len(all_findings)).font = Font(bold=True)

    # Severity breakdown
    row += 2
    ws2.cell(row=row, column=1, value="Severity Breakdown").font = Font(bold=True, size=12)
    row += 1
    for sev in ['Critical', 'High', 'Medium', 'Low']:
        ws2.cell(row=row, column=1, value=sev)
        ws2.cell(row=row, column=2, value=severity_counts.get(sev, 0))
        if sev in severity_fills:
            ws2.cell(row=row, column=1).fill = severity_fills[sev]
            ws2.cell(row=row, column=1).font = severity_fonts[sev]
        row += 1

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12

    # --- Grade Summary Sheet ---
    ws3 = wb.create_sheet("Grade Summary")

    ws3.cell(row=1, column=1, value="ASM-NG Grade Summary").font = Font(bold=True, size=14)
    ws3.cell(row=2, column=1, value=f"Score Formula: max(0, 100 + (raw_score * weight))")
    ws3.cell(row=3, column=1, value="Grade Thresholds: A=90+, B=80+, C=70+, D=60+, F=<60")

    grade_detail_headers = ['Category', 'Weight', 'Raw Score', 'Weighted Score', 'Grade', 'Findings', 'Critical', 'High', 'Medium', 'Low']
    for col, h in enumerate(grade_detail_headers, 1):
        cell = ws3.cell(row=5, column=col, value=h.upper())
        cell.fill = header_fill
        cell.font = header_font

    row = 6
    cat_to_group = {
        'Network Security': 1, 'Web App Security': 2, 'Information Leakage': 3,
        'General Health': 4, 'External Account Exposure': 5, 'DNS Health': 6, 'IP Reputation': 7
    }
    for cat_name in GRADE_CATEGORIES:
        cs = category_scores[cat_name]
        g = cat_to_group[cat_name]
        gs = group_stats.get(g, {})
        ws3.cell(row=row, column=1, value=cat_name)
        ws3.cell(row=row, column=2, value=cs['weight'])
        ws3.cell(row=row, column=3, value=cs['raw_score'])
        ws3.cell(row=row, column=4, value=cs['score'])
        grade_cell = ws3.cell(row=row, column=5, value=cs['grade'])
        if cs['grade'] in grade_colors:
            grade_cell.fill = PatternFill(start_color=grade_colors[cs['grade']], end_color=grade_colors[cs['grade']], fill_type="solid")
        ws3.cell(row=row, column=6, value=cs['findings'])
        ws3.cell(row=row, column=7, value=gs.get('critical', 0))
        ws3.cell(row=row, column=8, value=gs.get('high', 0))
        ws3.cell(row=row, column=9, value=gs.get('medium', 0))
        ws3.cell(row=row, column=10, value=gs.get('low', 0))
        row += 1

    for col in range(1, 11):
        ws3.column_dimensions[chr(64 + col)].width = 18

    # Save
    xlsx_path = os.path.join(FINAL_DIR, 'CONSOLIDATED_FINDINGS.xlsx')
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")
else:
    print("\n  Skipping Excel output (openpyxl not installed)")

# ============================================================================
# FINAL VERIFICATION
# ============================================================================
print("\n=== Final Verification ===")

output_files = [
    'CONSOLIDATED_FINDINGS.csv',
    'CONSOLIDATED_FINDINGS.xlsx' if HAS_OPENPYXL else None,
    'EXECUTIVE_SUMMARY.md',
    'analysis_progress_tracker.md',
]

for fname in output_files:
    if fname is None:
        continue
    fpath = os.path.join(FINAL_DIR, fname)
    if os.path.exists(fpath):
        size = os.path.getsize(fpath)
        print(f"  OK: {fname} ({size:,} bytes)")
    else:
        print(f"  MISSING: {fname}")

for g in range(1, 8):
    fpath = os.path.join(OUTPUT_DIR, f'findings_GROUP{g}_FULL_CONSOLIDATED.csv')
    if os.path.exists(fpath):
        size = os.path.getsize(fpath)
        print(f"  OK: findings_GROUP{g}_FULL_CONSOLIDATED.csv ({size:,} bytes)")

detail_count = len(os.listdir(os.path.join(OUTPUT_DIR, 'findings_details')))
print(f"  OK: {detail_count} detail files in findings_details/")

print("\n" + "=" * 70)
print("FINAL CONSOLIDATION COMPLETE")
print(f"Overall Grade: {overall_grade} ({overall_score}/100)")
print("=" * 70)
