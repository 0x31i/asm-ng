#!/usr/bin/env python3
"""
ASM Data Obfuscation Pipeline - Advanced Tokenization
Replaces sensitive data (IPs, domains, emails, URLs) with reversible tokens
before processing through Claude Code.

SECURITY PRINCIPLE: Claude never sees actual sensitive identifiers.
Even if data is retained by Anthropic, tokens are meaningless without mapping file.

SUPPORTED FILE TYPES:
- Excel (.xlsx, .xlsm)
- CSV (.csv)
- HTML (.html, .htm) - Burp Suite reports
- Nessus (.nessus) - Nessus XML scan files
- Text (.txt, .log, .xml, .json)
"""

import re
import json
import hashlib
import ipaddress
from pathlib import Path
from datetime import datetime
from typing import Dict, Set, Tuple
import csv
import xml.etree.ElementTree as ET

# ANSI Color Codes
class Colors:
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    BOLD = '\033[1m'
    RESET = '\033[0m'

# Excel support
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print(f"{Colors.YELLOW}WARNING: openpyxl not installed. Excel obfuscation disabled.{Colors.RESET}")
    print("Install with: pip install openpyxl")

# HTML parsing
try:
    from bs4 import BeautifulSoup
    HTML_AVAILABLE = True
except ImportError:
    HTML_AVAILABLE = False
    print(f"{Colors.YELLOW}WARNING: BeautifulSoup4 not installed. HTML obfuscation disabled.{Colors.RESET}")
    print("Install with: pip3 install beautifulsoup4")


class TokenMapper:
    """Manages reversible token mappings for obfuscated data."""
    
    def __init__(self, mapping_file: str = "obfuscation_mapping.json.enc"):
        self.mapping_file = Path(mapping_file)
        self.mappings: Dict[str, Dict[str, str]] = {
            'ip_addresses': {},      # real_ip -> token
            'domains': {},           # real_domain -> token
            'urls': {},              # real_url -> token
            'emails': {},            # real_email -> token
            'hostnames': {},         # real_hostname -> token
            'paths': {},             # real_path -> token
        }
        self.reverse_mappings: Dict[str, Dict[str, str]] = {}
        self.counters = {
            'ip_addresses': 1,
            'domains': 1,
            'urls': 1,
            'emails': 1,
            'hostnames': 1,
            'paths': 1,
        }
        
    def _generate_token(self, category: str, value: str) -> str:
        """Generate deterministic token for a value."""
        # Check if we've already tokenized this value
        if value in self.mappings[category]:
            return self.mappings[category][value]
        
        # Map category names to token prefixes
        category_to_prefix = {
            'ip_addresses': 'IP_ADDRESS',
            'domains': 'DOMAIN',
            'urls': 'URL',
            'emails': 'EMAIL',
            'hostnames': 'HOSTNAME',
            'paths': 'PATH',
        }
        
        # Generate new token
        counter = self.counters[category]
        prefix = category_to_prefix.get(category, category.upper())
        token = f"{prefix}_TOKEN_{counter:04d}"
        
        # Store mapping
        self.mappings[category][value] = token
        self.counters[category] += 1
        
        return token
    
    def tokenize_ip(self, ip: str) -> str:
        """Replace IP address with token."""
        return self._generate_token('ip_addresses', ip)
    
    def tokenize_domain(self, domain: str) -> str:
        """Replace domain with token."""
        return self._generate_token('domains', domain)
    
    def tokenize_url(self, url: str) -> str:
        """Replace URL with token."""
        return self._generate_token('urls', url)
    
    def tokenize_email(self, email: str) -> str:
        """Replace email with token."""
        return self._generate_token('emails', email)
    
    def tokenize_hostname(self, hostname: str) -> str:
        """Replace hostname with token."""
        return self._generate_token('hostnames', hostname)
    
    def tokenize_path(self, path: str) -> str:
        """Replace file path with token."""
        return self._generate_token('paths', path)
    
    def save_mappings(self, output_path: Path):
        """Save mappings to encrypted JSON file."""
        mapping_data = {
            'timestamp': datetime.utcnow().isoformat(),
            'mappings': self.mappings,
            'counters': self.counters,
            'stats': {
                category: len(mappings) 
                for category, mappings in self.mappings.items()
            }
        }
        
        # Save as JSON (in production, encrypt this file!)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(mapping_data, f, indent=2)
        
        # Create reverse mappings for de-obfuscation
        self.reverse_mappings = {
            category: {token: value for value, token in mappings.items()}
            for category, mappings in self.mappings.items()
        }
        
        reverse_path = output_path.parent / f"reverse_{output_path.name}"
        with open(reverse_path, 'w', encoding='utf-8') as f:
            json.dump({
                'timestamp': datetime.utcnow().isoformat(),
                'reverse_mappings': self.reverse_mappings
            }, f, indent=2)
        
        print(f"\n{Colors.GREEN}[SUCCESS] Mappings saved to: {output_path}{Colors.RESET}")
        print(f"{Colors.GREEN}[SUCCESS] Reverse mappings saved to: {reverse_path}{Colors.RESET}")
        print(f"\n{Colors.YELLOW}[CRITICAL] Keep these files secure and LOCAL. Never upload to Claude!{Colors.RESET}")
    
    def load_mappings(self, mapping_path: Path):
        """Load existing mappings from file."""
        with open(mapping_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            self.mappings = data['mappings']
            self.counters = data['counters']
        print(f"{Colors.GREEN}[SUCCESS] Loaded {sum(len(m) for m in self.mappings.values())} existing mappings{Colors.RESET}")


class ASMObfuscator:
    """Obfuscates sensitive data in ASM scan results."""
    
    # Regex patterns for sensitive data
    IP_PATTERN = re.compile(
        r'\b(?:'
        r'(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}'
        r'(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)'
        r'\b'
    )
    
    DOMAIN_PATTERN = re.compile(
        r'\b(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,}\b',
        re.IGNORECASE
    )
    
    EMAIL_PATTERN = re.compile(
        r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    )
    
    URL_PATTERN = re.compile(
        r'https?://[^\s<>"{}|\\^`\[\]]+',
        re.IGNORECASE
    )
    
    HOSTNAME_PATTERN = re.compile(
        r'\b(?:server|host|node|vm|pc|workstation|dev|prod|staging|test)[-.a-z0-9]+\b',
        re.IGNORECASE
    )
    
    PATH_PATTERN = re.compile(
        r'(?:[A-Z]:\\|/)[^\s<>"|?*]+',
        re.IGNORECASE
    )
    
    # Common false positives to ignore
    IGNORE_DOMAINS = {
        'localhost', 'example.com', 'example.org', 'test.com',
        'w3.org', 'ietf.org', 'rfc-editor.org',
        'microsoft.com', 'apple.com', 'google.com',  # Common tool references
        'owasp.org', 'cwe.mitre.org', 'nvd.nist.gov',  # Vulnerability references
        'nessus.org', 'tenable.com', 'tenable.io',  # Nessus references
        'nist.gov', 'mitre.org', 'first.org'  # Security standards references
    }
    
    IGNORE_IPS = {
        '0.0.0.0', '127.0.0.1', '255.255.255.255',
        '192.0.2.1', '198.51.100.1', '203.0.113.1'  # RFC 5737 documentation IPs
    }
    
    def __init__(self, mapper: TokenMapper):
        self.mapper = mapper
        self.stats = {
            'ips_obfuscated': 0,
            'domains_obfuscated': 0,
            'urls_obfuscated': 0,
            'emails_obfuscated': 0,
            'hostnames_obfuscated': 0,
            'paths_obfuscated': 0,
        }
    
    def _is_private_ip(self, ip: str) -> bool:
        """Check if IP is private/reserved."""
        try:
            ip_obj = ipaddress.ip_address(ip)
            return ip_obj.is_private or ip_obj.is_loopback or ip_obj.is_reserved
        except ValueError:
            return False
    
    def obfuscate_text(self, text: str) -> str:
        """Obfuscate all sensitive data in text."""
        if not text or not isinstance(text, str):
            return text
        
        original_text = text
        
        # 1. Obfuscate URLs first (to avoid breaking domain extraction)
        def replace_url(match):
            url = match.group(0)
            token = self.mapper.tokenize_url(url)
            self.stats['urls_obfuscated'] += 1
            return token
        
        text = self.URL_PATTERN.sub(replace_url, text)
        
        # 2. Obfuscate emails
        def replace_email(match):
            email = match.group(0)
            token = self.mapper.tokenize_email(email)
            self.stats['emails_obfuscated'] += 1
            return token
        
        text = self.EMAIL_PATTERN.sub(replace_email, text)
        
        # 3. Obfuscate IP addresses
        def replace_ip(match):
            ip = match.group(0)
            # Skip ignored IPs
            if ip in self.IGNORE_IPS:
                return ip
            token = self.mapper.tokenize_ip(ip)
            self.stats['ips_obfuscated'] += 1
            return token
        
        text = self.IP_PATTERN.sub(replace_ip, text)
        
        # 4. Obfuscate domains (after URLs to avoid double-processing)
        def replace_domain(match):
            domain = match.group(0).lower()
            # Skip ignored domains
            if domain in self.IGNORE_DOMAINS:
                return match.group(0)
            # Skip if it's part of a token we already created
            if 'TOKEN' in match.group(0).upper():
                return match.group(0)
            token = self.mapper.tokenize_domain(domain)
            self.stats['domains_obfuscated'] += 1
            return token
        
        text = self.DOMAIN_PATTERN.sub(replace_domain, text)
        
        # 5. Obfuscate hostnames
        def replace_hostname(match):
            hostname = match.group(0)
            # Skip if already a token
            if 'TOKEN' in hostname.upper():
                return hostname
            token = self.mapper.tokenize_hostname(hostname)
            self.stats['hostnames_obfuscated'] += 1
            return token
        
        text = self.HOSTNAME_PATTERN.sub(replace_hostname, text)
        
        # 6. Obfuscate file paths
        def replace_path(match):
            path = match.group(0)
            # Skip if already a token
            if 'TOKEN' in path.upper():
                return path
            # Skip very short paths (likely false positives)
            if len(path) < 10:
                return path
            token = self.mapper.tokenize_path(path)
            self.stats['paths_obfuscated'] += 1
            return token
        
        text = self.PATH_PATTERN.sub(replace_path, text)
        
        return text
    
    def obfuscate_csv(self, input_path: Path, output_path: Path):
        """Obfuscate sensitive data in CSV files."""
        import time
        start_time = time.time()
        
        print(f"\n{Colors.CYAN}[INFO] Processing CSV: {input_path.name}{Colors.RESET}")
        print(f"    [1/3] Reading CSV file...", end='', flush=True)
        
        with open(input_path, 'rb') as raw:
            clean = raw.read().replace(b'\x00', b'')
        import io
        reader = csv.reader(io.StringIO(clean.decode('utf-8', errors='ignore')))
        rows = list(reader)
        
        total_rows = len(rows)
        file_size_kb = input_path.stat().st_size / 1024
        print(f" Done! ({total_rows:,} rows, {file_size_kb:.1f} KB)")
        
        print(f"    [2/3] Obfuscating data...")
        obfuscated_rows = []
        last_percent = -1
        
        for row_num, row in enumerate(rows):
            obfuscated_row = [self.obfuscate_text(cell) for cell in row]
            obfuscated_rows.append(obfuscated_row)
            
            # Update progress every 5%
            if total_rows > 100:
                percent = int((row_num / total_rows) * 100)
                if percent != last_percent and percent % 5 == 0:
                    elapsed = time.time() - start_time
                    rows_per_sec = (row_num + 1) / elapsed if elapsed > 0 else 0
                    eta_sec = (total_rows - row_num) / rows_per_sec if rows_per_sec > 0 else 0
                    print(f"    Progress: {percent}% ({row_num:,}/{total_rows:,} rows) [{rows_per_sec:.0f} rows/sec, ETA: {eta_sec:.0f}s]", 
                          end='\r', flush=True)
                    last_percent = percent
        
        if total_rows > 100:
            print()  # New line after progress updates
        
        print(f"    [3/3] Writing obfuscated CSV...", end='', flush=True)
        with open(output_path, 'w', encoding='utf-8', newline='') as outfile:
            writer = csv.writer(outfile)
            writer.writerows(obfuscated_rows)
        
        elapsed = time.time() - start_time
        print(f" Done!")
        print(f"{Colors.GREEN}[SUCCESS] CSV obfuscated: {len(rows):,} rows processed in {elapsed:.1f}s{Colors.RESET}")
    
    def obfuscate_excel(self, input_path: Path, output_path: Path):
        """Obfuscate sensitive data in Excel files."""
        if not EXCEL_AVAILABLE:
            print(f"{Colors.YELLOW}[SKIPPED] Excel file (openpyxl not installed): {input_path.name}{Colors.RESET}")
            return
        
        print(f"\n{Colors.CYAN}[INFO] Processing Excel: {input_path.name}{Colors.RESET}")
        
        # Load workbook
        print(f"    Loading workbook...", end='', flush=True)
        workbook = openpyxl.load_workbook(input_path)
        print(f" Done! Found {len(workbook.sheetnames)} sheet(s)")
        
        total_cells_processed = 0
        
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
            sheet = workbook[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            total_cells = max_row * max_col
            
            print(f"\n    [Sheet {sheet_idx}/{len(workbook.sheetnames)}] {sheet_name}")
            print(f"    Size: {max_row} rows x {max_col} columns ({total_cells:,} cells)")
            
            cells_processed = 0
            last_percent = -1
            
            for row_num, row in enumerate(sheet.iter_rows(), 1):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell.value = self.obfuscate_text(cell.value)
                        cells_processed += 1
                
                # Update progress every 5%
                percent = int((row_num / max_row) * 100)
                if percent != last_percent and percent % 5 == 0:
                    print(f"    Processing: {percent}% ({row_num:,}/{max_row:,} rows)", end='\r', flush=True)
                    last_percent = percent
            
            print(f"    Completed: 100% ({max_row:,}/{max_row:,} rows) - {cells_processed:,} cells obfuscated")
            total_cells_processed += cells_processed
        
        print(f"\n    Saving workbook...", end='', flush=True)
        workbook.save(output_path)
        print(f" Done!")
        
        print(f"{Colors.GREEN}[SUCCESS] Excel obfuscated: {len(workbook.sheetnames)} sheets, {total_cells_processed:,} total cells processed{Colors.RESET}")
    
    def obfuscate_html(self, input_path: Path, output_path: Path):
        """Obfuscate sensitive data in HTML files (Burp reports)."""
        if not HTML_AVAILABLE:
            print(f"{Colors.YELLOW}[SKIPPED] HTML file (beautifulsoup4 not installed): {input_path.name}{Colors.RESET}")
            return
        
        print(f"\n{Colors.CYAN}[INFO] Processing HTML: {input_path.name}{Colors.RESET}")
        
        # Read file
        print(f"    [1/5] Reading file...", end='', flush=True)
        with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
            html_content = f.read()
        file_size_kb = len(html_content) / 1024
        print(f" Done! ({file_size_kb:.1f} KB)")
        
        # Parse HTML
        print(f"    [2/5] Parsing HTML structure...", end='', flush=True)
        soup = BeautifulSoup(html_content, 'html.parser')
        print(f" Done!")
        
        # Count elements
        total_text_nodes = len(soup.find_all(string=True))
        total_tags = len(soup.find_all())
        print(f"    Found {total_text_nodes:,} text nodes and {total_tags:,} HTML tags")
        
        # Obfuscate text nodes
        print(f"    [3/5] Obfuscating text content...")
        text_nodes_processed = 0
        for text_node in soup.find_all(string=True):
            if text_node.parent.name not in ['script', 'style']:
                obfuscated = self.obfuscate_text(str(text_node))
                text_node.replace_with(obfuscated)
                text_nodes_processed += 1
                
                # Update every 10%
                if text_nodes_processed % max(1, total_text_nodes // 10) == 0:
                    percent = int((text_nodes_processed / total_text_nodes) * 100)
                    print(f"    Progress: {percent}% ({text_nodes_processed:,}/{total_text_nodes:,} nodes)", end='\r', flush=True)
        
        print(f"    Completed: 100% ({text_nodes_processed:,}/{total_text_nodes:,} nodes)")
        
        # Obfuscate attributes
        print(f"    [4/5] Obfuscating HTML attributes...", end='', flush=True)
        sensitive_attrs = ['href', 'src', 'data-url', 'data-host', 'data-ip']
        attrs_processed = 0
        for tag in soup.find_all():
            for attr in sensitive_attrs:
                if attr in tag.attrs:
                    tag[attr] = self.obfuscate_text(tag[attr])
                    attrs_processed += 1
        print(f" Done! ({attrs_processed} attributes)")
        
        # Save file
        print(f"    [5/5] Writing obfuscated HTML...", end='', flush=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(str(soup))
        output_size_kb = output_path.stat().st_size / 1024
        print(f" Done! ({output_size_kb:.1f} KB)")
        
        print(f"{Colors.GREEN}[SUCCESS] HTML obfuscated: {text_nodes_processed:,} text nodes + {attrs_processed} attributes processed{Colors.RESET}")
    
    def obfuscate_nessus(self, input_path: Path, output_path: Path):
        """
        Obfuscate sensitive data in Nessus XML files (.nessus).
        
        Nessus files contain:
        - ReportHost[@name] - target IP/hostname
        - HostProperties/tag[@name='host-ip'] - IP address
        - HostProperties/tag[@name='hostname'] - hostname
        - HostProperties/tag[@name='host-fqdn'] - FQDN
        - HostProperties/tag[@name='netbios-name'] - NetBIOS name
        - ReportItem plugin_output - may contain IPs, domains, paths
        - Various description/solution text with URLs and references
        """
        import time
        start_time = time.time()
        
        print(f"\n{Colors.CYAN}[INFO] Processing Nessus XML: {input_path.name}{Colors.RESET}")
        
        # Read file
        print(f"    [1/6] Reading file...", end='', flush=True)
        with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
            xml_content = f.read()
        file_size_kb = len(xml_content) / 1024
        print(f" Done! ({file_size_kb:.1f} KB)")
        
        # Parse XML
        print(f"    [2/6] Parsing XML structure...", end='', flush=True)
        try:
            tree = ET.parse(input_path)
            root = tree.getroot()
            print(f" Done!")
        except ET.ParseError as e:
            print(f" Failed!")
            print(f"{Colors.RED}[ERROR] Failed to parse Nessus XML: {e}{Colors.RESET}")
            return
        
        # Count elements
        report_hosts = root.findall('.//ReportHost')
        report_items = root.findall('.//ReportItem')
        print(f"    Found {len(report_hosts)} hosts and {len(report_items):,} findings")
        
        # Track obfuscation stats for this file
        hosts_processed = 0
        items_processed = 0
        properties_processed = 0
        
        # ===================================================================
        # STEP 3: Obfuscate ReportHost names (target IPs/hostnames)
        # ===================================================================
        print(f"    [3/6] Obfuscating host identifiers...")
        
        for host in report_hosts:
            # Obfuscate the 'name' attribute (usually IP or hostname)
            host_name = host.get('name', '')
            if host_name:
                obfuscated_name = self.obfuscate_text(host_name)
                host.set('name', obfuscated_name)
                hosts_processed += 1
            
            # Obfuscate HostProperties tags
            host_properties = host.find('HostProperties')
            if host_properties is not None:
                sensitive_properties = [
                    'host-ip', 'hostname', 'host-fqdn', 'netbios-name',
                    'host-rdns', 'operating-system', 'mac-address',
                    'bios-uuid', 'system-type'
                ]
                
                for tag in host_properties.findall('tag'):
                    tag_name = tag.get('name', '')
                    
                    # Always obfuscate known sensitive properties
                    if tag_name in sensitive_properties:
                        if tag.text:
                            tag.text = self.obfuscate_text(tag.text)
                            properties_processed += 1
                    
                    # Also check tag text for IPs/domains even if property name isn't recognized
                    elif tag.text:
                        original = tag.text
                        obfuscated = self.obfuscate_text(tag.text)
                        if original != obfuscated:
                            tag.text = obfuscated
                            properties_processed += 1
            
            # Progress update
            if hosts_processed % max(1, len(report_hosts) // 10) == 0:
                percent = int((hosts_processed / len(report_hosts)) * 100)
                print(f"    Progress: {percent}% ({hosts_processed}/{len(report_hosts)} hosts)", end='\r', flush=True)
        
        print(f"    Completed: {hosts_processed} hosts, {properties_processed} properties obfuscated")
        
        # ===================================================================
        # STEP 4: Obfuscate ReportItem content
        # ===================================================================
        print(f"    [4/6] Obfuscating finding details...")
        
        # Elements that may contain sensitive data
        sensitive_elements = [
            'plugin_output', 'description', 'solution', 'synopsis',
            'plugin_name', 'fname', 'script_version', 'see_also'
        ]
        
        for idx, item in enumerate(report_items):
            # Obfuscate attributes
            for attr in ['svc_name', 'protocol']:
                if attr in item.attrib:
                    original = item.get(attr)
                    obfuscated = self.obfuscate_text(original)
                    if original != obfuscated:
                        item.set(attr, obfuscated)
            
            # Obfuscate child element text
            for elem_name in sensitive_elements:
                elem = item.find(elem_name)
                if elem is not None and elem.text:
                    elem.text = self.obfuscate_text(elem.text)
            
            items_processed += 1
            
            # Progress update every 10%
            if len(report_items) > 100 and idx % max(1, len(report_items) // 10) == 0:
                percent = int((idx / len(report_items)) * 100)
                elapsed = time.time() - start_time
                items_per_sec = (idx + 1) / elapsed if elapsed > 0 else 0
                eta_sec = (len(report_items) - idx) / items_per_sec if items_per_sec > 0 else 0
                print(f"    Progress: {percent}% ({idx:,}/{len(report_items):,} items) [{items_per_sec:.0f}/sec, ETA: {eta_sec:.0f}s]", end='\r', flush=True)
        
        print(f"    Completed: {items_processed:,} findings processed")
        
        # ===================================================================
        # STEP 5: Obfuscate any remaining text elements
        # ===================================================================
        print(f"    [5/6] Scanning for additional sensitive data...", end='', flush=True)
        
        additional_obfuscated = 0
        
        # Walk all elements and check text content
        for elem in root.iter():
            # Check element text
            if elem.text:
                original = elem.text
                obfuscated = self.obfuscate_text(original)
                if original != obfuscated:
                    elem.text = obfuscated
                    additional_obfuscated += 1
            
            # Check element tail (text after closing tag)
            if elem.tail:
                original = elem.tail
                obfuscated = self.obfuscate_text(original)
                if original != obfuscated:
                    elem.tail = obfuscated
                    additional_obfuscated += 1
        
        print(f" Done! ({additional_obfuscated} additional items)")
        
        # ===================================================================
        # STEP 6: Write obfuscated XML
        # ===================================================================
        print(f"    [6/6] Writing obfuscated Nessus file...", end='', flush=True)
        
        # Write with XML declaration
        tree.write(output_path, encoding='utf-8', xml_declaration=True)
        
        output_size_kb = output_path.stat().st_size / 1024
        elapsed = time.time() - start_time
        print(f" Done! ({output_size_kb:.1f} KB)")
        
        print(f"{Colors.GREEN}[SUCCESS] Nessus file obfuscated in {elapsed:.1f}s:{Colors.RESET}")
        print(f"    • Hosts: {hosts_processed}")
        print(f"    • Host Properties: {properties_processed}")
        print(f"    • Findings: {items_processed:,}")
        print(f"    • Additional: {additional_obfuscated}")
    
    def obfuscate_file(self, input_path: Path, output_dir: Path):
        """Obfuscate a single file based on extension."""
        import time
        from datetime import datetime
        
        output_path = output_dir / input_path.name
        extension = input_path.suffix.lower()
        
        # Visual separator
        print(f"\n{'='*70}")
        print(f"FILE: {input_path.name}")
        print(f"TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"TYPE: {extension.upper() if extension else 'NO EXTENSION'}")
        print(f"{'='*70}")
        
        start_time = time.time()
        
        if extension == '.csv':
            self.obfuscate_csv(input_path, output_path)
        elif extension in ['.xlsx', '.xlsm']:
            self.obfuscate_excel(input_path, output_path)
        elif extension in ['.html', '.htm']:
            self.obfuscate_html(input_path, output_path)
        elif extension == '.nessus':
            self.obfuscate_nessus(input_path, output_path)
        elif extension in ['.txt', '.log', '.xml', '.json']:
            # Plain text processing with progress
            print(f"\n{Colors.CYAN}[INFO] Processing text file: {input_path.name}{Colors.RESET}")
            print(f"    [1/3] Reading file...", end='', flush=True)
            with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            file_size_kb = len(content) / 1024
            print(f" Done! ({file_size_kb:.1f} KB)")
            
            print(f"    [2/3] Obfuscating content...", end='', flush=True)
            obfuscated = self.obfuscate_text(content)
            print(f" Done!")
            
            print(f"    [3/3] Writing output...", end='', flush=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(obfuscated)
            print(f" Done!")
            
            print(f"{Colors.GREEN}[SUCCESS] Text file obfuscated: {input_path.name}{Colors.RESET}")
        else:
            print(f"\n{Colors.RED}[ERROR] Unsupported file type: {extension or 'NO EXTENSION'}{Colors.RESET}")
            print(f"    Supported types: .csv, .xlsx, .xlsm, .html, .htm, .nessus, .txt, .log, .xml, .json")
            print(f"    This file should have been filtered out earlier. Skipping...")
            return
        
        elapsed = time.time() - start_time
        output_size_kb = output_path.stat().st_size / 1024 if output_path.exists() else 0
        
        print(f"\n{'-'*70}")
        print(f"{Colors.GREEN}COMPLETED in {elapsed:.1f}s | Output: {output_size_kb:.1f} KB{Colors.RESET}")
        print(f"{'-'*70}")
    
    def print_stats(self):
        """Print obfuscation statistics."""
        print(f"\n{Colors.BOLD}OBFUSCATION STATISTICS:{Colors.RESET}")
        print(f"{'-'*70}")
        
        total_items = sum(self.stats.values())
        
        for key, value in self.stats.items():
            label = key.replace('_', ' ').title()
            percentage = (value / total_items * 100) if total_items > 0 else 0
            bar_length = int(percentage / 2)  # Scale to 50 chars max
            bar = '█' * bar_length + '░' * (50 - bar_length)
            print(f"  {label:<20} {value:>8,}  {bar} {percentage:>5.1f}%")
        
        print(f"{'-'*70}")
        print(f"  {'TOTAL OBFUSCATED':<20} {total_items:>8,}")
        print(f"{'-'*70}")


class Deobfuscator:
    """Reverses obfuscation using mapping file."""
    
    def __init__(self, reverse_mapping_file: Path):
        with open(reverse_mapping_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            self.reverse_mappings = data['reverse_mappings']
        
        # Create combined regex for all tokens
        all_tokens = []
        for category_mappings in self.reverse_mappings.values():
            all_tokens.extend(category_mappings.keys())
        
        # Sort by length (longest first) to avoid partial replacements
        all_tokens.sort(key=len, reverse=True)
        
        self.token_pattern = re.compile('|'.join(re.escape(token) for token in all_tokens))
        
        print(f"{Colors.GREEN}[SUCCESS] Loaded {len(all_tokens)} token mappings for de-obfuscation{Colors.RESET}")
    
    def deobfuscate_text(self, text: str) -> str:
        """Replace all tokens with original values."""
        if not text or not isinstance(text, str):
            return text
        
        def replace_token(match):
            token = match.group(0)
            # Find token in reverse mappings
            for category_mappings in self.reverse_mappings.values():
                if token in category_mappings:
                    return category_mappings[token]
            return token  # Shouldn't happen, but return original if not found
        
        return self.token_pattern.sub(replace_token, text)
    
    def deobfuscate_file(self, input_path: Path, output_path: Path):
        """De-obfuscate a file."""
        print(f"\n{'='*70}")
        print(f"{Colors.CYAN}[INFO] De-obfuscating: {input_path.name}{Colors.RESET}")
        print(f"{'='*70}")
        
        print(f"    [1/3] Reading input file...", end='', flush=True)
        with open(input_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        file_size_kb = len(content) / 1024
        print(f" Done! ({file_size_kb:.1f} KB)")
        
        print(f"    [2/3] Restoring original values...", end='', flush=True)
        deobfuscated = self.deobfuscate_text(content)
        tokens_restored = len(self.token_pattern.findall(content))
        print(f" Done! ({tokens_restored:,} tokens restored)")
        
        print(f"    [3/3] Writing output file...", end='', flush=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(deobfuscated)
        output_size_kb = output_path.stat().st_size / 1024
        print(f" Done! ({output_size_kb:.1f} KB)")
        
        print(f"\n{Colors.GREEN}[SUCCESS] De-obfuscated: {output_path}{Colors.RESET}")
        print(f"{'='*70}")


def main():
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Obfuscate/De-obfuscate ASM scan data for secure AI processing'
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Command to execute')
    
    # Obfuscate command
    obfuscate_parser = subparsers.add_parser('obfuscate', help='Obfuscate sensitive data')
    obfuscate_parser.add_argument('input_dir', help='Directory containing files to obfuscate')
    obfuscate_parser.add_argument('output_dir', help='Directory for obfuscated output')
    obfuscate_parser.add_argument('--mapping-file', default='./mappings/obfuscation_mapping.json',
                                   help='Path to save mapping file (default: ./mappings/obfuscation_mapping.json)')
    obfuscate_parser.add_argument('--pattern', default='*',
                                   help='File pattern to process (default: all files)')
    
    # De-obfuscate command
    deobfuscate_parser = subparsers.add_parser('deobfuscate', 
                                                help='Restore original values from Claude output')
    deobfuscate_parser.add_argument('--input', dest='input_path', default='./output',
                                     help='File or directory to de-obfuscate (default: ./output)')
    deobfuscate_parser.add_argument('--output-dir', dest='output_dir', default='./final',
                                     help='Directory for de-obfuscated files (default: ./final)')
    deobfuscate_parser.add_argument('--mapping-file', default='./mappings/reverse_obfuscation_mapping.json',
                                     help='Path to reverse mapping file (default: ./mappings/reverse_obfuscation_mapping.json)')
    deobfuscate_parser.add_argument('--pattern', default='*',
                                     help='File pattern to process if input is directory (default: all files)')
    deobfuscate_parser.add_argument('--exclude', nargs='*', default=['analysis_progress_tracker.md'],
                                     help='Files to exclude (default: analysis_progress_tracker.md)')
    
    args = parser.parse_args()
    
    if args.command == 'obfuscate':
        import time
        from datetime import datetime
        
        overall_start = time.time()
        
        input_dir = Path(args.input_dir)
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Header
        print(f"\n{'='*70}")
        print(f"{'='*70}")
        print(f"{Colors.BOLD}   ASM DATA OBFUSCATION STARTING{Colors.RESET}")
        print(f"{'='*70}")
        print(f"{'='*70}")
        print(f"\nSTART TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"INPUT DIR:  {input_dir.absolute()}")
        print(f"OUTPUT DIR: {output_dir.absolute()}")
        print(f"PATTERN:    {args.pattern}")
        
        # Create mapper and obfuscator
        mapper = TokenMapper(args.mapping_file)
        obfuscator = ASMObfuscator(mapper)
        
        # Count files first (excluding hidden files)
        all_files = list(input_dir.glob(args.pattern))
        files_to_process = [f for f in all_files if f.is_file() and not f.name.startswith('.')]
        hidden_files_skipped = [f for f in all_files if f.is_file() and f.name.startswith('.')]
        
        total_files = len(files_to_process)
        
        print(f"\nFILES FOUND: {total_files}")
        
        if hidden_files_skipped:
            print(f"{Colors.YELLOW}HIDDEN FILES SKIPPED: {len(hidden_files_skipped)}{Colors.RESET}")
            for hidden_file in hidden_files_skipped:
                print(f"  Skipped: {hidden_file.name} (hidden file)")
        
        if total_files == 0:
            print(f"\n{Colors.RED}[ERROR] No files found matching pattern '{args.pattern}' in {input_dir}{Colors.RESET}")
            print(f"{Colors.YELLOW}[WARNING] Make sure you have files in the input directory.{Colors.RESET}")
            if hidden_files_skipped:
                print(f"{Colors.YELLOW}[WARNING] Note: {len(hidden_files_skipped)} hidden file(s) were skipped (.DS_Store, etc.){Colors.RESET}")
            return 1
        
        # Process all files
        files_processed = 0
        files_skipped = 0
        
        for file_path in files_to_process:
            # Check if file type is supported
            extension = file_path.suffix.lower()
            # UPDATED: Added .nessus to supported extensions
            supported_extensions = ['.csv', '.xlsx', '.xlsm', '.html', '.htm', '.nessus', '.txt', '.log', '.xml', '.json']
            
            if extension not in supported_extensions:
                print(f"\n{Colors.YELLOW}[SKIPPED] {file_path.name} (unsupported type: {extension}){Colors.RESET}")
                files_skipped += 1
                continue
            
            obfuscator.obfuscate_file(file_path, output_dir)
            files_processed += 1
        
        # Save mappings
        print(f"\n{'='*70}")
        print(f"{Colors.CYAN}[INFO] SAVING MAPPING FILES...{Colors.RESET}")
        print(f"{'='*70}")
        mapping_path = Path(args.mapping_file)
        mapping_path.parent.mkdir(parents=True, exist_ok=True)
        mapper.save_mappings(mapping_path)
        
        overall_elapsed = time.time() - overall_start
        
        # Print statistics
        print(f"\n{'='*70}")
        print(f"{'='*70}")
        print(f"{Colors.BOLD}{Colors.GREEN}   OBFUSCATION COMPLETE{Colors.RESET}")
        print(f"{'='*70}")
        print(f"{'='*70}")
        print(f"\n{Colors.BOLD}SUMMARY:{Colors.RESET}")
        print(f"{'-'*70}")
        print(f"  Files Processed:      {files_processed}")
        if files_skipped > 0:
            print(f"  Files Skipped:        {files_skipped} (unsupported types)")
        if hidden_files_skipped:
            print(f"  Hidden Files Ignored: {len(hidden_files_skipped)} (.DS_Store, etc.)")
        print(f"  Total Time:           {overall_elapsed:.1f}s ({overall_elapsed/60:.1f} minutes)")
        if files_processed > 0:
            print(f"  Average per File:     {overall_elapsed/files_processed:.1f}s")
        print(f"{'-'*70}")
        
        if files_processed > 0:
            obfuscator.print_stats()
        
        print(f"\n{'='*70}")
        print(f"END TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*70}")
        
        print(f"\n{Colors.YELLOW}SECURITY REMINDER:{Colors.RESET}")
        print(f"{'-'*70}")
        print(f"  [*] Keep mapping files LOCAL and SECURE")
        print(f"  [*] Never upload mapping files to Claude")
        print(f"  [*] Process obfuscated files through Claude")
        print(f"  [*] Use 'deobfuscate' command on Claude's output")
        print(f"{'-'*70}")
    
    elif args.command == 'deobfuscate':
        import time
        from datetime import datetime
        
        overall_start = time.time()
        
        input_path = Path(args.input_path)
        output_dir = Path(args.output_dir)
        mapping_file = Path(args.mapping_file)
        
        # Header
        print(f"\n{'='*70}")
        print(f"{'='*70}")
        print(f"{Colors.BOLD}   ASM DATA DE-OBFUSCATION STARTING{Colors.RESET}")
        print(f"{'='*70}")
        print(f"{'='*70}")
        print(f"\nSTART TIME:    {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"INPUT:         {input_path.absolute()}")
        print(f"OUTPUT DIR:    {output_dir.absolute()}")
        print(f"MAPPING FILE:  {mapping_file}")
        
        # Check mapping file exists
        if not mapping_file.exists():
            print(f"\n{Colors.RED}[ERROR] Mapping file not found: {mapping_file}{Colors.RESET}")
            print(f"    Make sure you have the reverse mapping file from obfuscation step")
            print(f"\n    Expected location: ./mappings/reverse_<mapping-name>.json")
            print(f"    Or specify with: --mapping-file /path/to/reverse_mapping.json")
            return 1
        
        # Create output directory
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize deobfuscator
        print(f"\n{Colors.CYAN}[INFO] Loading token mappings...{Colors.RESET}")
        deobfuscator = Deobfuscator(mapping_file)
        
        # Determine if input is file or directory
        files_to_process = []
        excluded_files = []
        
        # Files to exclude (internal tracking files)
        exclude_patterns = args.exclude if hasattr(args, 'exclude') and args.exclude else ['analysis_progress_tracker.md']
        
        if input_path.is_file():
            # Check if single file is in exclusion list
            if input_path.name in exclude_patterns:
                print(f"\n{Colors.YELLOW}[WARNING] File is in exclusion list: {input_path.name}{Colors.RESET}")
                print(f"    Use --exclude '' to process this file")
                return 1
            files_to_process = [input_path]
            print(f"\nMODE: Single file")
        elif input_path.is_dir():
            all_files = list(input_path.glob(args.pattern))
            
            # Filter out hidden files and excluded files
            for f in all_files:
                if not f.is_file():
                    continue
                if f.name.startswith('.'):
                    continue
                if f.name in exclude_patterns:
                    excluded_files.append(f)
                    continue
                files_to_process.append(f)
            
            hidden_skipped = [f for f in all_files if f.is_file() and f.name.startswith('.')]
            
            print(f"\nMODE: Directory scan")
            print(f"PATTERN: {args.pattern}")
            print(f"FILES FOUND: {len(files_to_process)}")
            
            if excluded_files:
                print(f"{Colors.YELLOW}EXCLUDED FILES: {len(excluded_files)}{Colors.RESET}")
                for ef in excluded_files:
                    print(f"  Excluded: {ef.name} (internal tracking file)")
            
            if hidden_skipped:
                print(f"{Colors.YELLOW}HIDDEN FILES SKIPPED: {len(hidden_skipped)}{Colors.RESET}")
        else:
            print(f"\n{Colors.RED}[ERROR] Input path does not exist: {input_path}{Colors.RESET}")
            return 1
        
        if not files_to_process:
            print(f"\n{Colors.RED}[ERROR] No files found to de-obfuscate{Colors.RESET}")
            print(f"    Check that files exist in: {input_path}")
            return 1
        
        # Process files
        files_processed = 0
        total_tokens_restored = 0
        
        for file_path in files_to_process:
            output_file = output_dir / file_path.name
            extension = file_path.suffix.lower()
            
            print(f"\n{'='*70}")
            print(f"FILE: {file_path.name}")
            print(f"TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"TYPE: {extension.upper()}")
            print(f"{'='*70}")
            
            print(f"\n{Colors.CYAN}[INFO] De-obfuscating: {file_path.name}{Colors.RESET}")
            
            tokens_in_file = 0
            
            if extension in ['.xlsx', '.xlsm']:
                # Excel file de-obfuscation
                if not EXCEL_AVAILABLE:
                    print(f"{Colors.YELLOW}[SKIPPED] Excel file (openpyxl not installed): {file_path.name}{Colors.RESET}")
                    continue
                
                print(f"    [1/3] Loading Excel workbook...", end='', flush=True)
                workbook = openpyxl.load_workbook(file_path)
                print(f" Done! ({len(workbook.sheetnames)} sheets)")
                
                print(f"    [2/3] Restoring original values...")
                for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
                    sheet = workbook[sheet_name]
                    sheet_tokens = 0
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                original = cell.value
                                cell.value = deobfuscator.deobfuscate_text(cell.value)
                                if original != cell.value:
                                    sheet_tokens += len(deobfuscator.token_pattern.findall(original))
                    tokens_in_file += sheet_tokens
                    print(f"        Sheet {sheet_idx}/{len(workbook.sheetnames)}: {sheet_name} ({sheet_tokens} tokens)")
                
                print(f"    [3/3] Saving workbook...", end='', flush=True)
                workbook.save(output_file)
                print(f" Done!")
                
            elif extension == '.csv':
                # CSV file de-obfuscation
                print(f"    [1/3] Reading CSV file...", end='', flush=True)
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                print(f" Done! ({len(rows)} rows)")
                
                print(f"    [2/3] Restoring original values...", end='', flush=True)
                deobfuscated_rows = []
                for row in rows:
                    new_row = []
                    for cell in row:
                        if cell:
                            tokens_in_file += len(deobfuscator.token_pattern.findall(cell))
                            new_row.append(deobfuscator.deobfuscate_text(cell))
                        else:
                            new_row.append(cell)
                    deobfuscated_rows.append(new_row)
                print(f" Done! ({tokens_in_file} tokens)")
                
                print(f"    [3/3] Writing CSV file...", end='', flush=True)
                with open(output_file, 'w', encoding='utf-8', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerows(deobfuscated_rows)
                print(f" Done!")
            
            elif extension == '.nessus':
                # Nessus XML file de-obfuscation (treat as text/XML)
                print(f"    [1/3] Reading Nessus file...", end='', flush=True)
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                file_size_kb = len(content) / 1024
                print(f" Done! ({file_size_kb:.1f} KB)")
                
                print(f"    [2/3] Restoring original values...", end='', flush=True)
                tokens_in_file = len(deobfuscator.token_pattern.findall(content))
                deobfuscated = deobfuscator.deobfuscate_text(content)
                print(f" Done! ({tokens_in_file:,} tokens)")
                
                print(f"    [3/3] Writing Nessus file...", end='', flush=True)
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(deobfuscated)
                output_size_kb = output_file.stat().st_size / 1024
                print(f" Done! ({output_size_kb:.1f} KB)")
                
            else:
                # Text-based files (md, txt, json, xml, html, etc.)
                print(f"    [1/3] Reading input file...", end='', flush=True)
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                file_size_kb = len(content) / 1024
                print(f" Done! ({file_size_kb:.1f} KB)")
                
                print(f"    [2/3] Restoring original values...", end='', flush=True)
                tokens_in_file = len(deobfuscator.token_pattern.findall(content))
                deobfuscated = deobfuscator.deobfuscate_text(content)
                print(f" Done! ({tokens_in_file:,} tokens)")
                
                print(f"    [3/3] Writing output file...", end='', flush=True)
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(deobfuscated)
                output_size_kb = output_file.stat().st_size / 1024
                print(f" Done! ({output_size_kb:.1f} KB)")
            
            print(f"\n{Colors.GREEN}[SUCCESS] De-obfuscated: {output_file.name}{Colors.RESET}")
            
            files_processed += 1
            total_tokens_restored += tokens_in_file
        
        overall_elapsed = time.time() - overall_start
        
        # Summary
        print(f"\n{'='*70}")
        print(f"{'='*70}")
        print(f"{Colors.BOLD}{Colors.GREEN}   DE-OBFUSCATION COMPLETE{Colors.RESET}")
        print(f"{'='*70}")
        print(f"{'='*70}")
        
        print(f"\n{Colors.BOLD}SUMMARY:{Colors.RESET}")
        print(f"{'-'*70}")
        print(f"  Files Processed:       {files_processed}")
        if excluded_files:
            print(f"  Files Excluded:        {len(excluded_files)} (internal tracking)")
        print(f"  Total Tokens Restored: {total_tokens_restored:,}")
        print(f"  Total Time:            {overall_elapsed:.1f}s")
        print(f"  Output Directory:      {output_dir.absolute()}")
        print(f"{'-'*70}")
        
        print(f"\n{'='*70}")
        print(f"END TIME: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*70}")
        
        print(f"\n{Colors.CYAN}[INFO] Files are ready in: {output_dir.absolute()}{Colors.RESET}")
    
    else:
        parser.print_help()
        return 1


if __name__ == '__main__':
    main()