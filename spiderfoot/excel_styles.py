"""Excel formatting and styling utilities for scan exports.

Provides openpyxl styling helpers, color definitions, and sheet builders
for producing professionally formatted Excel workbooks with tab colors,
severity color-coding, and an Executive Summary with grade visualization.
"""

import re

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# Regex matching XML-illegal control characters that openpyxl rejects.
# Covers: 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F, 0x7F-0x84, 0x86-0x9F
_ILLEGAL_XML_CHARS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f]'
)


def _safe_str(value) -> str:
    """Convert a value to a string safe for openpyxl worksheet cells.

    Strips illegal XML control characters that cause
    openpyxl.utils.exceptions.IllegalCharacterError.
    """
    s = str(value)
    return _ILLEGAL_XML_CHARS_RE.sub('', s)


# ============================================================================
# COLOR UTILITIES
# ============================================================================

def hex_to_argb(hex_color: str) -> str:
    """Convert '#RRGGBB' or 'RRGGBB' to openpyxl ARGB string 'FFRRGGBB'."""
    c = hex_color.lstrip('#')
    return f"FF{c.upper()}"


def set_tab_color(ws, hex_color: str):
    """Set the worksheet tab color.

    Args:
        ws: openpyxl Worksheet
        hex_color: '#RRGGBB' or 'RRGGBB' hex color string
    """
    ws.sheet_properties.tabColor = hex_color.lstrip('#')


# ============================================================================
# SEVERITY / RISK COLORS (mirrors CSS .sev-* classes)
# ============================================================================

SEVERITY_COLORS = {
    'CRITICAL': '#8b5cf6',
    'HIGH':     '#ef4444',
    'MEDIUM':   '#f59e0b',
    'LOW':      '#3b82f6',
    'INFO':     '#22c55e',
}

SEVERITY_TEXT_COLORS = {
    'CRITICAL': '#FFFFFF',
    'HIGH':     '#FFFFFF',
    'MEDIUM':   '#000000',
    'LOW':      '#FFFFFF',
    'INFO':     '#FFFFFF',
}

RISK_COLORS = {
    'CRITICAL':  '#8b5cf6',
    'HIGH':     '#ef4444',
    'MEDIUM':   '#f59e0b',
    'LOW':      '#3b82f6',
    'INFO':     '#22c55e',
}

GRADE_METRIC_COLORS = {
    'A': '#22c55e',
    'B': '#3b82f6',
    'C': '#f59e0b',
    'D': '#f97316',
    'F': '#ef4444',
}

# ============================================================================
# COMMON STYLE ELEMENTS
# ============================================================================

HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    bottom=Side(style='thin', color='FFE5E7EB'),
)

DATA_FONT = Font(name='Calibri', size=10)
DATA_ALIGNMENT = Alignment(horizontal='left', vertical='top', wrap_text=True)

ALT_ROW_FILL = PatternFill(start_color='FFF9FAFB', end_color='FFF9FAFB', fill_type='solid')
DEFAULT_FONT = Font()


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def apply_header_row(ws, headers: list, row: int = 1):
    """Apply dark header styling to a row of cells."""
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_severity_fill(cell, severity: str):
    """Apply severity-appropriate background fill and font color to a cell."""
    sev = severity.upper().strip()
    bg = SEVERITY_COLORS.get(sev)
    fg = SEVERITY_TEXT_COLORS.get(sev)
    if bg:
        cell.fill = PatternFill(
            start_color=hex_to_argb(bg),
            end_color=hex_to_argb(bg),
            fill_type='solid',
        )
    if fg:
        cell.font = Font(name='Calibri', size=10, bold=True, color=hex_to_argb(fg))


def freeze_header(ws, row: int = 2):
    """Freeze panes below a header row."""
    ws.freeze_panes = f'A{row}'


def apply_alternating_rows(ws, start_row: int, end_row: int):
    """Apply subtle alternating row fill for readability.

    Only applies to even rows and skips cells that already have an
    explicit fill (e.g., severity color-coding).
    """
    for row_idx in range(start_row, end_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                # Skip cells that already have an explicit fill applied.
                # A cell with no explicit fill has fill_type=None.
                if cell.fill.fill_type is None:
                    cell.fill = ALT_ROW_FILL


def _make_table_name(base_name: str) -> str:
    """Create a valid Excel table displayName from a base string.

    Table names must start with a letter/underscore and contain only
    alphanumeric characters, underscores, and periods.
    """
    name = ''.join(c if c.isalnum() or c == '_' else '_' for c in base_name)
    if not name or name[0].isdigit():
        name = f"T_{name}"
    return name[:255]


def add_data_table(ws, header_row: int, num_data_rows: int, num_columns: int):
    """Add an Excel Table (ListObject) for sorting and filtering.

    Creates a proper Excel table starting at the header row, which enables
    column sorting, auto-filter dropdowns, and structured references.
    Uses a minimal table style so our manual cell formatting is preserved.

    Args:
        ws: openpyxl Worksheet
        header_row: 1-indexed row number where the headers are
        num_data_rows: number of data rows (excluding the header row)
        num_columns: total number of columns
    """
    if num_data_rows < 1:
        return

    try:
        last_row = header_row + num_data_rows
        last_col = get_column_letter(num_columns)
        ref = f"A{header_row}:{last_col}{last_row}"

        table_name = _make_table_name(ws.title)

        style = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )

        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = style
        ws.add_table(table)
    except Exception:
        pass  # Non-fatal: table is cosmetic, data is still there


# ============================================================================
# EXECUTIVE SUMMARY SHEET BUILDER
# ============================================================================

def build_executive_summary(ws, grade_data: dict, scan_info: dict,
                            findings_rows: list = None,
                            correlation_rows: list = None):
    """Build the Executive Summary worksheet with grade visualization.

    Args:
        ws: openpyxl Worksheet (should be the active/first sheet)
        grade_data: output from calculate_full_grade() -- must contain
                    'overall_grade', 'overall_score', 'overall_grade_color',
                    'overall_grade_bg', and 'categories' dict
        scan_info: dict with keys 'name', 'target', 'date'
        findings_rows: list of finding rows for summary statistics
        correlation_rows: list of correlation rows for summary statistics
    """
    findings_rows = findings_rows or []
    correlation_rows = correlation_rows or []

    set_tab_color(ws, '16a34a')

    dark_fill = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
    subtle_border = Border(bottom=Side(style='thin', color='FFD1D5DB'))
    section_border = Border(bottom=Side(style='medium', color='FF9CA3AF'))
    label_font = Font(name='Calibri', size=10, bold=True, color='FF6B7280')
    value_font = Font(name='Calibri', size=10, color='FF374151')

    # ── Row 1: Title banner ──────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'EXECUTIVE SUMMARY'
    title_cell.font = Font(name='Calibri', size=18, bold=True, color='FFFFFFFF')
    title_cell.fill = dark_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 36
    for c in range(2, 8):
        ws.cell(row=1, column=c).fill = dark_fill

    # ── Row 2: Scan metadata (single row) ────────────────────────────────
    ws.row_dimensions[2].height = 22
    ws.cell(row=2, column=1, value='Target:').font = label_font
    ws.cell(row=2, column=2, value=scan_info.get('target', '')).font = value_font
    ws.cell(row=2, column=4, value='Scan:').font = label_font
    ws.cell(row=2, column=5, value=scan_info.get('name', '')).font = value_font
    ws.cell(row=2, column=6, value='Date:').font = label_font
    ws.cell(row=2, column=7, value=scan_info.get('date', '')).font = value_font
    for c in range(1, 8):
        ws.cell(row=2, column=c).border = subtle_border

    # ── Rows 3-5: Grade block ────────────────────────────────────────────
    overall_grade = grade_data.get('overall_grade', '-')
    overall_score = grade_data.get('overall_score', 0)
    grade_color = grade_data.get('overall_grade_color', '#6b7280')
    grade_bg = grade_data.get('overall_grade_bg', '#f3f4f6')

    ws.row_dimensions[3].height = 8  # spacer

    ws.merge_cells('A4:A6')
    grade_cell = ws['A4']
    grade_cell.value = overall_grade
    grade_cell.font = Font(name='Calibri', size=36, bold=True, color=hex_to_argb(grade_color))
    grade_cell.fill = PatternFill(
        start_color=hex_to_argb(grade_bg),
        end_color=hex_to_argb(grade_bg),
        fill_type='solid',
    )
    grade_cell.alignment = Alignment(horizontal='center', vertical='center')
    grade_bdr = Border(
        left=Side(style='thick', color=hex_to_argb(grade_color)),
        right=Side(style='thick', color=hex_to_argb(grade_color)),
        top=Side(style='thick', color=hex_to_argb(grade_color)),
        bottom=Side(style='thick', color=hex_to_argb(grade_color)),
    )
    grade_cell.border = grade_bdr
    for ri in range(4, 7):
        ws.row_dimensions[ri].height = 22

    ws.cell(row=4, column=2, value='Overall Score').font = Font(
        name='Calibri', size=9, color='FF9CA3AF')
    ws.cell(row=5, column=2, value=overall_score).font = Font(
        name='Calibri', size=20, bold=True, color=hex_to_argb(grade_color))
    ws.cell(row=5, column=2).number_format = '0.0'

    # Per-category mini scores beside the grade
    cat_results = grade_data.get('categories', {})
    sorted_cats = sorted(
        cat_results.items(),
        key=lambda x: (-x[1].get('weight', 0), x[0]),
    )
    if sorted_cats:
        col = 3
        for cat_name, cat_data in sorted_cats:
            if col > 7:
                break
            cat_c = cat_data.get('color', '#6b7280')
            cat_g = cat_data.get('grade', '-')
            ws.cell(row=4, column=col, value=cat_name).font = Font(
                name='Calibri', size=8, color=hex_to_argb(cat_c))
            ws.cell(row=4, column=col).alignment = Alignment(horizontal='center')
            grade_badge = ws.cell(row=5, column=col, value=cat_g)
            grade_badge.font = Font(name='Calibri', size=11, bold=True,
                                    color=hex_to_argb(cat_data.get('grade_color', '#6b7280')))
            grade_badge.fill = PatternFill(
                start_color=hex_to_argb(cat_data.get('grade_bg', '#ffffff')),
                end_color=hex_to_argb(cat_data.get('grade_bg', '#ffffff')),
                fill_type='solid',
            )
            grade_badge.alignment = Alignment(horizontal='center')
            sc = ws.cell(row=6, column=col, value=cat_data.get('score', 0))
            sc.font = Font(name='Calibri', size=8, color='FF9CA3AF')
            sc.number_format = '0.0'
            sc.alignment = Alignment(horizontal='center')
            col += 1

    # ── Row 7: Section divider ───────────────────────────────────────────
    ws.row_dimensions[7].height = 6
    for c in range(1, 8):
        ws.cell(row=7, column=c).border = section_border

    # ── FINDINGS SECTION (rows 8-12) ─────────────────────────────────────
    r = 8
    # Findings banner
    ws.merge_cells(f'A{r}:G{r}')
    findings_banner = ws.cell(row=r, column=1, value=f'FINDINGS  ({len(findings_rows)} total)')
    findings_banner.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    findings_banner.fill = PatternFill(start_color='FF111827', end_color='FF111827', fill_type='solid')
    findings_banner.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = PatternFill(
            start_color='FF111827', end_color='FF111827', fill_type='solid')
    ws.row_dimensions[r].height = 26
    r += 1

    # Severity counts for findings
    sev_counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0}
    for row in findings_rows:
        sev = str(row[0]).upper().strip() if row else ''
        if sev in sev_counts:
            sev_counts[sev] += 1

    # Row of severity labels
    sev_col = 1
    for sev_name in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO']:
        badge = ws.cell(row=r, column=sev_col, value=sev_name)
        badge.font = Font(name='Calibri', size=9, bold=True,
                          color=hex_to_argb(SEVERITY_TEXT_COLORS[sev_name]))
        badge.fill = PatternFill(
            start_color=hex_to_argb(SEVERITY_COLORS[sev_name]),
            end_color=hex_to_argb(SEVERITY_COLORS[sev_name]),
            fill_type='solid',
        )
        badge.alignment = Alignment(horizontal='center')
        sev_col += 1
    ws.row_dimensions[r].height = 20
    r += 1

    # Row of severity counts
    sev_col = 1
    for sev_name in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO']:
        cnt = ws.cell(row=r, column=sev_col, value=sev_counts[sev_name])
        cnt.font = Font(name='Calibri', size=14, bold=True, color='FF1F2937')
        cnt.alignment = Alignment(horizontal='center')
        cnt.border = subtle_border
        sev_col += 1
    ws.row_dimensions[r].height = 24
    r += 1

    # Spacer
    ws.row_dimensions[r].height = 10
    r += 1

    # ── CORRELATIONS SECTION (rows 12-16) ────────────────────────────────
    # Correlations banner - distinct color (dark slate/purple tint)
    ws.merge_cells(f'A{r}:G{r}')
    corr_banner = ws.cell(row=r, column=1, value=f'CORRELATIONS  ({len(correlation_rows)} total)')
    corr_banner.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    corr_banner.fill = PatternFill(start_color='FF374151', end_color='FF374151', fill_type='solid')
    corr_banner.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = PatternFill(
            start_color='FF374151', end_color='FF374151', fill_type='solid')
    ws.row_dimensions[r].height = 26
    r += 1

    # Risk counts for correlations
    risk_counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0}
    for row in correlation_rows:
        risk = str(row[2]).upper().strip() if len(row) > 2 else ''
        if risk in risk_counts:
            risk_counts[risk] += 1

    # Row of risk labels
    risk_col = 1
    for risk_name in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO']:
        badge = ws.cell(row=r, column=risk_col, value=risk_name)
        badge.font = Font(name='Calibri', size=9, bold=True, color='FFFFFFFF')
        badge.fill = PatternFill(
            start_color=hex_to_argb(RISK_COLORS[risk_name]),
            end_color=hex_to_argb(RISK_COLORS[risk_name]),
            fill_type='solid',
        )
        badge.alignment = Alignment(horizontal='center')
        risk_col += 1
    ws.row_dimensions[r].height = 20
    r += 1

    # Row of risk counts
    risk_col = 1
    for risk_name in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'INFO']:
        cnt = ws.cell(row=r, column=risk_col, value=risk_counts[risk_name])
        cnt.font = Font(name='Calibri', size=14, bold=True, color='FF1F2937')
        cnt.alignment = Alignment(horizontal='center')
        cnt.border = subtle_border
        risk_col += 1
    ws.row_dimensions[r].height = 24
    r += 1

    # Spacer + section divider
    ws.row_dimensions[r].height = 6
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = section_border
    r += 1

    # ── Top Risks ──────────────────────────────────────────────────────
    ws.merge_cells(f'A{r}:G{r}')
    ws.cell(row=r, column=1, value='TOP RISKS').font = Font(
        name='Calibri', size=11, bold=True, color='FFDC2626')
    ws.cell(row=r, column=1).border = Border(
        bottom=Side(style='thin', color='FFDC2626'))
    for c in range(2, 8):
        ws.cell(row=r, column=c).border = Border(
            bottom=Side(style='thin', color='FFDC2626'))
    ws.row_dimensions[r].height = 22
    r += 1

    if sorted_cats:
        worst_cats = sorted(sorted_cats, key=lambda x: x[1].get('score', 100))[:3]
        for i, (cname, cdata) in enumerate(worst_cats):
            cscore = cdata.get('score', 0)
            cgrade = cdata.get('grade', '-')
            ccolor = cdata.get('color', '#6b7280')
            cgrade_color = cdata.get('grade_color', '#6b7280')
            cgrade_bg = cdata.get('grade_bg', '#ffffff')

            ws.cell(row=r, column=1, value=f'{i + 1}.').font = Font(
                name='Calibri', size=10, bold=True, color='FF9CA3AF')
            ws.cell(row=r, column=2, value=cname).font = Font(
                name='Calibri', size=10, bold=True, color=hex_to_argb(ccolor))
            ws.cell(row=r, column=3, value=cscore).font = Font(
                name='Calibri', size=10, color='FF6B7280')
            ws.cell(row=r, column=3).number_format = '0.0'
            g_cell = ws.cell(row=r, column=4, value=cgrade)
            g_cell.font = Font(name='Calibri', size=10, bold=True,
                               color=hex_to_argb(cgrade_color))
            g_cell.fill = PatternFill(
                start_color=hex_to_argb(cgrade_bg),
                end_color=hex_to_argb(cgrade_bg),
                fill_type='solid',
            )
            g_cell.alignment = Alignment(horizontal='center')
            desc = cdata.get('description', '')
            if desc:
                ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
                ws.cell(row=r, column=5, value=desc).font = Font(
                    name='Calibri', size=8, italic=True, color='FF9CA3AF')
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = subtle_border
            r += 1
    else:
        ws.cell(row=r, column=1, value='No category data available.').font = Font(
            name='Calibri', size=10, italic=True, color='FF9CA3AF')
        r += 1

    # ── Section divider ──────────────────────────────────────────────────
    ws.row_dimensions[r].height = 6
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = section_border
    r += 1

    # ── Category Breakdown table ─────────────────────────────────────────
    ws.merge_cells(f'A{r}:G{r}')
    ws.cell(row=r, column=1, value='CATEGORY BREAKDOWN').font = Font(
        name='Calibri', size=11, bold=True, color='FF1F2937')
    ws.cell(row=r, column=1).border = Border(
        bottom=Side(style='thin', color='FF1F2937'))
    for c in range(2, 8):
        ws.cell(row=r, column=c).border = Border(
            bottom=Side(style='thin', color='FF1F2937'))
    ws.row_dimensions[r].height = 22
    r += 1

    cat_header_row = r
    if not cat_results or not grade_data.get('enabled', True):
        ws.merge_cells(f'A{r}:G{r}')
        ws.cell(row=r, column=1, value='Grading data not available for this scan.').font = Font(
            name='Calibri', size=10, italic=True, color='FF9CA3AF')
        r += 1
    else:
        cat_headers = ['Category', 'Weight', 'Score', 'Grade', '', '', '']
        for col_num, header_text in enumerate(cat_headers, 1):
            if not header_text:
                continue
            cell = ws.cell(row=r, column=col_num, value=header_text)
            cell.font = Font(name='Calibri', size=9, bold=True, color='FF6B7280')
            cell.border = Border(bottom=Side(style='thin', color='FFD1D5DB'))
        r += 1

        for cat_name, cat_data in sorted_cats:
            cat_color = cat_data.get('color', '#6b7280')
            cat_grade_color = cat_data.get('grade_color', '#6b7280')
            cat_grade_bg = cat_data.get('grade_bg', '#ffffff')

            # Color indicator + name
            name_cell = ws.cell(row=r, column=1, value=cat_name)
            name_cell.font = Font(name='Calibri', size=10, bold=True, color=hex_to_argb(cat_color))
            name_cell.border = Border(
                left=Side(style='thick', color=hex_to_argb(cat_color)),
                bottom=Side(style='thin', color='FFE5E7EB'),
            )

            ws.cell(row=r, column=2, value=cat_data.get('weight', 0)).font = DATA_FONT
            ws.cell(row=r, column=2).number_format = '0.0'
            ws.cell(row=r, column=2).border = subtle_border

            score_cell = ws.cell(row=r, column=3, value=cat_data.get('score', 0))
            score_cell.font = Font(name='Calibri', size=10, bold=True, color='FF374151')
            score_cell.number_format = '0.0'
            score_cell.border = subtle_border

            cat_grade_cell = ws.cell(row=r, column=4, value=cat_data.get('grade', '-'))
            cat_grade_cell.font = Font(name='Calibri', size=10, bold=True,
                                       color=hex_to_argb(cat_grade_color))
            cat_grade_cell.fill = PatternFill(
                start_color=hex_to_argb(cat_grade_bg),
                end_color=hex_to_argb(cat_grade_bg),
                fill_type='solid',
            )
            cat_grade_cell.alignment = Alignment(horizontal='center')
            cat_grade_cell.border = subtle_border

            r += 1

    # Data bars on Score column (cosmetic — non-fatal)
    if sorted_cats and r > cat_header_row + 1:
        try:
            score_range = f"C{cat_header_row + 1}:C{r - 1}"
            rule = DataBarRule(
                start_type='num', start_value=0,
                end_type='num', end_value=100,
                color='3B82F6',
            )
            ws.conditional_formatting.add(score_range, rule)
        except Exception:
            pass

    # ── Column widths (left side) ────────────────────────────────────────
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14


# ============================================================================
# SNAPSHOT SHEET BUILDER (dedicated tab)
# ============================================================================

def build_snapshot_sheet(ws, snapshot_data: dict):
    """Build the SNAPSHOT worksheet — a dedicated security posture trend tab.

    Args:
        ws: openpyxl Worksheet
        snapshot_data: dict with keys 'snapshots', 'overall_change', 'trajectory',
                       'worst_category', 'best_category', 'scan_count',
                       'timeframe_days', 'first_date', 'last_date', and
                       per-snapshot 'category_scores', 'total_findings', 'unique_findings'
    """
    import time as _time

    set_tab_color(ws, '8b5cf6')  # purple tab

    # ── Style constants ────────────────────────────────────────────────────
    dark_fill = PatternFill(start_color='FF111827', end_color='FF111827', fill_type='solid')
    subtitle_fill = PatternFill(start_color='FF1F2937', end_color='FF1F2937', fill_type='solid')
    sub_banner_fill = PatternFill(start_color='FF374151', end_color='FF374151', fill_type='solid')
    alt_row_fill = PatternFill(start_color='FFF9FAFB', end_color='FFF9FAFB', fill_type='solid')
    card_label_font = Font(name='Calibri', size=9, bold=True, color='FF9CA3AF')
    card_sublabel_font = Font(name='Calibri', size=8, italic=True, color='FF9CA3AF')
    subtle_border = Border(bottom=Side(style='thin', color='FFD1D5DB'))
    section_border = Border(bottom=Side(style='medium', color='FF9CA3AF'))

    def _grade_color(grade_letter):
        return GRADE_METRIC_COLORS.get(grade_letter, '#6b7280')

    def _fill_row(row, col_start, col_end, fill):
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill

    # ── Column widths — dynamic: CATEGORY col wide, date cols narrow, TREND fixed
    # We don't know n_cols yet (depends on snapshot count) so set a broad default
    # and override narrow date columns after display_snaps is known (below).
    ws.column_dimensions['A'].width = 24  # CATEGORY
    for _ci in range(2, 20):              # pre-set all possible date + trend cols
        ws.column_dimensions[get_column_letter(_ci)].width = 8

    # ── Extract data ───────────────────────────────────────────────────────
    snapshots = snapshot_data.get('snapshots', [])
    overall_change = snapshot_data.get('overall_change', 0)
    trajectory = snapshot_data.get('trajectory', 'stable')
    _worst_raw = snapshot_data.get('worst_category') or ''
    _best_raw = snapshot_data.get('best_category') or ''
    # Accept both dict {'grade': 'F', 'name': 'Web App Security'} and plain string forms
    if isinstance(_worst_raw, dict):
        worst = _worst_raw
    else:
        worst = {'grade': snapshot_data.get('worst_grade', '-'), 'name': str(_worst_raw)}
    if isinstance(_best_raw, dict):
        best = _best_raw
    else:
        best = {'grade': snapshot_data.get('best_grade', '-'), 'name': str(_best_raw)}
    scan_count = snapshot_data.get('scan_count', len(snapshots))
    timeframe_days = snapshot_data.get('timeframe_days', 0)
    first_date = snapshot_data.get('first_date', 0)
    last_date = snapshot_data.get('last_date', 0)

    # Format date range string
    date_range_str = ''
    if first_date and last_date:
        try:
            d1 = _time.strftime('%b %d', _time.localtime(first_date))
            d2 = _time.strftime('%b %d', _time.localtime(last_date))
            date_range_str = f"{d1} \u2013 {d2}"
        except (OSError, ValueError):
            date_range_str = ''

    # Target name from scan info embedded in snapshot_data
    target_name = snapshot_data.get('target', '')

    # ── ROW 1: Title banner ────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'SECURITY POSTURE SNAPSHOT'
    title_cell.font = Font(name='Calibri', size=18, bold=True, color='FFFFFFFF')
    title_cell.fill = dark_fill
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    _fill_row(1, 2, 10, dark_fill)
    ws.row_dimensions[1].height = 40

    # ── ROW 2: Subtitle bar ───────────────────────────────────────────────
    ws.merge_cells('A2:J2')
    subtitle_parts = []
    if target_name:
        subtitle_parts.append(f"Target: {target_name}")
    if date_range_str:
        subtitle_parts.append(date_range_str)
    subtitle_parts.append(f"{scan_count} scans analyzed")
    subtitle_cell = ws['A2']
    subtitle_cell.value = ' | '.join(subtitle_parts)
    subtitle_cell.font = Font(name='Calibri', size=10, color='FF9CA3AF')
    subtitle_cell.fill = subtitle_fill
    subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
    _fill_row(2, 2, 10, subtitle_fill)
    ws.row_dimensions[2].height = 22

    # ── ROW 3: spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8

    # ── ROW 4: KEY METRICS section label ───────────────────────────────────
    ws.merge_cells('A4:J4')
    ws.cell(row=4, column=1, value='KEY METRICS').font = Font(
        name='Calibri', size=11, bold=True, color='FF1F2937')
    for c in range(1, 11):
        ws.cell(row=4, column=c).border = Border(
            bottom=Side(style='medium', color='FFD1D5DB'))
    ws.row_dimensions[4].height = 18

    # ── ROWS 5-10: 6 metric cards (3×2 grid) ──────────────────────────────
    # Card definitions: (label, value, sublabel, accent_color)
    change_str = f"+{overall_change}" if overall_change > 0 else str(overall_change)
    change_color = '#22c55e' if overall_change > 0 else ('#ef4444' if overall_change < 0 else '#6b7280')

    trend_map = {
        'improving': ('#22c55e', 'IMPROVING'),
        'degrading': ('#ef4444', 'DEGRADING'),
        'stable': ('#6b7280', 'STABLE'),
    }
    trend_color, trend_label = trend_map.get(trajectory, ('#6b7280', 'STABLE'))

    worst_grade = worst.get('grade', '-')
    worst_name = worst.get('name', '-')
    worst_color = _grade_color(worst_grade)

    best_grade = best.get('grade', '-')
    best_name = best.get('name', '-')
    best_color = _grade_color(best_grade)

    tf_val = f"{timeframe_days} days" if timeframe_days else '-'

    purple = '#8b5cf6'
    cyan = '#06b6d4'

    # Row 1 of cards: CHANGE, TREND, WORST (columns A:B, C:D, E:F)
    card_row1 = [
        ('CHANGE', change_str, 'vs first scan', change_color),
        ('TREND', trend_label, 'overall trajectory', trend_color),
        ('WORST', worst_grade, worst_name, worst_color),
    ]
    # Row 2 of cards: BEST, TIMEFRAME, SCANS
    card_row2 = [
        ('BEST', best_grade, best_name, best_color),
        ('TIMEFRAME', tf_val, date_range_str or '-', purple),
        ('SCANS', str(scan_count), 'data points', cyan),
    ]

    def _render_card_row(label_row, value_row, sublabel_row, cards_def):
        for i, (label, value, sublabel, accent) in enumerate(cards_def):
            col_start = 1 + i * 2  # A=1, C=3, E=5
            col_end = col_start + 1

            # Label row
            ws.merge_cells(start_row=label_row, start_column=col_start,
                           end_row=label_row, end_column=col_end)
            lbl = ws.cell(row=label_row, column=col_start, value=label)
            lbl.font = card_label_font
            lbl.alignment = Alignment(horizontal='left')

            # Value row
            ws.merge_cells(start_row=value_row, start_column=col_start,
                           end_row=value_row, end_column=col_end)
            val = ws.cell(row=value_row, column=col_start, value=value)
            val.font = Font(name='Calibri', size=16, bold=True, color=hex_to_argb(accent))
            val.alignment = Alignment(horizontal='left')
            val.border = Border(left=Side(style='thick', color=hex_to_argb(accent)))

            # Sublabel row
            ws.merge_cells(start_row=sublabel_row, start_column=col_start,
                           end_row=sublabel_row, end_column=col_end)
            sub = ws.cell(row=sublabel_row, column=col_start, value=sublabel)
            sub.font = card_sublabel_font
            sub.alignment = Alignment(horizontal='left')

    _render_card_row(5, 6, 7, card_row1)
    _render_card_row(8, 9, 10, card_row2)
    ws.row_dimensions[5].height = 14
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 14
    ws.row_dimensions[8].height = 14
    ws.row_dimensions[9].height = 28
    ws.row_dimensions[10].height = 14

    # ── ROW 11: spacer + divider ──────────────────────────────────────────
    ws.row_dimensions[11].height = 8
    for c in range(1, 11):
        ws.cell(row=11, column=c).border = section_border

    # ── ROW 12: SCAN HISTORY sub-banner ───────────────────────────────────
    ws.merge_cells('A12:J12')
    hist_banner = ws.cell(row=12, column=1, value='SCAN HISTORY')
    hist_banner.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    hist_banner.fill = dark_fill
    hist_banner.alignment = Alignment(horizontal='left', vertical='center')
    _fill_row(12, 2, 10, dark_fill)
    ws.row_dimensions[12].height = 26

    # ── ROW 13: table headers ─────────────────────────────────────────────
    hist_headers = ['DATE', 'GRADE', 'SCORE', 'FINDINGS', 'UNIQUE', 'CHANGE']
    header_font = Font(name='Calibri', size=9, bold=True, color='FF6B7280')
    header_border = Border(bottom=Side(style='thin', color='FFD1D5DB'))
    for ci, hdr in enumerate(hist_headers):
        cell = ws.cell(row=13, column=ci + 1, value=hdr)
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center' if ci > 0 else 'left')
    ws.row_dimensions[13].height = 18

    # ── ROWS 14+: scan data (most recent first, up to 15) ────────────────
    display_snapshots = list(reversed(snapshots))[:15]
    sr = 14
    prev_score = None
    for snap in display_snapshots:
        # Date
        snap_date = snap.get('date', 0)
        if snap_date:
            try:
                date_str = _time.strftime('%b %d', _time.localtime(snap_date))
            except (OSError, ValueError):
                date_str = str(snap_date)
        else:
            date_str = '-'

        snap_grade = snap.get('overall_grade', '-')
        snap_score = snap.get('overall_score', 0)
        total_findings = snap.get('total_findings', 0)
        unique_findings = snap.get('unique_findings', 0)

        # Delta from the next-most-recent scan (prev in reversed order = later in time)
        if prev_score is not None:
            delta = round(snap_score - prev_score, 1)
            delta_str = f"+{delta}" if delta > 0 else str(delta)
            delta_color = '#22c55e' if delta > 0 else ('#ef4444' if delta < 0 else '#6b7280')
        else:
            delta_str = '-'
            delta_color = '#6b7280'
        prev_score = snap_score

        # Alternating row fill
        row_fill = alt_row_fill if (sr - 14) % 2 == 1 else None

        # Date
        d_cell = ws.cell(row=sr, column=1, value=date_str)
        d_cell.font = Font(name='Calibri', size=9, color='FF374151')
        d_cell.border = subtle_border
        if row_fill:
            d_cell.fill = row_fill

        # Grade
        g_color = _grade_color(snap_grade)
        g_cell = ws.cell(row=sr, column=2, value=snap_grade)
        g_cell.font = Font(name='Calibri', size=10, bold=True, color=hex_to_argb(g_color))
        g_cell.alignment = Alignment(horizontal='center')
        g_cell.border = subtle_border
        if row_fill:
            g_cell.fill = row_fill

        # Score
        s_cell = ws.cell(row=sr, column=3, value=snap_score)
        s_cell.font = Font(name='Calibri', size=10, bold=True, color='FF374151')
        s_cell.number_format = '0.0'
        s_cell.alignment = Alignment(horizontal='center')
        s_cell.border = subtle_border
        if row_fill:
            s_cell.fill = row_fill

        # Findings
        f_cell = ws.cell(row=sr, column=4, value=total_findings)
        f_cell.font = Font(name='Calibri', size=9, color='FF374151')
        f_cell.alignment = Alignment(horizontal='center')
        f_cell.border = subtle_border
        if row_fill:
            f_cell.fill = row_fill

        # Unique
        u_cell = ws.cell(row=sr, column=5, value=unique_findings)
        u_cell.font = Font(name='Calibri', size=9, color='FF374151')
        u_cell.alignment = Alignment(horizontal='center')
        u_cell.border = subtle_border
        if row_fill:
            u_cell.fill = row_fill

        # Change
        ch_cell = ws.cell(row=sr, column=6, value=delta_str)
        ch_cell.font = Font(name='Calibri', size=9, bold=True, color=hex_to_argb(delta_color))
        ch_cell.alignment = Alignment(horizontal='center')
        ch_cell.border = subtle_border
        if row_fill:
            ch_cell.fill = row_fill

        sr += 1

    # ── Spacer + section divider after history ────────────────────────────
    ws.row_dimensions[sr].height = 8
    for c in range(1, 11):
        ws.cell(row=sr, column=c).border = section_border
    sr += 1

    # Build category comparison from per-snapshot category_scores
    # Use up to 14 most recent scans (covers a full year of monthly data)
    display_snaps = snapshots[-14:] if len(snapshots) > 14 else snapshots
    # Total columns: CATEGORY + N date cols + TREND
    n_cols = max(10, len(display_snaps) + 2)

    # ── CATEGORY COMPARISON sub-banner ────────────────────────────────────
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=n_cols)
    cat_banner = ws.cell(row=sr, column=1, value='CATEGORY COMPARISON')
    cat_banner.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    cat_banner.fill = sub_banner_fill
    cat_banner.alignment = Alignment(horizontal='left', vertical='center')
    _fill_row(sr, 2, n_cols, sub_banner_fill)
    ws.row_dimensions[sr].height = 26
    sr += 1

    # Collect all category names across all snapshots
    all_cats = {}
    for snap in display_snaps:
        cats = snap.get('category_scores', {})
        for cn, cd in cats.items():
            if cd.get('weight', 0) > 0:
                all_cats[cn] = cd.get('weight', 0)

    if all_cats and len(display_snaps) >= 2:
        # Sort categories by weight descending
        sorted_cat_names = sorted(all_cats.keys(), key=lambda x: (-all_cats[x], x))

        # Header row: CATEGORY, then date columns, then TREND
        cat_header_font = Font(name='Calibri', size=9, bold=True, color='FF6B7280')
        cat_header_border = Border(bottom=Side(style='thin', color='FFD1D5DB'))

        ws.cell(row=sr, column=1, value='CATEGORY').font = cat_header_font
        ws.cell(row=sr, column=1).border = cat_header_border

        for si, snap in enumerate(display_snaps):
            col = si + 2
            snap_date = snap.get('date', 0)
            if snap_date:
                try:
                    hdr_date = _time.strftime('%m/%d', _time.localtime(snap_date))
                except (OSError, ValueError):
                    hdr_date = '-'
            else:
                hdr_date = '-'
            hdr_cell = ws.cell(row=sr, column=col, value=hdr_date)
            hdr_cell.font = cat_header_font
            hdr_cell.alignment = Alignment(horizontal='center')
            hdr_cell.border = cat_header_border

        trend_col = len(display_snaps) + 2
        ws.cell(row=sr, column=trend_col, value='TREND').font = cat_header_font
        ws.cell(row=sr, column=trend_col).border = cat_header_border
        ws.cell(row=sr, column=trend_col).alignment = Alignment(horizontal='center')
        ws.row_dimensions[sr].height = 18
        sr += 1

        # Category rows
        for cat_name in sorted_cat_names:
            cat_tab_color = CATEGORY_TAB_COLORS.get(cat_name, '#6b7280')

            # Category name with thick left border
            name_cell = ws.cell(row=sr, column=1, value=cat_name)
            name_cell.font = Font(name='Calibri', size=9, bold=True,
                                  color=hex_to_argb(cat_tab_color))
            name_cell.border = Border(
                left=Side(style='thick', color=hex_to_argb(cat_tab_color)),
                bottom=Side(style='thin', color='FFE5E7EB'),
            )

            # Grade cells per scan
            first_score = None
            last_score = None
            for si, snap in enumerate(display_snaps):
                col = si + 2
                cats = snap.get('category_scores', {})
                cd = cats.get(cat_name, {})
                grade = cd.get('grade', '-')
                score = cd.get('score', None)

                if score is not None and first_score is None:
                    first_score = score
                if score is not None:
                    last_score = score

                gc = _grade_color(grade)
                g_cell = ws.cell(row=sr, column=col, value=grade)
                g_cell.font = Font(name='Calibri', size=10, bold=True,
                                   color=hex_to_argb(gc))
                g_cell.alignment = Alignment(horizontal='center')
                g_cell.border = Border(bottom=Side(style='thin', color='FFE5E7EB'))

            # Trend arrow
            if first_score is not None and last_score is not None:
                diff = last_score - first_score
                if diff > 10:
                    arrow, arrow_color = '\u25b2\u25b2', '#22c55e'
                elif diff > 2:
                    arrow, arrow_color = '\u25b2', '#22c55e'
                elif diff < -10:
                    arrow, arrow_color = '\u25bc\u25bc', '#ef4444'
                elif diff < -2:
                    arrow, arrow_color = '\u25bc', '#ef4444'
                else:
                    arrow, arrow_color = '\u25c6', '#6b7280'
            else:
                arrow, arrow_color = '\u25c6', '#6b7280'

            t_cell = ws.cell(row=sr, column=trend_col, value=arrow)
            t_cell.font = Font(name='Calibri', size=10, bold=True,
                               color=hex_to_argb(arrow_color))
            t_cell.alignment = Alignment(horizontal='center')
            t_cell.border = Border(bottom=Side(style='thin', color='FFE5E7EB'))

            sr += 1

        # Overall row
        overall_border_top = Border(
            top=Side(style='thick', color='FF374151'),
            bottom=Side(style='thin', color='FFE5E7EB'),
        )
        ov_cell = ws.cell(row=sr, column=1, value='Overall')
        ov_cell.font = Font(name='Calibri', size=10, bold=True, color='FF1F2937')
        ov_cell.border = Border(
            left=Side(style='thick', color='FF374151'),
            top=Side(style='thick', color='FF374151'),
            bottom=Side(style='thin', color='FFE5E7EB'),
        )

        first_ov_score = None
        last_ov_score = None
        for si, snap in enumerate(display_snaps):
            col = si + 2
            grade = snap.get('overall_grade', '-')
            score = snap.get('overall_score', 0)

            if first_ov_score is None:
                first_ov_score = score
            last_ov_score = score

            gc = _grade_color(grade)
            g_cell = ws.cell(row=sr, column=col, value=grade)
            g_cell.font = Font(name='Calibri', size=10, bold=True,
                               color=hex_to_argb(gc))
            g_cell.alignment = Alignment(horizontal='center')
            g_cell.border = overall_border_top

            # Highlight the most recent scan's overall grade
            if si == len(display_snaps) - 1:
                grade_bg = GRADE_METRIC_COLORS.get(grade, '#6b7280')
                # Use a lighter version for the fill
                g_cell.fill = PatternFill(
                    start_color=hex_to_argb(grade_bg),
                    end_color=hex_to_argb(grade_bg),
                    fill_type='solid',
                )
                g_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')

        # Overall trend
        if first_ov_score is not None and last_ov_score is not None:
            ov_diff = last_ov_score - first_ov_score
            if ov_diff > 10:
                ov_arrow, ov_color = '\u25b2\u25b2', '#22c55e'
            elif ov_diff > 2:
                ov_arrow, ov_color = '\u25b2', '#22c55e'
            elif ov_diff < -10:
                ov_arrow, ov_color = '\u25bc\u25bc', '#ef4444'
            elif ov_diff < -2:
                ov_arrow, ov_color = '\u25bc', '#ef4444'
            else:
                ov_arrow, ov_color = '\u25c6', '#6b7280'
        else:
            ov_arrow, ov_color = '\u25c6', '#6b7280'

        ov_t = ws.cell(row=sr, column=trend_col, value=ov_arrow)
        ov_t.font = Font(name='Calibri', size=10, bold=True,
                         color=hex_to_argb(ov_color))
        ov_t.alignment = Alignment(horizontal='center')
        ov_t.border = overall_border_top

    # ── SCORE TREND CHART ────────────────────────────────────────────────
    # Write chart source data in chronological order in a hidden block
    # below all visible content, then build a dark-themed LineChart.
    if len(snapshots) >= 2:
        from openpyxl.chart.shapes import GraphicalProperties as ChartGP
        from openpyxl.chart.text import RichText
        from openpyxl.chart.title import Title as ChartTitle
        from openpyxl.chart.marker import Marker
        from openpyxl.chart.axis import ChartLines
        from openpyxl.drawing.text import (
            Paragraph, ParagraphProperties, CharacterProperties,
            RegularTextRun, Font as DrawingFont,
        )

        sr += 1  # one spacer row
        chart_anchor_row = sr  # chart will be placed here visually
        sr += 1  # another spacer (chart floats over these)
        chart_data_start = sr

        # Header row for chart data (invisible text)
        _hide_font = Font(name='Calibri', size=1, color='FFF9FAFB')
        ws.cell(row=sr, column=1, value='Date').font = _hide_font
        ws.cell(row=sr, column=2, value='Overall Score').font = _hide_font
        ws.row_dimensions[sr].height = 1
        sr += 1

        for snap in snapshots:  # chronological order
            snap_date = snap.get('date', 0)
            if snap_date:
                try:
                    date_label = _time.strftime('%b %d', _time.localtime(snap_date))
                except (OSError, ValueError):
                    date_label = '-'
            else:
                date_label = '-'
            ws.cell(row=sr, column=1, value=date_label).font = _hide_font
            ws.cell(row=sr, column=2, value=snap.get('overall_score', 0))
            ws.row_dimensions[sr].height = 1
            sr += 1

        chart_data_end = sr - 1

        # ── Build the dark-themed LineChart ────────────────────────────
        chart = LineChart()
        chart.width = 30.86  # 12.15" — matches full A:J banner width
        chart.height = 16   # cm — extra height gives title room above plot area

        # Dark chart area background (#111827) with border
        chart_frame = ChartGP()
        chart_frame.solidFill = '111827'
        chart_frame.line.solidFill = '4B5563'
        chart_frame.line.width = 12700  # 1pt
        chart.graphical_properties = chart_frame

        # Dark plot area background (#1F2937)
        plot_frame = ChartGP()
        plot_frame.solidFill = '1F2937'
        plot_frame.line.noFill = True
        chart.plot_area.graphicalProperties = plot_frame

        # ── Axis styling (light text on dark) ──────────────────────────
        _axis_font = CharacterProperties(
            latin=DrawingFont(typeface='Calibri'), sz=900, solidFill='D1D5DB')
        _axis_text = RichText(
            p=[Paragraph(pPr=ParagraphProperties(defRPr=_axis_font),
                         endParaRPr=_axis_font)])

        # Y-axis: 0-100, dark gridlines, no title
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.y_axis.numFmt = '0'
        chart.y_axis.txPr = _axis_text
        chart.y_axis.title = None
        chart.y_axis.delete = False
        # Subtle dark gridlines
        grid_lines = ChartLines()
        grid_gp = ChartGP()
        grid_gp.line.solidFill = '374151'
        grid_gp.line.width = 6350  # 0.5pt
        grid_lines.graphicalProperties = grid_gp
        chart.y_axis.majorGridlines = grid_lines

        # X-axis: date labels, light text
        _x_font = CharacterProperties(
            latin=DrawingFont(typeface='Calibri'), sz=900, b=True,
            solidFill='D1D5DB')
        chart.x_axis.txPr = RichText(
            p=[Paragraph(pPr=ParagraphProperties(defRPr=_x_font),
                         endParaRPr=_x_font)])
        chart.x_axis.title = None
        chart.x_axis.delete = False

        # ── Title (large, bold, white — built with explicit styled runs) ─
        _title_font = CharacterProperties(
            latin=DrawingFont(typeface='Calibri'), sz=1800, b=True,
            solidFill='FFFFFF')
        _title_obj = ChartTitle()
        _title_obj.overlay = False  # push plot area down; title sits above, not over
        _title_obj.tx.rich.paragraphs = [
            Paragraph(
                pPr=ParagraphProperties(defRPr=_title_font),
                r=[RegularTextRun(t='SCORE TREND', rPr=_title_font)],
            )
        ]
        chart.title = _title_obj

        # ── No legend (single series, self-explanatory) ───────────────
        chart.legend = None

        # ── Data + categories ──────────────────────────────────────────
        cats = Reference(ws, min_col=1, min_row=chart_data_start + 1,
                         max_row=chart_data_end)
        score_data = Reference(ws, min_col=2, min_row=chart_data_start,
                               max_row=chart_data_end)
        chart.add_data(score_data, titles_from_data=True)
        chart.set_categories(cats)

        # ── Style the score line ───────────────────────────────────────
        s0 = chart.series[0]
        s0.graphicalProperties.line.width = 32000  # ~2.8pt thick
        s0.graphicalProperties.line.solidFill = 'f59e0b'  # amber
        s0.smooth = True

        # Large circle markers: amber fill, thick white ring
        s0.marker = Marker(symbol='circle', size=12)
        s0_marker_gp = ChartGP()
        s0_marker_gp.solidFill = 'f59e0b'
        s0_marker_gp.line.solidFill = 'ffffff'
        s0_marker_gp.line.width = 19050  # 1.5pt white ring
        s0.marker.graphicalProperties = s0_marker_gp

        # Data labels: score values positioned above markers
        _dlbl_font = CharacterProperties(
            latin=DrawingFont(typeface='Calibri'), sz=1000, b=True,
            solidFill='FFFFFF')
        s0.dLbls = DataLabelList()
        s0.dLbls.showVal = True
        s0.dLbls.showCatName = False
        s0.dLbls.showSerName = False
        s0.dLbls.numFmt = '0.0'
        s0.dLbls.dLblPos = 't'  # position labels above markers
        s0.dLbls.txPr = RichText(
            p=[Paragraph(pPr=ParagraphProperties(defRPr=_dlbl_font),
                         endParaRPr=_dlbl_font)])

        # ── Anchor chart below content, full width of A:J ─────────────
        ws.add_chart(chart, f'A{chart_anchor_row}')

    # ── Print setup (landscape, fit to page) ──────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


# ============================================================================
# FINDINGS SHEET BUILDER
# ============================================================================

def build_findings_sheet(ws, findings_rows: list):
    """Build the Findings worksheet with severity color-coding.

    Args:
        ws: openpyxl Worksheet
        findings_rows: list of [priority, category, tab, item, description, recommendation]
    """
    set_tab_color(ws, '000000')

    headers = ['Priority', 'Category', 'Tab', 'Item', 'Description', 'Recommendation']

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws['A1']
    title.value = f'FINDINGS  ({len(findings_rows)} results)'
    title.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, len(headers) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
    ws.row_dimensions[1].height = 28

    apply_header_row(ws, headers, row=2)
    freeze_header(ws, row=3)

    for row_num, row_data in enumerate(findings_rows, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=_safe_str(cell_value))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # Color-code the Priority cell
        priority = str(row_data[0]).upper().strip() if row_data else ''
        if priority in SEVERITY_COLORS:
            apply_severity_fill(ws.cell(row=row_num, column=1), priority)

    # Alternating rows
    if findings_rows:
        apply_alternating_rows(ws, 3, len(findings_rows) + 2)

    # Excel Table for sort/filter
    add_data_table(ws, header_row=2, num_data_rows=len(findings_rows), num_columns=len(headers))

    col_widths = [12, 22, 18, 40, 50, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# CORRELATIONS SHEET BUILDER
# ============================================================================

def build_correlations_sheet(ws, correlation_rows: list):
    """Build the Correlations worksheet with risk color-coding.

    Args:
        ws: openpyxl Worksheet
        correlation_rows: list of [title, rule_name, risk, description, rule_logic, event_count, event_types]
    """
    set_tab_color(ws, '374151')

    headers = ['Correlation', 'Rule Name', 'Risk', 'Description', 'Rule Logic', 'Event Count', 'Event Types']

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws['A1']
    title.value = f'CORRELATIONS  ({len(correlation_rows)} results)'
    title.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, len(headers) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
    ws.row_dimensions[1].height = 28

    apply_header_row(ws, headers, row=2)
    freeze_header(ws, row=3)

    for row_num, row_data in enumerate(correlation_rows, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=_safe_str(cell_value))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # Color-code the Risk cell (column 3)
        risk = str(row_data[2]).upper().strip() if len(row_data) > 2 else ''
        if risk in RISK_COLORS:
            risk_cell = ws.cell(row=row_num, column=3)
            bg = RISK_COLORS[risk]
            risk_cell.fill = PatternFill(
                start_color=hex_to_argb(bg),
                end_color=hex_to_argb(bg),
                fill_type='solid',
            )
            risk_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')

    if correlation_rows:
        apply_alternating_rows(ws, 3, len(correlation_rows) + 2)

    # Excel Table for sort/filter
    add_data_table(ws, header_row=2, num_data_rows=len(correlation_rows), num_columns=len(headers))

    col_widths = [30, 25, 10, 50, 40, 12, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# EXT-VULNS (NESSUS) SHEET BUILDER
# ============================================================================

def build_nessus_sheet(ws, nessus_rows: list):
    """Build the EXT-VULNS worksheet with severity color-coding and styled headers.

    Args:
        ws: openpyxl Worksheet
        nessus_rows: list of [severity, severity_number, plugin_name, plugin_id,
                     host_ip, host_name, operating_system, description, synopsis,
                     solution, see_also, service_name, port, protocol, request,
                     plugin_output, cvss3_base_score, tracking]
    """
    set_tab_color(ws, 'dc2626')  # red

    headers = [
        "Severity", "Severity Number", "Plugin Name", "Plugin ID",
        "Host IP", "Host Name", "Operating System", "Description",
        "Synopsis", "Solution", "See Also", "Service Name", "Port",
        "Protocol", "Request", "Plugin Output", "CVSS3 Base Score", "Tracking"
    ]

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws['A1']
    title.value = f'EXTERNAL VULNERABILITIES  ({len(nessus_rows)} results)'
    title.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, len(headers) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
    ws.row_dimensions[1].height = 28

    apply_header_row(ws, headers, row=2)
    freeze_header(ws, row=3)

    for row_num, row_data in enumerate(nessus_rows, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=_safe_str(cell_value))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # Color-code the Severity cell (column 1)
        severity = str(row_data[0]).upper().strip() if row_data else ''
        if severity in SEVERITY_COLORS:
            apply_severity_fill(ws.cell(row=row_num, column=1), severity)

    if nessus_rows:
        apply_alternating_rows(ws, 3, len(nessus_rows) + 2)

    # Excel Table for sort/filter
    add_data_table(ws, header_row=2, num_data_rows=len(nessus_rows), num_columns=len(headers))

    col_widths = [12, 8, 30, 10, 14, 20, 18, 50, 40, 40, 30, 12, 8, 8, 40, 40, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# WEBAPP-VULNS (BURP) SHEET BUILDER
# ============================================================================

def build_burp_sheet(ws, burp_rows: list):
    """Build the WEBAPP-VULNS worksheet with severity color-coding and styled headers.

    Args:
        ws: openpyxl Worksheet
        burp_rows: list of [severity, severity_number, host_ip, host_name,
                   plugin_name, issue_type, path, location, confidence,
                   issue_background, issue_detail, solutions, see_also,
                   references, vulnerability_classifications, request,
                   response, tracking]
    """
    set_tab_color(ws, 'ea580c')  # orange

    headers = [
        "Severity", "Severity Number", "Host IP", "Host Name",
        "Plugin Name", "Issue Type", "Path", "Location", "Confidence",
        "Issue Background", "Issue Detail", "Solutions", "See Also",
        "References", "Vulnerability Classifications",
        "Request", "Response", "Tracking"
    ]

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws['A1']
    title.value = f'WEB APPLICATION VULNERABILITIES  ({len(burp_rows)} results)'
    title.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, len(headers) + 1):
        ws.cell(row=1, column=c).fill = HEADER_FILL
    ws.row_dimensions[1].height = 28

    apply_header_row(ws, headers, row=2)
    freeze_header(ws, row=3)

    for row_num, row_data in enumerate(burp_rows, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=_safe_str(cell_value))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # Color-code the Severity cell (column 1)
        severity = str(row_data[0]).upper().strip() if row_data else ''
        if severity in SEVERITY_COLORS:
            apply_severity_fill(ws.cell(row=row_num, column=1), severity)

    if burp_rows:
        apply_alternating_rows(ws, 3, len(burp_rows) + 2)

    # Excel Table for sort/filter
    add_data_table(ws, header_row=2, num_data_rows=len(burp_rows), num_columns=len(headers))

    col_widths = [12, 8, 14, 20, 30, 14, 30, 20, 12, 50, 50, 40, 30, 30, 30, 40, 40, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# CATEGORY TAB BUILDER
# ============================================================================

def build_category_sheet(ws, cat_name: str, cat_color: str, cat_data: dict):
    """Build a per-category worksheet showing the grading detail breakdown.

    Args:
        ws: openpyxl Worksheet
        cat_name: category name (e.g. 'Network Security')
        cat_color: hex color for the tab (e.g. '#dc2626')
        cat_data: per-category dict from calculate_full_grade() containing
                  'score', 'grade', 'grade_color', 'grade_bg', 'weight',
                  'description', and 'details' list
    """
    set_tab_color(ws, cat_color)

    # Category title
    ws.merge_cells('A1:F1')
    title = ws['A1']
    score = cat_data.get('score', 0)
    grade = cat_data.get('grade', '-')
    title.value = f'{cat_name}  \u2014  Score: {score}  ({grade})'
    title.font = Font(name='Calibri', size=14, bold=True, color=hex_to_argb(cat_color))
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30

    # Description
    desc = cat_data.get('description', '')
    if desc:
        ws.merge_cells('A2:F2')
        ws['A2'].value = desc
        ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='FF6B7280')

    # Weight & score summary
    ws.cell(row=3, column=1, value='Weight:').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=3, column=2, value=cat_data.get('weight', 0)).number_format = '0.0'
    ws.cell(row=3, column=3, value='Adj Score:').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=3, column=4, value=cat_data.get('adj_score', 0)).number_format = '0.0'

    # Detail table header
    detail_header_row = 5
    detail_headers = ['Event Type', 'Count', 'Points', 'Logic', 'Rank']
    apply_header_row(ws, detail_headers, detail_header_row)
    freeze_header(ws, detail_header_row + 1)

    details = cat_data.get('details', [])
    current_row = detail_header_row + 1

    if not details:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value='No findings in this category.').font = Font(
            name='Calibri', size=10, italic=True, color='FF9CA3AF',
        )
        current_row += 1
    else:
        for detail in details:
            ws.cell(row=current_row, column=1, value=detail.get('type', '')).font = DATA_FONT
            ws.cell(row=current_row, column=2, value=detail.get('count', 0)).font = DATA_FONT

            pts = detail.get('points', 0)
            pts_cell = ws.cell(row=current_row, column=3, value=pts)
            if pts < 0:
                pts_cell.font = Font(name='Calibri', size=10, color='FFDC2626')
            else:
                pts_cell.font = DATA_FONT
            pts_cell.number_format = '0.0'

            ws.cell(row=current_row, column=4, value=detail.get('logic', '')).font = DATA_FONT
            ws.cell(row=current_row, column=5, value=detail.get('rank', 5)).font = DATA_FONT

            for col in range(1, 6):
                ws.cell(row=current_row, column=col).border = THIN_BORDER

            current_row += 1

        apply_alternating_rows(ws, detail_header_row + 1, current_row - 1)

    col_widths = [35, 10, 10, 20, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# EVENT TYPE DATA TAB BUILDER
# ============================================================================

# Category → tab color mapping (mirrors grade_config DEFAULT_GRADE_CATEGORIES)
CATEGORY_TAB_COLORS = {
    'Network Security':           '#dc2626',
    'Web App Security':           '#ea580c',
    'Information Leakage':        '#eab308',
    'General Health':             '#3b82f6',
    'External Account Exposure':  '#06b6d4',
    'DNS Health':                 '#22c55e',
    'IP Reputation':              '#d946ef',
    'AI Security':                '#8b5cf6',
    'Information / Reference':    '#6b7280',
}


def build_event_type_sheet(ws, event_type_name: str, rows: list, tab_color: str = '#6b7280'):
    """Build a styled per-event-type data worksheet.

    Args:
        ws: openpyxl Worksheet
        event_type_name: human-readable event type (e.g. 'IP Address')
        rows: list of [Updated, Module, Source, F/P, Data]
        tab_color: hex color for the tab
    """
    set_tab_color(ws, tab_color)

    # Title row
    ws.merge_cells('A1:E1')
    title = ws['A1']
    title.value = f'{event_type_name}  ({len(rows)} results)'
    title.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, 6):
        ws.cell(row=1, column=c).fill = HEADER_FILL
    ws.row_dimensions[1].height = 28

    # Column headers
    headers = ['Updated', 'Module', 'Source', 'F/P', 'Data']
    apply_header_row(ws, headers, row=2)
    freeze_header(ws, row=3)

    # Data rows
    fp_labels = {0: '', 1: 'FP', 2: 'OK'}
    fp_fill_red = PatternFill(start_color='FFFEE2E2', end_color='FFFEE2E2', fill_type='solid')
    fp_font_red = Font(name='Calibri', size=9, bold=True, color='FFDC2626')
    fp_fill_green = PatternFill(start_color='FFDCFCE7', end_color='FFDCFCE7', fill_type='solid')
    fp_font_green = Font(name='Calibri', size=9, bold=True, color='FF16A34A')

    for row_num, row_data in enumerate(rows, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=_safe_str(cell_value))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # Style the F/P column (column 4)
        fp_cell = ws.cell(row=row_num, column=4)
        fp_val = row_data[3] if len(row_data) > 3 else 0
        fp_cell.value = fp_labels.get(fp_val, str(fp_val))
        if fp_val == 1:
            fp_cell.fill = fp_fill_red
            fp_cell.font = fp_font_red
        elif fp_val == 2:
            fp_cell.fill = fp_fill_green
            fp_cell.font = fp_font_green
        fp_cell.alignment = Alignment(horizontal='center')

    # Alternating rows
    if rows:
        apply_alternating_rows(ws, 3, len(rows) + 2)

    # Excel Table for sort/filter (starts at row 2 header, below the title banner)
    add_data_table(ws, header_row=2, num_data_rows=len(rows), num_columns=len(headers))

    # Column widths
    col_widths = [18, 20, 40, 6, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sanitize_sheet_name(name: str) -> str:
    """Sanitize a string for use as an Excel sheet name.

    Excel sheet names cannot exceed 31 characters or contain: \\ / ? * [ ] :
    Leading single quotes are also prohibited.
    """
    for ch in ('\\', '/', '?', '*', '[', ']', ':'):
        name = name.replace(ch, '-')
    name = name.strip().lstrip("'")
    return name[:31] if name else 'Sheet'
