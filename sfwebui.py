# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         sfwebui
# Purpose:      User interface class for use with a web browser
#
# Author:       Steve Micallef <steve@binarypool.com>
#
# Created:      30/09/2012
# Copyright:    (c) Steve Micallef 2012
# License:      MIT
# -----------------------------------------------------------------

import base64
import csv
import hashlib
import html
import json
import logging
import multiprocessing as mp
import openpyxl
import random
import re
import string
import threading
import time
import os
import warnings
import uuid
import xml.etree.ElementTree as ET
from datetime import datetime as dt_datetime
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter

import cherrypy
import secure
from cherrypy import _cperror
from mako.lookup import TemplateLookup
from mako.template import Template

import markdown

from sflib import SpiderFoot
from sfscan import startSpiderFootScanner
from spiderfoot import SpiderFootDb
from spiderfoot import SpiderFootHelpers
from spiderfoot import __version__
from spiderfoot.logger import logListenerSetup, logWorkerSetup
from spiderfoot.workspace import SpiderFootWorkspace
from spiderfoot.event_type_mapping import translate_event_type
from spiderfoot.grade_config import (
    calculate_full_grade,
    load_grade_config_overrides,
    get_event_grading,
    DEFAULT_GRADE_CATEGORIES,
)
from spiderfoot.excel_styles import (
    build_executive_summary,
    build_findings_sheet,
    build_correlations_sheet,
    build_nessus_sheet,
    build_burp_sheet,
    build_event_type_sheet,
    sanitize_sheet_name,
    CATEGORY_TAB_COLORS,
    _safe_str,
)

# Use a spawn context for scan subprocesses only, rather than setting it
# globally.  The global setting forces mp.Queue() (used for logging) to use
# heavyweight spawn-mode semaphores/pipes even for same-process threading,
# which leaks semaphores and can crash the CherryPy web server under load.
_spawn_ctx = mp.get_context("spawn")


class SpiderFootWebUi:
    """SpiderFoot web interface."""

    lookup = TemplateLookup(directories=[''], filesystem_checks=True, collection_size=0)
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''
    _correlation_jobs = {}
    _correlation_jobs_lock = threading.Lock()
    # Track scan subprocess Process objects for kill capability
    _scan_processes = {}
    _scan_processes_lock = threading.Lock()

    def __init__(self: 'SpiderFootWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): SpiderFoot config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(
                f"web_config is {type(web_config)}; expected dict()")
        if not web_config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')

        # Register the auth_check CherryPy tool early, before database
        # operations that might fail.  The closure captures 'self' by
        # reference so self.config is resolved at request time (after it
        # has been fully initialised below).
        def check_auth():
            """Before handler: redirect to login if not authenticated."""
            path = cherrypy.request.path_info
            if path.endswith('/login') or path.endswith('/logout') or '/static/' in path or path.endswith('/ping'):
                return
            if not cherrypy.session.get('username'):
                raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
            if not cherrypy.session.get('user_role'):
                try:
                    dbh = SpiderFootDb(self.config)
                    cherrypy.session['user_role'] = dbh.userGetRole(cherrypy.session.get('username'))
                except Exception:
                    cherrypy.session['user_role'] = 'analyst'

        cherrypy.tools.auth_check = cherrypy.Tool('before_handler', check_auth)

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        with SpiderFootDb(self.defaultConfig, init=True) as dbh:
            sf = SpiderFoot(self.defaultConfig)
            self.config = sf.configUnserialize(dbh.configGet(), self.defaultConfig)

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"spiderfoot.{__name__}")

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        # Create default admin user if no users exist
        with SpiderFootDb(self.config, init=True) as dbh_init:
            if dbh_init.userCount() == 0:
                import secrets as _secrets
                default_password = _secrets.token_urlsafe(12)
                dbh_init.userCreate('admin', default_password, display_name='Administrator', role='admin')
                self._default_password = default_password
                self.log.info(f"Created default admin user. Password: {default_password}")
                print("")
                print("*************************************************************")
                print(f" Default login credentials created:")
                print(f"   Username: admin")
                print(f"   Password: {default_password}")
                print(f" Please change this password after first login!")
                print("*************************************************************")
                print("")

        csp = (
            secure.ContentSecurityPolicy()
                .default_src("'self'")
                .script_src("'self'", "'unsafe-inline'", "blob:")
                .style_src("'self'", "'unsafe-inline'")
                .base_uri("'self'")
                .connect_src("'self'", "data:")
                .frame_src("'self'", 'data:')
                .img_src("'self'", "data:")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        # Build header list compatible with secure 0.x, 1.x, and edge versions.
        # CherryPy expects a list of (str, str) tuples.
        header_list = []
        try:
            if hasattr(secure_headers, 'framework'):
                # secure 0.x
                header_list = secure_headers.framework.cherrypy()
            else:
                # secure >= 1.0
                hdrs = secure_headers.headers
                if callable(hdrs):
                    hdrs = hdrs()
                if isinstance(hdrs, dict):
                    header_list = [(str(k), str(v)) for k, v in hdrs.items()]
                elif isinstance(hdrs, (list, tuple)):
                    for item in hdrs:
                        if isinstance(item, (list, tuple)) and len(item) == 2:
                            header_list.append((str(item[0]), str(item[1])))
                        elif hasattr(item, 'header_name') and hasattr(item, 'value'):
                            header_list.append((str(item.header_name), str(item.value)))
        except Exception:
            pass

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": header_list
        })

        # Kill orphaned multiprocessing workers from a previous run.
        # These can hold open SQLite connections that lock the database,
        # preventing startup cleanup from succeeding.
        try:
            import subprocess
            current_pid = os.getpid()
            db_path = self.config.get('__database', '')
            # Derive the project directory from the DB path to identify our workers
            project_dir = os.path.dirname(os.path.dirname(db_path)) if db_path else ''
            if project_dir:
                result = subprocess.run(
                    ['pgrep', '-f', project_dir],
                    capture_output=True, text=True, timeout=5
                )
                if result.stdout.strip():
                    for pid_str in result.stdout.strip().split('\n'):
                        try:
                            stale_pid = int(pid_str.strip())
                            if stale_pid != current_pid and stale_pid > 1:
                                import signal as _sig
                                os.kill(stale_pid, _sig.SIGKILL)
                                self.log.warning(
                                    f"Killed orphaned process {stale_pid} "
                                    f"from previous run"
                                )
                        except (ValueError, OSError):
                            pass
        except Exception:
            pass  # pgrep may not be available on all platforms

        # Mark orphaned scans (RUNNING/STARTING/STARTED) as ABORTED on
        # startup.  If the server is starting, no scan process can still be
        # alive, so any "active" status in the DB is stale.
        # Kill any lingering scan processes first, then update the database.
        active_statuses = ("RUNNING", "STARTING", "STARTED", "ABORT-REQUESTED", "INITIALIZING")
        cleanup_done = False

        for attempt in range(5):
            try:
                with SpiderFootDb(self.config) as dbh_cleanup:
                    scans = dbh_cleanup.scanInstanceList()
                    for scan in scans:
                        if scan[6] in active_statuses:
                            scan_id = scan[0]
                            old_status = scan[6]

                            # Try to kill any lingering scan process by PID
                            try:
                                pid = dbh_cleanup.scanInstanceGetPid(scan_id)
                                if pid and pid > 1:
                                    import signal
                                    try:
                                        os.kill(pid, 0)  # Check if alive
                                        self.log.warning(
                                            f"Startup cleanup: killing orphaned scan "
                                            f"process {scan_id} (pid={pid})"
                                        )
                                        os.kill(pid, signal.SIGKILL)
                                        time.sleep(0.5)
                                    except OSError:
                                        pass  # Process already dead
                            except Exception:
                                pass  # PID column may not exist yet

                            dbh_cleanup.scanInstanceSet(
                                scan_id, status="ABORTED", ended=time.time() * 1000
                            )
                            self.log.warning(
                                f"Startup cleanup: marked orphaned scan {scan_id} "
                                f"as ABORTED (was {old_status})"
                            )
                cleanup_done = True
                break
            except Exception as e:
                if attempt < 4:
                    self.log.warning(
                        f"Startup cleanup attempt {attempt + 1} failed: {e}, retrying..."
                    )
                    time.sleep(2 ** attempt)  # 1s, 2s, 4s, 8s
                else:
                    self.log.error(f"Startup cleanup of orphaned scans failed: {e}")

        # Last resort: use direct SQLite connection if shared handler failed
        if not cleanup_done:
            self.log.warning("Startup cleanup: falling back to direct SQLite connection")
            try:
                import sqlite3
                db_path = self.config.get('__database')
                if db_path:
                    conn = sqlite3.connect(db_path, timeout=30)
                    cursor = conn.execute(
                        "SELECT guid, status FROM tbl_scan_instance WHERE status IN (?, ?, ?, ?, ?)",
                        active_statuses
                    )
                    orphans = cursor.fetchall()
                    for scan_id, old_status in orphans:
                        # Kill process by PID if possible
                        try:
                            pid_cursor = conn.execute(
                                "SELECT pid FROM tbl_scan_instance WHERE guid = ?",
                                (scan_id,)
                            )
                            pid_row = pid_cursor.fetchone()
                            if pid_row and pid_row[0] and pid_row[0] > 1:
                                import signal
                                try:
                                    os.kill(pid_row[0], signal.SIGKILL)
                                    time.sleep(0.5)
                                except OSError:
                                    pass
                        except Exception:
                            pass

                        conn.execute(
                            "UPDATE tbl_scan_instance SET status = 'ABORTED', ended = ? WHERE guid = ?",
                            (time.time() * 1000, scan_id)
                        )
                        self.log.warning(
                            f"Startup cleanup (direct): marked orphaned scan "
                            f"{scan_id} as ABORTED (was {old_status})"
                        )
                    conn.commit()
                    conn.close()
            except Exception as e:
                self.log.error(f"Startup cleanup (direct) also failed: {e}")

    def currentUser(self: 'SpiderFootWebUi') -> str:
        """Get the currently logged-in username from the session.

        Returns:
            str: username or None if not authenticated
        """
        return cherrypy.session.get('username')

    def currentUserRole(self: 'SpiderFootWebUi') -> str:
        """Get the role of the currently logged-in user from the session.

        Returns:
            str: 'admin' or 'analyst'
        """
        return cherrypy.session.get('user_role', 'analyst')

    def requireAuth(self: 'SpiderFootWebUi') -> None:
        """Redirect to login page if user is not authenticated."""
        if not self.currentUser():
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login")

    def requireAdmin(self: 'SpiderFootWebUi') -> None:
        """Redirect to index if user is not an admin."""
        if self.currentUserRole() != 'admin':
            raise cherrypy.HTTPRedirect(f"{self.docroot}/")

    def clientIP(self: 'SpiderFootWebUi') -> str:
        """Get the client IP address from the request.

        Returns:
            str: client IP address
        """
        return cherrypy.request.remote.ip

    def _kill_scan_process(self: 'SpiderFootWebUi', scan_id: str) -> bool:
        """Kill a scan subprocess by scan ID.

        Tries the in-memory Process object first, then falls back to the
        PID stored in the database.  Uses SIGTERM first, then SIGKILL.

        Args:
            scan_id (str): scan instance ID

        Returns:
            bool: True if the process was killed or is already dead
        """
        import signal

        killed = False

        # Try the in-memory Process object first
        with self._scan_processes_lock:
            proc = self._scan_processes.get(scan_id)

        if proc is not None:
            try:
                if proc.is_alive():
                    self.log.warning(f"Terminating scan process {scan_id} (pid={proc.pid})")
                    proc.terminate()
                    proc.join(timeout=3)
                    if proc.is_alive():
                        self.log.warning(f"Force-killing scan process {scan_id} (pid={proc.pid})")
                        proc.kill()
                        proc.join(timeout=2)
                    killed = True
                else:
                    killed = True  # Already dead
            except Exception as e:
                self.log.error(f"Error killing scan process {scan_id} via Process object: {e}")

            # Clean up the dict entry
            with self._scan_processes_lock:
                self._scan_processes.pop(scan_id, None)

            if killed:
                return True

        # Fallback: try PID from database
        pid = 0
        try:
            dbh = SpiderFootDb(self.config)
            pid = dbh.scanInstanceGetPid(scan_id)
        except Exception:
            pass

        if pid and pid > 1:
            try:
                # Check if process is alive
                os.kill(pid, 0)
                self.log.warning(f"Terminating scan process {scan_id} via PID {pid}")
                os.kill(pid, signal.SIGTERM)
                # Wait briefly for it to die
                for _ in range(30):  # 3 seconds
                    time.sleep(0.1)
                    try:
                        os.kill(pid, 0)
                    except OSError:
                        return True  # Process is dead
                # Still alive, force kill
                self.log.warning(f"Force-killing scan process {scan_id} via PID {pid}")
                os.kill(pid, signal.SIGKILL)
                time.sleep(0.5)
                return True
            except OSError:
                return True  # Process already dead
            except Exception as e:
                self.log.error(f"Error killing scan process {scan_id} via PID {pid}: {e}")

        return killed

    def _force_scan_status(self: 'SpiderFootWebUi', scan_id: str, status: str) -> bool:
        """Force-set a scan status using a fresh direct SQLite connection.

        Bypasses the shared SpiderFootDb connection and its class-level lock.
        Used as a last resort when the normal path fails due to a locked
        database.

        Args:
            scan_id (str): scan instance ID
            status (str): terminal status to set (FINISHED, ABORTED, ERROR-FAILED)

        Returns:
            bool: True if the update succeeded
        """
        import sqlite3

        db_path = self.config.get('__database')
        if not db_path:
            return False

        for attempt in range(5):
            try:
                conn = sqlite3.connect(db_path, timeout=10)
                conn.execute(
                    "UPDATE tbl_scan_instance SET status = ?, ended = ? WHERE guid = ?",
                    (status, time.time() * 1000, scan_id)
                )
                conn.commit()
                conn.close()
                return True
            except sqlite3.OperationalError as e:
                if "locked" in str(e) and attempt < 4:
                    time.sleep(2 ** attempt)
                    continue
                self.log.error(f"_force_scan_status failed for {scan_id}: {e}")
                return False
            except Exception as e:
                self.log.error(f"_force_scan_status failed for {scan_id}: {e}")
                return False

        return False

    @cherrypy.expose
    def login(self: 'SpiderFootWebUi', username: str = None, password: str = None) -> str:
        """Login page and handler.

        Args:
            username (str): username for POST
            password (str): password for POST

        Returns:
            str: login page HTML or redirect
        """
        error = None

        if cherrypy.request.method == 'POST' and username and password:
            dbh = SpiderFootDb(self.config)
            if dbh.userVerify(username, password):
                cherrypy.session['username'] = username
                cherrypy.session['user_role'] = dbh.userGetRole(username)
                dbh.userUpdateLastLogin(username)
                dbh.auditLog(username, 'LOGIN', detail='Successful login', ip_address=self.clientIP())
                self.log.info(f"User '{username}' logged in from {self.clientIP()}")
                raise cherrypy.HTTPRedirect(f"{self.docroot}/")
            else:
                dbh.auditLog(username, 'LOGIN_FAILED', detail='Invalid credentials', ip_address=self.clientIP())
                self.log.warning(f"Failed login attempt for user '{username}' from {self.clientIP()}")
                error = "Invalid username or password."

        templ = Template(
            filename='spiderfoot/templates/login.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__, error=error)

    @cherrypy.expose
    def logout(self: 'SpiderFootWebUi') -> str:
        """Logout the current user."""
        username = self.currentUser()
        if username:
            dbh = SpiderFootDb(self.config)
            dbh.auditLog(username, 'LOGOUT', ip_address=self.clientIP())
            self.log.info(f"User '{username}' logged out")
        cherrypy.session.clear()

        raise cherrypy.HTTPRedirect(f"{self.docroot}/login")

    def error_page(self: 'SpiderFootWebUi') -> None:
        """Error page."""
        cherrypy.response.status = 500

        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(
                status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = Template(
            filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'SpiderFootWebUi', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'SpiderFootWebUi', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = Template(
            filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__,
                            pageid='ERROR', user_role=self.currentUserRole())

    def cleanUserInput(self: 'SpiderFootWebUi', inputList: list) -> list:
        """Convert data to HTML entities; except quotes and ampersands.

        Args:
            inputList (list): list of strings to sanitize

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid

        Todo:
            Review all uses of this function, then remove it.
            Use of this function is overloaded.
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append("")
                continue
            
            c = html.escape(item, True)

            # Decode '&' and '"' HTML entities
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def searchBase(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def _compute_fp_flag(self, row_fp_field, event_type, event_data, source_data, targetFps):
        """Compute the false positive flag value.

        Args:
            row_fp_field: the false_positive field from the DB row
            event_type: event type string
            event_data: event data string
            source_data: source data string
            targetFps: set of (event_type, event_data, source_data) tuples from target-level FPs

        Returns:
            int: 0=Unverified, 1=False Positive, 2=Verified
        """
        if row_fp_field == 1 or (event_type, event_data, source_data) in targetFps:
            return 1
        if row_fp_field == 2:
            return 2
        return 0

    @staticmethod
    def _export_filename(scan_name: str, scan_id: str, suffix: str, ext: str) -> str:
        """Build a sanitised export filename using the scan name and a short hash.

        Example: ``FHCSD-ASM-2026_01-FINDINGS-a3c8d7e1.xlsx``

        Args:
            scan_name: human-readable scan name (e.g. ``FHCSD-ASM-2026_01``)
            scan_id:   full scan UUID / hash
            suffix:    descriptive label (e.g. ``REPORT``, ``FINDINGS-CSV``)
            ext:       file extension without dot (e.g. ``xlsx``, ``csv``, ``zip``)

        Returns:
            str: filename safe for Content-Disposition headers
        """
        import re as _re
        # Sanitise scan name: keep alphanumerics, hyphens, underscores, dots
        safe_name = _re.sub(r'[^\w\-.]', '_', scan_name or 'Export').strip('_') or 'Export'
        short_id = (scan_id or 'unknown')[:8]
        return f"{safe_name}-{suffix}-{short_id}.{ext}"

    def buildExcel(self: 'SpiderFootWebUi', data: list, columnNames: list, sheetNameIndex: int = 0, prepend_sheets: list = None) -> str:
        """Convert supplied raw data into Excel format.

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD
            prepend_sheets (list): optional list of dicts with keys "name", "headers", "rows"
                                   to insert as sheets at the front of the workbook

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'

        for row in data:
            sheetName = "".join(
                [c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])[:31]
            try:
                worksheet = workbook[sheetName]
            except KeyError:
                worksheet = workbook.create_sheet(sheetName)
                rowNums[sheetName] = 1
                # Write headers
                for col_num, header in enumerate(columnNames, 1):
                    worksheet.cell(row=1, column=col_num, value=header)
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=rowNums[sheetName], column=col_num, value=_safe_str(cell_value))

            rowNums[sheetName] += 1

        if rowNums or prepend_sheets:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Add any prepend sheets at the beginning of the workbook
        if prepend_sheets:
            for sheet_info in reversed(prepend_sheets):
                ws = workbook.create_sheet(sheet_info["name"], 0)
                for col_num, header in enumerate(sheet_info["headers"], 1):
                    ws.cell(row=1, column=col_num, value=header)
                for row_num, row_data in enumerate(sheet_info["rows"], 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=_safe_str(cell_value))

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'SpiderFootWebUi', id: str, dialect: str = "excel") -> bytes:
        """Get scan log.

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanLogs(id)
        except Exception:
            return json.dumps(self.jsonify_error("404", "Scan ID not found")).encode("utf-8")

        if not data:
            return json.dumps(self.jsonify_error("404", "No scan logs found")).encode("utf-8")

        scan = dbh.scanInstanceGet(id)
        scan_name = scan[0] if scan else ''

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
        for row in data:
            parser.writerow([str(x) for x in row])
        fname = self._export_filename(scan_name, id, 'LOGS', 'csv')
        cherrypy.response.headers[
            'Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan correlation data in CSV, Excel, or HTML format.

        Args:
            id (str): scan ID
            filetype (str): type of file ("xlsx|excel", "csv", or "html")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV, Excel, or HTML format
        """
        dbh = SpiderFootDb(self.config)

        try:
            raw = dbh.scanCorrelationList(id)
            # Transform from scanCorrelationList format to export format
            # DB returns: [id, title, rule_id, rule_risk, rule_name, rule_descr, rule_logic, event_count, event_types]
            # Export needs: [Rule Name, Correlation, Risk, Description]
            data = [[row[4], row[1], row[3], row[5]] for row in raw]
        except Exception:
            return self.error("Scan ID not found")

        try:
            scan = dbh.scanInstanceGet(id)
        except Exception:
            return self.error("Scan ID not found")

        headings = ["Rule Name", "Correlation", "Risk", "Description"]
        scan_name = scan[0] if scan else ''

        if filetype.lower() in ["xlsx", "excel"]:
            fname = self._export_filename(scan_name, id, 'CORRELATIONS', 'xlsx')
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(data, headings)

        if filetype.lower() == 'csv':
            fname = self._export_filename(scan_name, id, 'CORRELATIONS', 'csv')
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"

            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)
            for row in data:
                parser.writerow([str(x) for x in row])
            return fileobj.getvalue()

        if filetype.lower() == 'html':
            # Generate HTML report for correlations
            scan_name = scan[0] if scan else "Unknown"
            html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SpiderFoot Correlations Report</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #333; border-bottom: 3px solid #ffc107; padding-bottom: 15px; margin-bottom: 20px; }
        .summary { background: #fff3cd; padding: 15px; border-radius: 5px; margin-bottom: 20px; }
        .summary p { margin: 5px 0; color: #856404; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 14px; }
        th { background: #343a40; color: white; padding: 12px 8px; text-align: left; }
        td { padding: 12px 8px; border-bottom: 1px solid #dee2e6; vertical-align: top; }
        tr:hover { background: #f8f9fa; }
        tr:nth-child(even) { background: #fafafa; }
        .risk-high { background: #dc3545; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold; }
        .risk-medium { background: #ffc107; color: #333; padding: 3px 8px; border-radius: 3px; font-weight: bold; }
        .risk-low { background: #28a745; color: white; padding: 3px 8px; border-radius: 3px; font-weight: bold; }
        .risk-info { background: #17a2b8; color: white; padding: 3px 8px; border-radius: 3px; }
        .rule-name { font-weight: 600; color: #495057; }
        footer { margin-top: 30px; padding-top: 20px; border-top: 1px solid #dee2e6; color: #6c757d; font-size: 12px; text-align: center; }
    </style>
</head>
<body>
    <div class="container">
        <h1>SpiderFoot Correlations Report</h1>
        <div class="summary">
            <p><strong>Scan:</strong> """ + scan_name + """</p>
            <p><strong>Scan ID:</strong> """ + id + """</p>
            <p><strong>Generated:</strong> """ + time.strftime("%Y-%m-%d %H:%M:%S") + """</p>
            <p><strong>Total Correlations:</strong> """ + str(len(data)) + """</p>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Rule Name</th>
                    <th>Correlation</th>
                    <th>Risk Level</th>
                    <th>Description</th>
                </tr>
            </thead>
            <tbody>
"""
            for row in data:
                rule_name = str(row[0]) if len(row) > 0 else ""
                correlation = str(row[1]) if len(row) > 1 else ""
                risk = str(row[2]) if len(row) > 2 else ""
                description = str(row[3]) if len(row) > 3 else ""

                # Escape HTML entities
                correlation = correlation.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                description = description.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

                # Determine risk class
                risk_lower = risk.lower()
                if 'high' in risk_lower:
                    risk_class = 'risk-high'
                elif 'medium' in risk_lower:
                    risk_class = 'risk-medium'
                elif 'low' in risk_lower:
                    risk_class = 'risk-low'
                else:
                    risk_class = 'risk-info'

                html_content += f"""                <tr>
                    <td class="rule-name">{rule_name}</td>
                    <td>{correlation}</td>
                    <td><span class="{risk_class}">{risk}</span></td>
                    <td>{description}</td>
                </tr>
"""

            html_content += """            </tbody>
        </table>
        <footer>
            <p>Generated by SpiderFoot - Open Source Intelligence Automation</p>
        </footer>
    </div>
</body>
</html>"""

            fname = self._export_filename(scan_name, id, 'CORRELATIONS', 'html')
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "text/html; charset=utf-8"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return html_content.encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexport(self: 'SpiderFootWebUi', id: str, type: str, filetype: str = "csv", dialect: str = "excel", export_mode: str = "full", legacy: str = "0") -> str:
        """Get scan event result data in CSV, Excel, or HTML format.

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel", "csv", or "html")
            dialect (str): CSV dialect (default: excel)
            export_mode (str): "full" (all data), "analysis" (no FPs),
                               or "analysis_correlations" (no FPs + correlations tab, Excel only)
            legacy (str): "1" to use legacy v4.0 type mapping, "0" for native types (default)

        Returns:
            str: results in CSV, Excel, or HTML format
        """
        use_legacy = (str(legacy) == "1")
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, type)
        filter_fps = export_mode in ("analysis", "analysis_correlations")

        # Force Excel for analysis_correlations mode (correlations tab requires .xlsx)
        if export_mode == "analysis_correlations":
            filetype = "excel"

        # Get target-level false positives for this scan
        scanInfo = dbh.scanInstanceGet(id)
        target = scanInfo[1] if scanInfo else None
        targetFps = set()
        if target:
            try:
                targetFps = dbh.targetFalsePositivesForTarget(target)
            except Exception:
                pass  # Table may not exist in older databases

        # Build prepend sheets for analysis_correlations mode (Findings + Correlations)
        prepend_sheets = None
        if export_mode == "analysis_correlations":
            # Findings sheet
            findings_rows = []
            try:
                findings_data = dbh.scanFindingsList(id)
                for f_row in findings_data:
                    findings_rows.append([
                        str(f_row[1]),   # Priority
                        str(f_row[2]),   # Category
                        str(f_row[3]),   # Tab
                        str(f_row[4]),   # Item
                        str(f_row[5]),   # Description
                        str(f_row[6]),   # Recommendation
                    ])
            except Exception:
                pass

            # Correlations sheet
            correlation_rows = []
            try:
                corr_data = dbh.scanCorrelationList(id)
                for corr_row in corr_data:
                    correlation_rows.append([
                        str(corr_row[1]),   # Title
                        str(corr_row[4]),   # Rule Name
                        str(corr_row[3]),   # Risk
                        str(corr_row[5]),   # Description
                        str(corr_row[6]),   # Rule Logic
                        str(corr_row[7]),   # Event Count
                        str(corr_row[8] or ''),  # Event Types
                    ])
            except Exception:
                pass
            prepend_sheets = [
                {
                    "name": "Findings",
                    "headers": ["Priority", "Category", "Tab", "Item", "Description", "Recommendation"],
                    "rows": findings_rows
                },
                {
                    "name": "Correlations",
                    "headers": ["Correlation", "Rule Name", "Risk", "Description", "Rule Logic", "Event Count", "Event Types"],
                    "rows": correlation_rows
                }
            ]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                fp_flag = self._compute_fp_flag(row[13], row[4], row[1], row[2], targetFps)
                if filter_fps and fp_flag == 1:
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]), use_legacy=use_legacy)
                rows.append([lastseen, event_type, str(row[3]),
                            str(row[2]), fp_flag, datafield])

            _scan_name = scanInfo[0] if scanInfo else ''
            if export_mode == "analysis_correlations":
                fname = self._export_filename(_scan_name, id, 'ANALYSIS-CORRELATIONS', 'xlsx')
            elif export_mode == "analysis":
                fname = self._export_filename(_scan_name, id, 'ANALYSIS', 'xlsx')
            else:
                fname = self._export_filename(_scan_name, id, 'DATA', 'xlsx')
            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1,
                                   prepend_sheets=prepend_sheets)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(
                ["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                fp_flag = self._compute_fp_flag(row[13], row[4], row[1], row[2], targetFps)
                if filter_fps and fp_flag == 1:
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]), use_legacy=use_legacy)
                parser.writerow([lastseen, event_type, str(
                    row[3]), str(row[2]), fp_flag, datafield])

            _scan_name = scanInfo[0] if scanInfo else ''
            if export_mode == "analysis":
                fname = self._export_filename(_scan_name, id, 'ANALYSIS', 'csv')
            else:
                fname = self._export_filename(_scan_name, id, 'DATA', 'csv')
            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        if filetype.lower() == 'html':
            # Generate HTML report
            scan_name = scanInfo[0] if scanInfo else "Unknown"
            html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SpiderFoot Scan Report</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }
        .container { max-width: 1400px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #333; border-bottom: 3px solid #28a745; padding-bottom: 15px; margin-bottom: 20px; }
        .summary { background: #e8f5e9; padding: 15px; border-radius: 5px; margin-bottom: 20px; }
        .summary p { margin: 5px 0; color: #2e7d32; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 14px; }
        th { background: #343a40; color: white; padding: 12px 8px; text-align: left; position: sticky; top: 0; }
        td { padding: 10px 8px; border-bottom: 1px solid #dee2e6; vertical-align: top; }
        tr:hover { background: #f8f9fa; }
        tr:nth-child(even) { background: #fafafa; }
        tr:nth-child(even):hover { background: #f0f0f0; }
        .fp-yes { color: #dc3545; font-weight: bold; }
        .fp-no { color: #28a745; }
        .data-cell { max-width: 400px; word-wrap: break-word; font-family: monospace; font-size: 12px; }
        .type-badge { background: #007bff; color: white; padding: 2px 8px; border-radius: 3px; font-size: 12px; white-space: nowrap; }
        .module-badge { background: #6c757d; color: white; padding: 2px 6px; border-radius: 3px; font-size: 11px; }
        .timestamp { color: #666; font-size: 12px; white-space: nowrap; }
        footer { margin-top: 30px; padding-top: 20px; border-top: 1px solid #dee2e6; color: #6c757d; font-size: 12px; text-align: center; }
    </style>
</head>
<body>
    <div class="container">
        <h1>SpiderFoot Scan Report</h1>
        <div class="summary">
            <p><strong>Scan:</strong> """ + scan_name + """</p>
            <p><strong>Scan ID:</strong> """ + id + """</p>
            <p><strong>Generated:</strong> """ + time.strftime("%Y-%m-%d %H:%M:%S") + """</p>
            <p><strong>Total Results:</strong> """ + str(len([r for r in data if r[4] != "ROOT"])) + """</p>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Timestamp</th>
                    <th>Type</th>
                    <th>Module</th>
                    <th>Source</th>
                    <th>F/P</th>
                    <th>Data</th>
                </tr>
            </thead>
            <tbody>
"""
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                # Escape HTML entities
                datafield = datafield.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                event_type = translate_event_type(str(row[4]), use_legacy=use_legacy)
                fp_flag = self._compute_fp_flag(row[13], row[4], row[1], row[2], targetFps)
                fp_display = '<span class="fp-yes">Yes</span>' if fp_flag else '<span class="fp-no">No</span>'

                html_content += f"""                <tr>
                    <td class="timestamp">{lastseen}</td>
                    <td><span class="type-badge">{event_type}</span></td>
                    <td><span class="module-badge">{row[3]}</span></td>
                    <td>{row[2]}</td>
                    <td>{fp_display}</td>
                    <td class="data-cell">{datafield}</td>
                </tr>
"""

            html_content += """            </tbody>
        </table>
        <footer>
            <p>Generated by SpiderFoot - Open Source Intelligence Automation</p>
        </footer>
    </div>
</body>
</html>"""

            fname = self._export_filename(scan_name, id, 'REPORT', 'html')
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "text/html; charset=utf-8"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return html_content.encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scandiscoverypathexport(self: 'SpiderFootWebUi', id: str, eventType: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Export discovery path data as CSV or Excel.

        Flattens the hierarchical discovery path tree into table rows.
        Each node in the path gets its own set of columns (Type, Source Module, Data),
        ordered Root (left) to Leaf (right). Rows follow the same ordering as the
        full data view.

        Args:
            id (str): scan ID
            eventType (str): event type filter
            filetype (str): "csv" or "xlsx"/"excel"
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: CSV or Excel data as file download
        """
        dbh = SpiderFootDb(self.config)
        scan = dbh.scanInstanceGet(id)
        _scan_name = scan[0] if scan else ''

        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            fname = self._export_filename(_scan_name, id, 'DISCOVERY-PATH', 'csv')
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/csv'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return b''

        if 'ROOT' in pc:
            del pc['ROOT']
        tree = SpiderFootHelpers.dataParentChildToTree(pc)

        # Build leaf hash -> path (list of node dicts) by walking the tree
        leaf_paths = {}

        def walk(node, path_so_far):
            node_name = node.get('name', '')
            node_data = datamap.get(node_name)
            current_path = list(path_so_far)
            if node_data:
                data_str = str(node_data[1]).replace('<SFURL>', '').replace('</SFURL>', '')
                current_path.append({
                    'type': str(node_data[10]) if node_data[10] else str(node_data[4]),
                    'module': str(node_data[3]) if node_data[3] else '',
                    'data': data_str,
                    'is_root': str(node_data[4]) == 'ROOT'
                })

            children = node.get('children')
            if not children:
                if node_data and str(node_data[4]) != 'ROOT':
                    leaf_paths[str(node_data[8])] = current_path
            else:
                for child in children:
                    walk(child, current_path)

        if tree:
            walk(tree, [])

        # Find max path depth
        max_depth = 0
        for path in leaf_paths.values():
            if len(path) > max_depth:
                max_depth = len(path)

        if max_depth == 0:
            max_depth = 1

        # Build column headers: Status + per-node (Type, Source Module, Data)
        column_names = ["Status"]
        for d in range(max_depth):
            if d == 0:
                prefix = "Root"
            elif d == max_depth - 1 and max_depth > 1:
                prefix = "Leaf"
            elif max_depth <= 3:
                prefix = "Branch"
            else:
                prefix = f"Branch {d}"
            column_names.extend([f"{prefix} Type", f"{prefix} Source Module", f"{prefix} Data"])

        # Build rows in leafSet order (same as full data view)
        rows = []
        for leaf_row in leafSet:
            if str(leaf_row[4]) == 'ROOT':
                continue
            leaf_hash = str(leaf_row[8])
            fp_flag = leaf_row[13]
            status = "FALSE POSITIVE" if fp_flag == 1 else ("VALIDATED" if fp_flag == 2 else "UNVALIDATED")

            path = leaf_paths.get(leaf_hash, [])
            row = [status]
            for d in range(max_depth):
                if d < len(path):
                    node = path[d]
                    row.extend([node['type'], node['module'], node['data']])
                else:
                    row.extend(['', '', ''])
            rows.append(row)

        if filetype.lower() in ["xlsx", "excel"]:
            # For buildExcel, prepend the leaf type as sheet name column
            excel_rows = []
            for i, leaf_row in enumerate(leafSet):
                if str(leaf_row[4]) == 'ROOT':
                    continue
                leaf_hash = str(leaf_row[8])
                path = leaf_paths.get(leaf_hash, [])
                # Use the leaf's event type description as the sheet name
                sheet_name = str(leaf_row[10]) if leaf_row[10] else str(leaf_row[4])
                idx = len(excel_rows)
                if idx < len(rows):
                    excel_rows.append([sheet_name] + rows[idx])
            excel_col_names = ["_SheetName"] + column_names

            fname = self._export_filename(_scan_name, id, 'DISCOVERY-PATH', 'xlsx')
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            cherrypy.response.headers['Pragma'] = 'no-cache'
            return self.buildExcel(excel_rows, excel_col_names, sheetNameIndex=0)

        # Default: CSV
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(column_names)
        for row in rows:
            parser.writerow(row)

        fname = self._export_filename(_scan_name, id, 'DISCOVERY-PATH', 'csv')
        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
        cherrypy.response.headers['Content-Type'] = 'application/csv'
        cherrypy.response.headers['Pragma'] = 'no-cache'
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'SpiderFootWebUi', ids: str, filetype: str = "csv", dialect: str = "excel", export_mode: str = "full", legacy: str = "0") -> str:
        """Get scan event result data in CSV, Excel, or HTML format for multiple
        scans.

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel", "csv", or "html")
            dialect (str): CSV dialect (default: excel)
            export_mode (str): "full" (all data), "analysis" (no FPs),
                               or "analysis_correlations" (no FPs + correlations tab, Excel only)
            legacy (str): "1" to use legacy v4.0 type mapping, "0" for native types (default)

        Returns:
            str: results in CSV, Excel, or HTML format
        """
        use_legacy = (str(legacy) == "1")
        dbh = SpiderFootDb(self.config)
        scaninfo = dict()
        targetFpsPerScan = dict()  # Store target FPs per scan ID
        data = list()
        scan_name = ""
        filter_fps = export_mode in ("analysis", "analysis_correlations")

        # Force Excel for analysis_correlations mode (correlations tab requires .xlsx)
        if export_mode == "analysis_correlations":
            filetype = "excel"

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            # Get target-level false positives for this scan
            target = scaninfo[id][1] if scaninfo[id] else None
            targetFpsPerScan[id] = set()
            if target:
                try:
                    targetFpsPerScan[id] = dbh.targetFalsePositivesForTarget(target)
                except Exception:
                    pass  # Table may not exist in older databases
            data = data + dbh.scanResultEvent(id)

        # Build prepend sheets for analysis_correlations mode (Findings + Correlations)
        prepend_sheets = None
        if export_mode == "analysis_correlations":
            # Findings sheet (multi-scan)
            findings_rows = []
            for scan_id in ids.split(','):
                if scan_id not in scaninfo or scaninfo[scan_id] is None:
                    continue
                try:
                    findings_data = dbh.scanFindingsList(scan_id)
                    scan_name_display = scaninfo[scan_id][0] if scaninfo[scan_id] else "Unknown"
                    for f_row in findings_data:
                        findings_rows.append([
                            scan_name_display,
                            str(f_row[1]),   # Priority
                            str(f_row[2]),   # Category
                            str(f_row[3]),   # Tab
                            str(f_row[4]),   # Item
                            str(f_row[5]),   # Description
                            str(f_row[6]),   # Recommendation
                        ])
                except Exception:
                    pass

            # Correlations sheet (multi-scan)
            correlation_rows = []
            for scan_id in ids.split(','):
                if scan_id not in scaninfo or scaninfo[scan_id] is None:
                    continue
                try:
                    corr_data = dbh.scanCorrelationList(scan_id)
                    scan_name_display = scaninfo[scan_id][0] if scaninfo[scan_id] else "Unknown"
                    for corr_row in corr_data:
                        correlation_rows.append([
                            scan_name_display,
                            str(corr_row[1]),   # Title
                            str(corr_row[4]),   # Rule Name
                            str(corr_row[3]),   # Risk
                            str(corr_row[5]),   # Description
                            str(corr_row[6]),   # Rule Logic
                            str(corr_row[7]),   # Event Count
                            str(corr_row[8] or ''),  # Event Types
                        ])
                except Exception:
                    pass
            prepend_sheets = [
                {
                    "name": "Findings",
                    "headers": ["Scan Name", "Priority", "Category", "Tab", "Item", "Description", "Recommendation"],
                    "rows": findings_rows
                },
                {
                    "name": "Correlations",
                    "headers": ["Scan Name", "Correlation", "Rule Name", "Risk", "Description", "Rule Logic", "Event Count", "Event Types"],
                    "rows": correlation_rows
                }
            ]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                scan_id = row[12]
                targetFps = targetFpsPerScan.get(scan_id, set())
                fp_flag = self._compute_fp_flag(row[13], row[4], row[1], row[2], targetFps)
                if filter_fps and fp_flag == 1:
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]), use_legacy=use_legacy)
                rows.append([scaninfo[row[12]][0], lastseen, event_type, str(row[3]),
                            str(row[2]), fp_flag, datafield])

            _name = scan_name if scan_name and len(ids.split(',')) == 1 else 'Multi-Scan'
            _id = ids.split(',')[0] if ids else ''
            if export_mode == "analysis_correlations":
                fname = self._export_filename(_name, _id, 'ANALYSIS-CORRELATIONS', 'xlsx')
            elif export_mode == "analysis":
                fname = self._export_filename(_name, _id, 'ANALYSIS', 'xlsx')
            else:
                fname = self._export_filename(_name, _id, 'DATA', 'xlsx')

            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Scan Name", "Updated", "Type", "Module",
                                   "Source", "F/P", "Data"], sheetNameIndex=2,
                                   prepend_sheets=prepend_sheets)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Scan Name", "Updated", "Type",
                            "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                scan_id = row[12]
                targetFps = targetFpsPerScan.get(scan_id, set())
                fp_flag = self._compute_fp_flag(row[13], row[4], row[1], row[2], targetFps)
                if filter_fps and fp_flag == 1:
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]), use_legacy=use_legacy)
                parser.writerow([scaninfo[row[12]][0], lastseen, event_type, str(row[3]),
                                str(row[2]), fp_flag, datafield])

            _name = scan_name if scan_name and len(ids.split(',')) == 1 else 'Multi-Scan'
            _id = ids.split(',')[0] if ids else ''
            if export_mode == "analysis":
                fname = self._export_filename(_name, _id, 'ANALYSIS', 'csv')
            else:
                fname = self._export_filename(_name, _id, 'DATA', 'csv')

            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        if filetype.lower() == 'html':
            # Generate HTML report
            html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SpiderFoot Scan Report</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }
        .container { max-width: 1400px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #333; border-bottom: 3px solid #28a745; padding-bottom: 15px; margin-bottom: 20px; }
        .summary { background: #e8f5e9; padding: 15px; border-radius: 5px; margin-bottom: 20px; }
        .summary p { margin: 5px 0; color: #2e7d32; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 14px; }
        th { background: #343a40; color: white; padding: 12px 8px; text-align: left; position: sticky; top: 0; }
        td { padding: 10px 8px; border-bottom: 1px solid #dee2e6; vertical-align: top; }
        tr:hover { background: #f8f9fa; }
        tr:nth-child(even) { background: #fafafa; }
        tr:nth-child(even):hover { background: #f0f0f0; }
        .fp-yes { color: #dc3545; font-weight: bold; }
        .fp-no { color: #28a745; }
        .data-cell { max-width: 400px; word-wrap: break-word; font-family: monospace; font-size: 12px; }
        .type-badge { background: #007bff; color: white; padding: 2px 8px; border-radius: 3px; font-size: 12px; white-space: nowrap; }
        .module-badge { background: #6c757d; color: white; padding: 2px 6px; border-radius: 3px; font-size: 11px; }
        .timestamp { color: #666; font-size: 12px; white-space: nowrap; }
        .scan-name { font-weight: 500; color: #495057; }
        footer { margin-top: 30px; padding-top: 20px; border-top: 1px solid #dee2e6; color: #6c757d; font-size: 12px; text-align: center; }
    </style>
</head>
<body>
    <div class="container">
        <h1>SpiderFoot Scan Report</h1>
        <div class="summary">
            <p><strong>Generated:</strong> """ + time.strftime("%Y-%m-%d %H:%M:%S") + """</p>
            <p><strong>Scans Included:</strong> """ + ids + """</p>
            <p><strong>Total Results:</strong> """ + str(len([r for r in data if r[4] != "ROOT"])) + """</p>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Scan Name</th>
                    <th>Timestamp</th>
                    <th>Type</th>
                    <th>Module</th>
                    <th>Source</th>
                    <th>F/P</th>
                    <th>Data</th>
                </tr>
            </thead>
            <tbody>
"""
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                # Escape HTML entities
                datafield = datafield.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                event_type = translate_event_type(str(row[4]), use_legacy=use_legacy)
                scan_id = row[12]
                scan_name_display = scaninfo[scan_id][0] if scan_id in scaninfo and scaninfo[scan_id] else "Unknown"
                targetFps = targetFpsPerScan.get(scan_id, set())
                fp_flag = self._compute_fp_flag(row[13], row[4], row[1], row[2], targetFps)
                fp_display = '<span class="fp-yes">Yes</span>' if fp_flag else '<span class="fp-no">No</span>'

                html_content += f"""                <tr>
                    <td class="scan-name">{scan_name_display}</td>
                    <td class="timestamp">{lastseen}</td>
                    <td><span class="type-badge">{event_type}</span></td>
                    <td><span class="module-badge">{row[3]}</span></td>
                    <td>{row[2]}</td>
                    <td>{fp_display}</td>
                    <td class="data-cell">{datafield}</td>
                </tr>
"""

            html_content += """            </tbody>
        </table>
        <footer>
            <p>Generated by SpiderFoot - Open Source Intelligence Automation</p>
        </footer>
    </div>
</body>
</html>"""

            _name = scan_name if scan_name and len(ids.split(',')) == 1 else 'Multi-Scan'
            _id = ids.split(',')[0] if ids else ''
            fname = self._export_filename(_name, _id, 'REPORT', 'html')

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "text/html; charset=utf-8"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return html_content.encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scansearchresultexport(self: 'SpiderFootWebUi', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel", export_mode: str = "full", legacy: str = "0") -> str:
        """Get search result data in CSV or Excel format.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)
            export_mode (str): "full" (all data), "analysis" (no FPs),
                               or "analysis_correlations" (no FPs + correlations tab, Excel only)
            legacy (str): "1" to use legacy v4.0 type mapping, "0" for native types (default)

        Returns:
            str: results in CSV or Excel format
        """
        use_legacy = (str(legacy) == "1")
        data = self.searchBase(id, eventType, value)
        filter_fps = export_mode in ("analysis", "analysis_correlations")

        # Force Excel for analysis_correlations mode (correlations tab requires .xlsx)
        if export_mode == "analysis_correlations":
            filetype = "excel"

        # Get target-level false positives for this scan
        dbh = SpiderFootDb(self.config)
        scanInfo = dbh.scanInstanceGet(id)
        target = scanInfo[1] if scanInfo else None
        targetFps = set()
        if target:
            try:
                targetFps = dbh.targetFalsePositivesForTarget(target)
            except Exception:
                pass  # Table may not exist in older databases

        # Build prepend sheets for analysis_correlations mode (Findings + Correlations)
        prepend_sheets = None
        if export_mode == "analysis_correlations":
            # Findings sheet
            findings_rows = []
            try:
                findings_data = dbh.scanFindingsList(id)
                for f_row in findings_data:
                    findings_rows.append([
                        str(f_row[1]),   # Priority
                        str(f_row[2]),   # Category
                        str(f_row[3]),   # Tab
                        str(f_row[4]),   # Item
                        str(f_row[5]),   # Description
                        str(f_row[6]),   # Recommendation
                    ])
            except Exception:
                pass

            # Correlations sheet
            correlation_rows = []
            try:
                corr_data = dbh.scanCorrelationList(id)
                for corr_row in corr_data:
                    correlation_rows.append([
                        str(corr_row[1]),   # Title
                        str(corr_row[4]),   # Rule Name
                        str(corr_row[3]),   # Risk
                        str(corr_row[5]),   # Description
                        str(corr_row[6]),   # Rule Logic
                        str(corr_row[7]),   # Event Count
                        str(corr_row[8] or ''),  # Event Types
                    ])
            except Exception:
                pass
            prepend_sheets = [
                {
                    "name": "Findings",
                    "headers": ["Priority", "Category", "Tab", "Item", "Description", "Recommendation"],
                    "rows": findings_rows
                },
                {
                    "name": "Correlations",
                    "headers": ["Correlation", "Rule Name", "Risk", "Description", "Rule Logic", "Event Count", "Event Types"],
                    "rows": correlation_rows
                }
            ]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                fp_flag = self._compute_fp_flag(row[11], row[10], row[1], row[2], targetFps)
                if filter_fps and fp_flag == 1:
                    continue
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[10]), use_legacy=use_legacy)
                rows.append([row[0], event_type, str(row[3]),
                            str(row[2]), fp_flag, datafield])

            _scan_name = scanInfo[0] if scanInfo else 'Search'
            if export_mode == "analysis_correlations":
                fname = self._export_filename(_scan_name, id, 'ANALYSIS-CORRELATIONS', 'xlsx')
            elif export_mode == "analysis":
                fname = self._export_filename(_scan_name, id, 'ANALYSIS', 'xlsx')
            else:
                fname = self._export_filename(_scan_name, id, 'DATA', 'xlsx')
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1,
                                   prepend_sheets=prepend_sheets)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(
                ["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                fp_flag = self._compute_fp_flag(row[11], row[10], row[1], row[2], targetFps)
                if filter_fps and fp_flag == 1:
                    continue
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[10]), use_legacy=use_legacy)
                parser.writerow([row[0], event_type, str(
                    row[3]), str(row[2]), fp_flag, datafield])

            _scan_name = scanInfo[0] if scanInfo else 'Search'
            if export_mode == "analysis":
                fname = self._export_filename(_scan_name, id, 'ANALYSIS', 'csv')
            else:
                fname = self._export_filename(_scan_name, id, 'DATA', 'csv')
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        _name = scan_name if scan_name and len(ids.split(',')) == 1 else 'Multi-Scan'
        _id = ids.split(',')[0] if ids else ''
        fname = self._export_filename(_name, _id, 'DATA', 'json')

        cherrypy.response.headers[
            'Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'SpiderFootWebUi', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data or JSON
        """
        # For JSON requests, always return valid JSON (as bytes for CherryPy)
        if gexf == "0":
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            try:
                if not id:
                    return json.dumps({'nodes': [], 'edges': []}).encode('utf-8')

                dbh = SpiderFootDb(self.config)
                data = dbh.scanResultEvent(id, filterFp=True)
                scan = dbh.scanInstanceGet(id)

                # DEBUG: Log data retrieval info
                self.log.info(f"scanviz: Retrieved {len(data) if data else 0} rows for scan {id}")
                if data and len(data) > 0:
                    # Check event_type distribution (row[11] is t.event_type from tbl_event_types)
                    type_counts = {}
                    for row in data:
                        if len(row) >= 12:
                            event_type = row[11]
                            type_counts[event_type] = type_counts.get(event_type, 0) + 1
                    self.log.info(f"scanviz: Event type distribution: {type_counts}")

                if not scan:
                    return json.dumps({'nodes': [], 'edges': []}).encode('utf-8')

                root = scan[1]
                return SpiderFootHelpers.buildGraphJson([root], data).encode('utf-8')
            except Exception as e:
                self.log.error(f"scanviz JSON error: {e}")
                return json.dumps({'nodes': [], 'edges': []}).encode('utf-8')

        # For GEXF requests
        try:
            if not id:
                return ""

            dbh = SpiderFootDb(self.config)
            data = dbh.scanResultEvent(id, filterFp=True)
            scan = dbh.scanInstanceGet(id)

            if not scan:
                return ""

            scan_name = scan[0]
            root = scan[1]
            fname = self._export_filename(scan_name, id, 'GRAPH', 'gexf')

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/gexf"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", data)
        except Exception as e:
            self.log.error(f"scanviz GEXF error: {e}")
            return ""

    @cherrypy.expose
    def scanvizmulti(self: 'SpiderFootWebUi', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = SpiderFootDb(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        _name = scan_name if scan_name and len(ids.split(',')) == 1 else 'Multi-Scan'
        _id = ids.split(',')[0] if ids else ''
        fname = self._export_filename(_name, _id, 'GRAPH', 'gexf')

        cherrypy.response.headers[
            'Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            dict: scan options for the specified scan
        """
        dbh = SpiderFootDb(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(
                        key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'SpiderFootWebUi', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(id)
        
        if not info:
            return self.error("Invalid scan ID.")

        scanname = info[0]
        scantarget = info[1]
        
        # Validate that we have a valid target
        if not scantarget:
            return self.error(f"Scan {id} has no target defined.")

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"Error loading config from scan: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if not targetType:
            # It must then be a name, as a re-run scan should always have a clean
            # target. Put quotes around the target value and try to determine the
            # target type again.
            targetType = SpiderFootHelpers.targetTypeFromString(
                f'"{scantarget}"')

        # Final validation - ensure we have a valid target type
        if not targetType:
            self.log.error(f"Cannot determine target type for scan rerun: '{scantarget}'")
            return self.error(f"Cannot determine target type for scan rerun. Target '{scantarget}' is not recognized as a valid SpiderFoot target.")

        if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = _spawn_ctx.Process(target=startSpiderFootScanner, args=(
                self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}", exc_info=True)
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Track the Process object for kill capability
        with self._scan_processes_lock:
            self._scan_processes[scanId] = p

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(
            f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("Invalid scan ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            
            # Validate that we have a valid target
            if not scantarget:
                return self.error(f"Scan {id} has no target defined.")
            
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "sfp__stor_stdout" in modlist:
                modlist.remove("sfp__stor_stdout")
                
            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                # Should never be triggered for a re-run scan..
                self.log.error(f"Invalid target type for scan {id}: '{scantarget}' could not be recognized")
                return self.error(f"Invalid target type for scan {id}. Could not recognize '{scantarget}' as a target SpiderFoot supports.")

            # Start running a new scan
            scanId = SpiderFootHelpers.genScanInstanceId()
            try:
                p = _spawn_ctx.Process(target=startSpiderFootScanner, args=(
                    self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(
                    f"[-] Scan [{scanId}] failed: {e}", exc_info=True)
                return self.error(f"[-] Scan [{scanId}] failed: {e}")

            # Track the Process object for kill capability
            with self._scan_processes_lock:
                self._scan_processes[scanId] = p

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("Waiting for the scan to initialize...")
                time.sleep(1)

        templ = Template(
            filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__,
                            user_role=self.currentUserRole())

    @cherrypy.expose
    def newscan(self: 'SpiderFootWebUi') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        templ = Template(
            filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", scantarget="", version=__version__,
                            user_role=self.currentUserRole())

    @cherrypy.expose
    def importscans(self: 'SpiderFootWebUi') -> str:
        """Show the import data page.

        Returns:
            str: Import page HTML
        """
        self.requireAdmin()
        templ = Template(
            filename='spiderfoot/templates/import.tmpl', lookup=self.lookup)
        return templ.render(pageid='IMPORT', docroot=self.docroot,
                            version=__version__, user_role=self.currentUserRole())

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def processimport(self: 'SpiderFootWebUi', importfile=None,
                      import_type: str = 'legacy', scan_name: str = None,
                      target: str = None, dry_run: str = None) -> dict:
        """Process an uploaded file for import.

        Args:
            importfile: uploaded file (CherryPy file upload Part)
            import_type (str): type of import (legacy, scan, burp, nessus)
            scan_name (str): optional name for the imported scan
            target (str): optional target for the scan
            dry_run (str): if '1', validate only without importing

        Returns:
            dict: import results
        """
        if self.currentUserRole() != 'admin':
            return {'success': False, 'message': 'Unauthorized'}
        if importfile is None:
            return {'success': False, 'message': 'No file was uploaded.'}

        is_dry_run = dry_run == '1'

        # Read the uploaded file content
        try:
            raw = importfile.file.read()
            content = raw.decode('utf-8', errors='replace')
        except Exception as e:
            return {'success': False, 'message': f'Failed to read uploaded file: {e}'}

        if not content.strip():
            return {'success': False, 'message': 'Uploaded file is empty.'}

        # Handle Nessus import
        if import_type == 'nessus':
            return self._processNessusImport(content, scan_name, target, importfile, is_dry_run)

        # Handle Burp XML import
        if import_type == 'burp':
            return self._processBurpImport(content, scan_name, target, importfile, is_dry_run)

        # Handle Burp HTML Enhance (requires existing scan with Burp XML data)
        if import_type == 'burp_html':
            return self._processBurpHtmlEnhance(content, scan_name, target, importfile, is_dry_run)

        # Parse CSV for legacy/scan imports
        try:
            reader = csv.reader(StringIO(content))
            headers = next(reader)
        except Exception as e:
            return {'success': False, 'message': f'Failed to parse CSV headers: {e}'}

        # Detect format
        headers_lower = [h.lower().strip() for h in headers]

        fp_index = None
        for name in ['f/p', 'fp', 'status', 'validated']:
            if name in headers_lower:
                fp_index = headers_lower.index(name)
                break

        try:
            col_map = {}
            if 'scan name' in headers_lower:
                col_map['format'] = 'multi'
                col_map['scan_name'] = headers_lower.index('scan name')
            else:
                col_map['format'] = 'single'
            col_map['updated'] = headers_lower.index('updated')
            col_map['type'] = headers_lower.index('type')
            col_map['module'] = headers_lower.index('module')
            col_map['source'] = headers_lower.index('source')
            col_map['fp'] = fp_index
            col_map['data'] = headers_lower.index('data')
        except ValueError:
            return {
                'success': False,
                'message': f'CSV format not recognized. Expected columns: Updated, Type, Module, Source, Data. Found: {", ".join(headers)}'
            }

        rows = list(reader)
        if not rows:
            return {'success': False, 'message': 'CSV file contains no data rows.'}

        # Determine target
        if not target:
            first_row = rows[0]
            target = first_row[col_map['source']] if first_row[col_map['source']] else 'imported_target'

        # Default scan name
        if not scan_name:
            filename = getattr(importfile, 'filename', 'import')
            scan_name = f'Imported: {filename.rsplit(".", 1)[0] if "." in filename else filename}'

        stats = {
            'rows_read': len(rows),
            'rows_imported': 0,
            'rows_skipped': 0,
            'fps_imported': 0,
            'validated_imported': 0,
            'errors': [],
            'scan_id': None,
            'event_types': set(),
        }

        if is_dry_run:
            # Just validate and count
            for row in rows:
                try:
                    event_type = row[col_map['type']]
                    stats['event_types'].add(event_type)
                    if event_type == 'ROOT':
                        stats['rows_skipped'] += 1
                    else:
                        stats['rows_imported'] += 1
                except (IndexError, KeyError):
                    stats['rows_skipped'] += 1

            return {
                'success': True,
                'dry_run': True,
                'message': f'Validation passed. {stats["rows_imported"]} rows ready to import.',
                'rows_read': stats['rows_read'],
                'rows_imported': stats['rows_imported'],
                'rows_skipped': stats['rows_skipped'],
                'fps_imported': 0,
                'validated_imported': 0,
                'event_types_count': len(stats['event_types']),
                'errors': stats['errors'],
            }

        # Actual import
        dbh = SpiderFootDb(self.config)
        scan_id = str(uuid.uuid4())
        stats['scan_id'] = scan_id

        try:
            dbh.scanInstanceCreate(scan_id, scan_name, target)
        except Exception as e:
            return {'success': False, 'message': f'Failed to create scan instance: {e}'}

        # Create ROOT event for web UI browse compatibility
        try:
            root_qry = """INSERT INTO tbl_scan_results
                (scan_instance_id, hash, type, generated, confidence,
                visibility, risk, module, data, false_positive, source_event_hash)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
            root_qvals = [scan_id, 'ROOT', 'ROOT', int(time.time() * 1000),
                          100, 100, 0, '', target, 0, 'ROOT']
            dbh.dbh.execute(root_qry, root_qvals)
        except Exception as e:
            return {'success': False, 'message': f'Failed to create ROOT event: {e}'}

        # Build synthetic source events so that each imported event
        # points to the correct source data element instead of ROOT.
        source_hash_map = {}
        for row in rows:
            try:
                event_type = row[col_map['type']]
                if event_type == 'ROOT':
                    continue
                source_val = row[col_map['source']]
                if source_val and source_val not in source_hash_map:
                    src_hash_input = f"{scan_id}|SOURCE_EVENT|{source_val}"
                    src_hash = hashlib.sha256(src_hash_input.encode('utf-8')).hexdigest()[:32]
                    source_hash_map[source_val] = src_hash
            except (IndexError, KeyError):
                continue

        # Insert synthetic source events (children of ROOT)
        src_qry = """INSERT INTO tbl_scan_results
            (scan_instance_id, hash, type, generated, confidence,
            visibility, risk, module, data, false_positive, source_event_hash)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
        for source_val, src_hash in source_hash_map.items():
            try:
                src_qvals = [scan_id, src_hash, 'ROOT', int(time.time() * 1000),
                             100, 100, 0, '', source_val, 0, 'ROOT']
                dbh.dbh.execute(src_qry, src_qvals)
            except Exception:
                pass

        # Import each row
        for i, row in enumerate(rows):
            try:
                event_type = row[col_map['type']]
                module = row[col_map['module']]
                source = row[col_map['source']]
                data = row[col_map['data']]

                # Parse timestamp
                timestamp_str = row[col_map['updated']]
                timestamp = int(time.time() * 1000)
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%m/%d/%Y %H:%M:%S"]:
                    try:
                        dt = dt_datetime.strptime(timestamp_str.strip(), fmt)
                        timestamp = int(dt.timestamp() * 1000)
                        break
                    except ValueError:
                        continue

                # Parse status flag
                fp = 0
                if col_map['fp'] is not None and len(row) > col_map['fp']:
                    fp_val = row[col_map['fp']].strip().lower() if row[col_map['fp']] else ''
                    if fp_val in ('1', 'true', 'yes', 'fp', 'false positive'):
                        fp = 1
                    elif fp_val in ('2', 'validated', 'valid', 'confirmed'):
                        fp = 2
                    elif fp_val and fp_val.isdigit():
                        fp = int(fp_val) if int(fp_val) in (0, 1, 2) else 0

                # Skip ROOT events
                if event_type == 'ROOT':
                    stats['rows_skipped'] += 1
                    continue

                # Generate hash
                hash_input = f"{scan_id}|{event_type}|{data}|{source}|{time.time()}"
                event_hash = hashlib.sha256(hash_input.encode('utf-8')).hexdigest()[:32]

                # Resolve source event hash from synthetic source events
                source_event_hash = source_hash_map.get(source, 'ROOT')

                qry = """INSERT INTO tbl_scan_results
                    (scan_instance_id, hash, type, generated, confidence,
                    visibility, risk, module, data, false_positive, source_event_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
                qvals = [scan_id, event_hash, event_type, timestamp,
                         100, 100, 0, module, data, fp, source_event_hash]
                dbh.dbh.execute(qry, qvals)

                # Save target-level false positives
                if fp == 1:
                    try:
                        fp_qry = """INSERT OR IGNORE INTO tbl_target_false_positives
                            (target, event_type, event_data, source_data, date_added, notes)
                            VALUES (?, ?, ?, ?, ?, ?)"""
                        fp_qvals = [target, event_type, data, source, int(time.time() * 1000),
                                    f'Imported via web UI: {scan_name}']
                        dbh.dbh.execute(fp_qry, fp_qvals)
                        stats['fps_imported'] += 1
                    except Exception:
                        pass

                # Save target-level validated entries
                if fp == 2:
                    try:
                        val_qry = """INSERT OR IGNORE INTO tbl_target_validated
                            (target, event_type, event_data, source_data, date_added, notes)
                            VALUES (?, ?, ?, ?, ?, ?)"""
                        val_qvals = [target, event_type, data, source, int(time.time() * 1000),
                                     f'Imported via web UI: {scan_name}']
                        dbh.dbh.execute(val_qry, val_qvals)
                        stats['validated_imported'] += 1
                    except Exception:
                        pass

                stats['event_types'].add(event_type)
                stats['rows_imported'] += 1

            except Exception as e:
                stats['errors'].append(f'Row {i + 1}: {str(e)}')
                stats['rows_skipped'] += 1

        # Commit and finalize
        dbh.conn.commit()
        dbh.scanInstanceSet(scan_id, status='FINISHED', ended=time.time() * 1000)

        return {
            'success': True,
            'dry_run': False,
            'message': f'Successfully imported {stats["rows_imported"]} rows into scan "{scan_name}".',
            'scan_id': scan_id,
            'rows_read': stats['rows_read'],
            'rows_imported': stats['rows_imported'],
            'rows_skipped': stats['rows_skipped'],
            'fps_imported': stats['fps_imported'],
            'validated_imported': stats['validated_imported'],
            'event_types_count': len(stats['event_types']),
            'errors': stats['errors'],
        }

    def _processNessusImport(self, content: str, scan_name: str, target: str,
                              importfile=None, is_dry_run: bool = False,
                              existing_scan_id: str = None) -> dict:
        """Parse and import a Nessus .nessus XML file.

        Args:
            content (str): XML file content
            scan_name (str): optional scan name
            target (str): optional target
            importfile: uploaded file object (for filename)
            is_dry_run (bool): if True, validate only
            existing_scan_id (str): if set, import into existing scan

        Returns:
            dict: import results
        """
        try:
            root = ET.fromstring(content)
        except ET.ParseError as e:
            return {'success': False, 'message': f'Failed to parse Nessus XML: {e}'}

        # Validate root element
        if root.tag != 'NessusClientData_v2':
            return {'success': False, 'message': f'Invalid Nessus file. Expected NessusClientData_v2 root element, got: {root.tag}'}

        results = []
        hosts_seen = set()

        for report in root.findall('.//Report'):
            for report_host in report.findall('ReportHost'):
                host_name = report_host.get('name', '')
                hosts_seen.add(host_name)

                # Extract host properties
                host_props = {}
                host_properties_el = report_host.find('HostProperties')
                if host_properties_el is not None:
                    for tag in host_properties_el.findall('tag'):
                        tag_name = tag.get('name', '')
                        tag_val = tag.text or ''
                        host_props[tag_name] = tag_val

                host_ip = host_props.get('host-ip', host_name)
                host_fqdn = host_props.get('host-fqdn', host_name)
                operating_system = host_props.get('operating-system', '')

                for item in report_host.findall('ReportItem'):
                    severity_num = int(item.get('severity', '0'))
                    severity_map = {0: 'None', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Critical'}
                    severity = severity_map.get(severity_num, 'None')

                    result = {
                        'severity': severity,
                        'severity_number': severity_num,
                        'plugin_name': item.get('pluginName', ''),
                        'plugin_id': item.get('pluginID', ''),
                        'host_ip': host_ip,
                        'host_name': host_fqdn,
                        'operating_system': operating_system,
                        'description': (item.findtext('description') or '').strip(),
                        'synopsis': (item.findtext('synopsis') or '').strip(),
                        'solution': (item.findtext('solution') or '').strip(),
                        'see_also': (item.findtext('see_also') or '').strip(),
                        'service_name': item.get('svc_name', ''),
                        'port': int(item.get('port', '0')),
                        'protocol': item.get('protocol', ''),
                        'request': '',
                        'plugin_output': (item.findtext('plugin_output') or '').strip(),
                        'cvss3_base_score': item.findtext('cvss3_base_score') or '',
                    }
                    results.append(result)

        if not results:
            return {'success': False, 'message': 'No vulnerability items found in Nessus file.'}

        # Auto-detect target
        if not target:
            target = list(hosts_seen)[0] if hosts_seen else 'nessus_import'

        # Default scan name
        if not scan_name:
            filename = getattr(importfile, 'filename', 'nessus_import')
            scan_name = f'Nessus Import: {filename.rsplit(".", 1)[0] if "." in filename else filename}'

        if is_dry_run:
            severity_counts = {}
            for r in results:
                sev = r['severity']
                severity_counts[sev] = severity_counts.get(sev, 0) + 1

            return {
                'success': True,
                'dry_run': True,
                'message': f'Validation passed. {len(results)} vulnerability items found across {len(hosts_seen)} host(s).',
                'rows_read': len(results),
                'rows_imported': len(results),
                'rows_skipped': 0,
                'event_types_count': len(severity_counts),
                'errors': [],
            }

        # Actual import
        dbh = SpiderFootDb(self.config)
        scan_id = existing_scan_id

        if not scan_id:
            scan_id = str(uuid.uuid4())
            try:
                dbh.scanInstanceCreate(scan_id, scan_name, target)
                # Create ROOT event
                root_qry = """INSERT INTO tbl_scan_results
                    (scan_instance_id, hash, type, generated, confidence,
                    visibility, risk, module, data, false_positive, source_event_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
                root_qvals = [scan_id, 'ROOT', 'ROOT', int(time.time() * 1000),
                              100, 100, 0, '', target, 0, 'ROOT']
                dbh.dbh.execute(root_qry, root_qvals)
                dbh.conn.commit()
            except Exception as e:
                return {'success': False, 'message': f'Failed to create scan instance: {e}'}

        # Preserve tracking status for existing TICKETED/CLOSED findings on reimport
        trackedFindings = None
        if existing_scan_id:
            try:
                trackedFindings = dbh.scanNessusTrackedFindings(scan_id)
            except Exception:
                pass

        try:
            count = dbh.scanNessusStore(scan_id, results, trackedFindings=trackedFindings)
        except Exception as e:
            return {'success': False, 'message': f'Failed to store Nessus results: {e}'}

        if not existing_scan_id:
            dbh.scanInstanceSet(scan_id, status='FINISHED', ended=time.time() * 1000)

        return {
            'success': True,
            'dry_run': False,
            'message': f'Successfully imported {count} Nessus findings into scan "{scan_name}".',
            'scan_id': scan_id,
            'rows_read': len(results),
            'rows_imported': count,
            'rows_skipped': 0,
            'fps_imported': 0,
            'validated_imported': 0,
            'event_types_count': len(set(r['severity'] for r in results)),
            'errors': [],
        }

    def _processBurpImport(self, content: str, scan_name: str, target: str,
                            importfile=None, is_dry_run: bool = False,
                            existing_scan_id: str = None) -> dict:
        """Parse and import a Burp Suite XML file.

        Args:
            content (str): XML file content
            scan_name (str): optional scan name
            target (str): optional target
            importfile: uploaded file object (for filename)
            is_dry_run (bool): if True, validate only
            existing_scan_id (str): if set, import into existing scan

        Returns:
            dict: import results
        """
        try:
            root = ET.fromstring(content)
        except ET.ParseError as e:
            return {'success': False, 'message': f'Failed to parse Burp XML: {e}'}

        # Validate root element
        if root.tag != 'issues':
            return {'success': False, 'message': f'Invalid Burp file. Expected "issues" root element, got: {root.tag}'}

        results = []
        hosts_seen = set()

        for issue in root.findall('issue'):
            host_el = issue.find('host')
            host_ip = ''
            host_name = ''
            if host_el is not None:
                host_ip = host_el.get('ip', '')
                host_name = (host_el.text or '').strip()

            if host_name:
                hosts_seen.add(host_name)
            elif host_ip:
                hosts_seen.add(host_ip)

            severity_text = (issue.findtext('severity') or '').strip()
            severity_map = {'High': 3, 'Medium': 2, 'Low': 1, 'Information': 0, 'Info': 0}
            severity_number = severity_map.get(severity_text, 0)

            # Extract request/response, handling base64 encoding
            request_text = ''
            response_text = ''
            rr = issue.find('requestresponse')
            if rr is not None:
                req_el = rr.find('request')
                if req_el is not None:
                    if req_el.get('base64') == 'true' and req_el.text:
                        try:
                            request_text = base64.b64decode(req_el.text).decode('utf-8', errors='replace')
                        except Exception:
                            request_text = req_el.text or ''
                    else:
                        request_text = (req_el.text or '').strip()

                resp_el = rr.find('response')
                if resp_el is not None:
                    if resp_el.get('base64') == 'true' and resp_el.text:
                        try:
                            response_text = base64.b64decode(resp_el.text).decode('utf-8', errors='replace')
                        except Exception:
                            response_text = resp_el.text or ''
                    else:
                        response_text = (resp_el.text or '').strip()

            # Combine remediation fields for solutions
            remediation_bg = (issue.findtext('remediationBackground') or '').strip()
            remediation_detail = (issue.findtext('remediationDetail') or '').strip()
            solutions = remediation_bg
            if remediation_detail:
                solutions = (solutions + '\n\n' + remediation_detail).strip() if solutions else remediation_detail

            # Extract references and vulnerability classifications
            references = (issue.findtext('references') or '').strip()
            vuln_classifications = (issue.findtext('vulnerabilityClassifications') or '').strip()

            result = {
                'severity': severity_text,
                'severity_number': severity_number,
                'host_ip': host_ip,
                'host_name': host_name,
                'plugin_name': (issue.findtext('name') or '').strip(),
                'issue_type': (issue.findtext('type') or '').strip(),
                'path': (issue.findtext('path') or '').strip(),
                'location': (issue.findtext('location') or '').strip(),
                'confidence': (issue.findtext('confidence') or '').strip(),
                'issue_background': (issue.findtext('issueBackground') or '').strip(),
                'issue_detail': (issue.findtext('issueDetail') or '').strip(),
                'solutions': solutions,
                'see_also': '',
                'references': references,
                'vulnerability_classifications': vuln_classifications,
                'request': request_text,
                'response': response_text,
            }
            results.append(result)

        if not results:
            return {'success': False, 'message': 'No issues found in Burp XML file.'}

        # Auto-detect target
        if not target:
            target = list(hosts_seen)[0] if hosts_seen else 'burp_import'

        # Default scan name
        if not scan_name:
            filename = getattr(importfile, 'filename', 'burp_import')
            scan_name = f'Burp Import: {filename.rsplit(".", 1)[0] if "." in filename else filename}'

        if is_dry_run:
            severity_counts = {}
            for r in results:
                sev = r['severity']
                severity_counts[sev] = severity_counts.get(sev, 0) + 1

            return {
                'success': True,
                'dry_run': True,
                'message': f'Validation passed. {len(results)} issues found across {len(hosts_seen)} host(s).',
                'rows_read': len(results),
                'rows_imported': len(results),
                'rows_skipped': 0,
                'event_types_count': len(severity_counts),
                'errors': [],
            }

        # Actual import
        dbh = SpiderFootDb(self.config)
        scan_id = existing_scan_id

        if not scan_id:
            scan_id = str(uuid.uuid4())
            try:
                dbh.scanInstanceCreate(scan_id, scan_name, target)
                # Create ROOT event
                root_qry = """INSERT INTO tbl_scan_results
                    (scan_instance_id, hash, type, generated, confidence,
                    visibility, risk, module, data, false_positive, source_event_hash)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""
                root_qvals = [scan_id, 'ROOT', 'ROOT', int(time.time() * 1000),
                              100, 100, 0, '', target, 0, 'ROOT']
                dbh.dbh.execute(root_qry, root_qvals)
                dbh.conn.commit()
            except Exception as e:
                return {'success': False, 'message': f'Failed to create scan instance: {e}'}

        # Preserve tracking status for existing TICKETED/CLOSED findings on reimport
        trackedFindings = None
        if existing_scan_id:
            try:
                trackedFindings = dbh.scanBurpTrackedFindings(scan_id)
            except Exception:
                pass

        try:
            count = dbh.scanBurpStore(scan_id, results, trackedFindings=trackedFindings)
        except Exception as e:
            return {'success': False, 'message': f'Failed to store Burp results: {e}'}

        if not existing_scan_id:
            dbh.scanInstanceSet(scan_id, status='FINISHED', ended=time.time() * 1000)

        return {
            'success': True,
            'dry_run': False,
            'message': f'Successfully imported {count} Burp issues into scan "{scan_name}".',
            'scan_id': scan_id,
            'rows_read': len(results),
            'rows_imported': count,
            'rows_skipped': 0,
            'fps_imported': 0,
            'validated_imported': 0,
            'event_types_count': len(set(r['severity'] for r in results)),
            'errors': [],
        }

    def _processBurpHtmlEnhance(self, content: str, scan_name: str, target: str,
                                importfile=None, is_dry_run: bool = False,
                                existing_scan_id: str = None) -> dict:
        """Parse a Burp Suite Pro HTML report and enhance existing Burp XML results.

        Parses the classic Burp Pro HTML report format and matches findings
        against existing Burp XML results by plugin_name. For each match,
        empty fields in the existing record are filled with data from the
        HTML report (issue_detail, issue_background, solutions, references,
        vulnerability_classifications, host_ip, host_name, request, response).

        Structure:
          BODH0 = issue type group (e.g. "1. TLS cookie without secure flag set")
            Group-level: Issue background, Issue remediation, Vulnerability
                         classifications, References
          BODH1 = individual instance (e.g. "1.1. https://host.com/")
            Instance-level: Summary table, Issue detail, Request, Response

        Args:
            content (str): HTML file content
            scan_name (str): optional scan name (unused for enhance)
            target (str): optional target (unused for enhance)
            importfile: uploaded file object (for filename)
            is_dry_run (bool): if True, validate only
            existing_scan_id (str): scan ID to enhance (required)

        Returns:
            dict: enhance results
        """
        try:
            from bs4 import BeautifulSoup, Tag
        except ImportError:
            return {'success': False, 'message': 'BeautifulSoup4 is required for HTML import. Install with: pip install beautifulsoup4'}

        import re as _re
        from urllib.parse import urlparse

        try:
            soup = BeautifulSoup(content, 'html.parser')
        except Exception as e:
            return {'success': False, 'message': f'Failed to parse HTML: {e}'}

        # Validate it looks like a Burp report
        bodh0_elements = soup.find_all('span', class_='BODH0')
        if not bodh0_elements:
            return {'success': False, 'message': 'Invalid Burp HTML report. No issue groups (BODH0) found. Ensure this is a Burp Suite Pro HTML export.'}

        results = []
        hosts_seen = set()

        def _extract_host_parts(host_val):
            """Extract hostname and derive IP-like field from a Host value.

            Burp HTML uses full URLs like 'https://admin.example.com' for Host.
            Returns (host_name, host_ip) where host_name is the full URL and
            host_ip is the extracted domain/hostname.
            """
            host_val = (host_val or '').strip()
            if not host_val:
                return '', ''
            # Parse as URL to extract hostname
            if '://' in host_val:
                parsed = urlparse(host_val)
                hostname = parsed.hostname or ''
                return host_val, hostname
            return host_val, host_val

        def _get_element_classes(el):
            """Safely get class list from an element."""
            if not isinstance(el, Tag):
                return []
            return el.get('class') or []

        # Build a flat ordered list of significant elements for sequential traversal.
        # This is more robust than find_next_sibling() which breaks with nested structures.
        body = soup.body or soup
        significant_elements = []
        for el in body.descendants:
            if not isinstance(el, Tag):
                continue
            classes = el.get('class') or []
            if el.name == 'span' and ('BODH0' in classes or 'BODH1' in classes or 'TEXT' in classes):
                significant_elements.append(el)
            elif el.name == 'h2':
                significant_elements.append(el)
            elif el.name == 'table' and 'summary_table' in classes:
                significant_elements.append(el)
            elif el.name == 'div' and 'rr_div' in classes:
                significant_elements.append(el)

        # State machine: iterate through significant elements in document order
        current_group_name = ''
        group_issue_background = ''
        group_remediation = ''
        group_references = ''
        group_classifications = ''
        current_h2 = ''
        in_instance = False
        current_instance = None
        instances_in_group = []

        def _new_instance(group_name):
            return {
                'plugin_name': group_name,
                'severity': '',
                'severity_number': 0,
                'confidence': '',
                'host_ip': '',
                'host_name': '',
                'path': '',
                'location': '',
                'issue_type': '',
                'issue_background': '',
                'issue_detail': '',
                'solutions': '',
                'see_also': '',
                'references': '',
                'vulnerability_classifications': '',
                'request': '',
                'response': '',
            }

        def _finalize_group():
            """Apply group-level shared data to instances and add to results."""
            nonlocal instances_in_group
            for inst in instances_in_group:
                # Apply group-level data to instances that don't have their own
                if not inst['issue_background'] and group_issue_background:
                    inst['issue_background'] = group_issue_background
                if not inst['solutions'] and group_remediation:
                    inst['solutions'] = group_remediation
                if not inst['references'] and group_references:
                    inst['references'] = group_references
                if not inst['vulnerability_classifications'] and group_classifications:
                    inst['vulnerability_classifications'] = group_classifications
                # If issue_detail is empty, use issue_background as fallback
                if not inst['issue_detail'] and inst['issue_background']:
                    inst['issue_detail'] = inst['issue_background']
            results.extend(instances_in_group)
            instances_in_group = []

        for el in significant_elements:
            classes = _get_element_classes(el)

            # --- BODH0: New issue type group ---
            if el.name == 'span' and 'BODH0' in classes:
                # Finalize previous instance and group
                if current_instance:
                    instances_in_group.append(current_instance)
                    current_instance = None
                _finalize_group()

                # Start new group
                raw_name = el.get_text(strip=True)
                current_group_name = _re.sub(r'^\d+\.\s*', '', raw_name)
                group_issue_background = ''
                group_remediation = ''
                group_references = ''
                group_classifications = ''
                current_h2 = ''
                in_instance = False

            # --- BODH1: New individual instance ---
            elif el.name == 'span' and 'BODH1' in classes:
                # Save previous instance
                if current_instance:
                    instances_in_group.append(current_instance)
                current_instance = _new_instance(current_group_name)
                current_h2 = ''
                in_instance = True

            # --- H2: Section header ---
            elif el.name == 'h2':
                current_h2 = el.get_text(strip=True).lower()

            # --- Summary table: severity, confidence, host, path ---
            elif el.name == 'table' and 'summary_table' in classes:
                if current_instance:
                    rows = el.find_all('tr')
                    for row in rows:
                        tds = row.find_all('td')
                        for td in tds:
                            # Skip icon cells
                            if 'icon' in (td.get('class') or []):
                                continue
                            text = td.get_text(strip=True)
                            # Get the value - could be in <b>, <a>, or just text
                            b_tag = td.find('b')
                            a_tag = td.find('a')
                            if b_tag:
                                val = b_tag.get_text(strip=True)
                            elif a_tag:
                                val = a_tag.get_text(strip=True)
                            else:
                                # Try to extract value after the label
                                val = text

                            if text.startswith('Severity:'):
                                sev_val = val.replace('Severity:', '').strip() if val == text else val
                                current_instance['severity'] = sev_val
                                sev_map = {'High': 3, 'Medium': 2, 'Low': 1, 'Information': 0, 'Info': 0}
                                current_instance['severity_number'] = sev_map.get(sev_val, 0)
                            elif text.startswith('Confidence:'):
                                conf_val = val.replace('Confidence:', '').strip() if val == text else val
                                current_instance['confidence'] = conf_val
                            elif text.startswith('Host:'):
                                host_val = val.replace('Host:', '').strip() if val == text else val
                                # Also check for <a> href which may have the full URL
                                if a_tag and a_tag.get('href'):
                                    host_val = a_tag.get('href').strip() or host_val
                                host_name, host_ip = _extract_host_parts(host_val)
                                current_instance['host_name'] = host_name
                                current_instance['host_ip'] = host_ip
                                if host_ip:
                                    hosts_seen.add(host_ip)
                            elif text.startswith('Path:'):
                                path_val = val.replace('Path:', '').strip() if val == text else val
                                current_instance['path'] = path_val

                    # Try to extract severity from icon class if not found in text
                    if not current_instance['severity']:
                        for icon_td in el.find_all('td', class_='icon'):
                            div = icon_td.find('div')
                            if div:
                                cls = ' '.join(div.get('class') or [])
                                if 'high' in cls:
                                    current_instance['severity'] = 'High'
                                    current_instance['severity_number'] = 3
                                elif 'medium' in cls:
                                    current_instance['severity'] = 'Medium'
                                    current_instance['severity_number'] = 2
                                elif 'low' in cls:
                                    current_instance['severity'] = 'Low'
                                    current_instance['severity_number'] = 1
                                elif 'info' in cls:
                                    current_instance['severity'] = 'Information'
                                    current_instance['severity_number'] = 0

            # --- TEXT spans: section body content ---
            elif el.name == 'span' and 'TEXT' in classes:
                text_content = el.get_text(separator='\n', strip=True)

                # Extract links for reference-type sections
                links = el.find_all('a')
                link_texts = [a.get_text(strip=True) for a in links if a.get_text(strip=True)]

                if in_instance and current_instance:
                    # Instance-level sections
                    if 'issue detail' in current_h2:
                        current_instance['issue_detail'] = (current_instance['issue_detail'] + '\n\n' + text_content).strip() if current_instance['issue_detail'] else text_content
                    elif 'issue background' in current_h2:
                        current_instance['issue_background'] = (current_instance['issue_background'] + '\n\n' + text_content).strip() if current_instance['issue_background'] else text_content
                    elif 'remediation' in current_h2:
                        current_instance['solutions'] = (current_instance['solutions'] + '\n\n' + text_content).strip() if current_instance['solutions'] else text_content
                    elif 'vulnerability classif' in current_h2:
                        val = '\n'.join(link_texts) if link_texts else text_content
                        current_instance['vulnerability_classifications'] = (current_instance['vulnerability_classifications'] + '\n' + val).strip() if current_instance['vulnerability_classifications'] else val
                    elif 'reference' in current_h2:
                        val = '\n'.join(link_texts) if link_texts else text_content
                        current_instance['references'] = (current_instance['references'] + '\n' + val).strip() if current_instance['references'] else val
                else:
                    # Group-level shared sections (before any BODH1)
                    if 'issue background' in current_h2:
                        group_issue_background = (group_issue_background + '\n\n' + text_content).strip() if group_issue_background else text_content
                    elif 'remediation' in current_h2:
                        group_remediation = (group_remediation + '\n\n' + text_content).strip() if group_remediation else text_content
                    elif 'vulnerability classif' in current_h2:
                        val = '\n'.join(link_texts) if link_texts else text_content
                        group_classifications = (group_classifications + '\n' + val).strip() if group_classifications else val
                    elif 'reference' in current_h2:
                        val = '\n'.join(link_texts) if link_texts else text_content
                        group_references = (group_references + '\n' + val).strip() if group_references else val

            # --- Request/Response divs ---
            elif el.name == 'div' and 'rr_div' in classes:
                if current_instance:
                    rr_text = el.get_text()
                    if 'request' in current_h2:
                        current_instance['request'] = (current_instance['request'] + '\n\n' + rr_text.strip()).strip() if current_instance['request'] else rr_text.strip()
                    elif 'response' in current_h2:
                        current_instance['response'] = (current_instance['response'] + '\n\n' + rr_text.strip()).strip() if current_instance['response'] else rr_text.strip()

        # Finalize last instance and group
        if current_instance:
            instances_in_group.append(current_instance)
        _finalize_group()

        if not results:
            return {'success': False, 'message': 'No individual issue instances found in the Burp HTML report.'}

        # Auto-detect target
        if not target:
            target = list(hosts_seen)[0] if hosts_seen else 'burp_html_import'

        # Default scan name
        if not scan_name:
            filename = getattr(importfile, 'filename', 'burp_html_import')
            scan_name = f'Burp HTML Import: {filename.rsplit(".", 1)[0] if "." in filename else filename}'

        if not existing_scan_id:
            return {'success': False, 'message': 'HTML ENHANCE requires an existing scan with Burp XML data. Import Burp XML first, then use HTML ENHANCE to add details.'}

        if is_dry_run:
            severity_counts = {}
            for r in results:
                sev = r['severity'] or 'Unknown'
                severity_counts[sev] = severity_counts.get(sev, 0) + 1

            # Check how many existing records could be matched
            dbh = SpiderFootDb(self.config)
            existing_count = dbh.scanBurpCount(existing_scan_id)
            html_names = set(r['plugin_name'] for r in results if r['plugin_name'])

            return {
                'success': True,
                'dry_run': True,
                'message': f'Validation passed. {len(results)} issues parsed from HTML across {len(hosts_seen)} host(s). {existing_count} existing Burp records will be checked for enhancement.',
                'rows_read': len(results),
                'rows_imported': len(results),
                'rows_skipped': 0,
                'event_types_count': len(html_names),
                'errors': [],
            }

        # Actual enhance
        dbh = SpiderFootDb(self.config)

        try:
            stats = dbh.scanBurpEnhance(existing_scan_id, results)
        except Exception as e:
            return {'success': False, 'message': f'Failed to enhance Burp results: {e}'}

        return {
            'success': True,
            'dry_run': False,
            'message': f'Enhanced {stats["enhanced"]} existing Burp records with HTML data. {stats["skipped"]} HTML issues had no matching XML record.',
            'scan_id': existing_scan_id,
            'rows_read': len(results),
            'rows_imported': stats['enhanced'],
            'rows_skipped': stats['skipped'],
            'fps_imported': 0,
            'validated_imported': 0,
            'event_types_count': len(set(r['plugin_name'] for r in results if r['plugin_name'])),
            'errors': [],
        }

    @cherrypy.expose
    def clonescan(self: 'SpiderFootWebUi', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)
        
        if not info:
            return self.error("Invalid scan ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        
        # Validate that we have a valid target
        if not scantarget:
            return self.error(f"Scan {id} has no target defined.")
        
        targetType = None
        
        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"
            # Re-check target type after wrapping
            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                self.log.error(f"Invalid target type for scan {id}: '{scantarget}' could not be recognized")
                return self.error(f"Invalid target type for scan {id}. Could not recognize '{scantarget}' as a target SpiderFoot supports.")

        modlist = scanconfig['_modulesenabled'].split(',')

        templ = Template(
            filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            scanname=str(scanname),
                            scantarget=str(scantarget), version=__version__,
                            user_role=self.currentUserRole())

    @cherrypy.expose
    def index(self: 'SpiderFootWebUi') -> str:
        """Show scan list page.

        Returns:
            str: Scan list page HTML
        """
        templ = Template(
            filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__,
                            user_role=self.currentUserRole())

    @cherrypy.expose
    def scaninfo(self: 'SpiderFootWebUi', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        dbh = SpiderFootDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = Template(filename='spiderfoot/templates/scaninfo.tmpl',
                         lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST", current_user=self.currentUser(), seedtarget=res[1],
                            user_role=self.currentUserRole())

    @cherrypy.expose
    def opts(self: 'SpiderFootWebUi', updated: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated successfully

        Returns:
            str: scan options page HTML
        """
        self.requireAdmin()
        templ = Template(
            filename='spiderfoot/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        current_user = self.currentUser()
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, docroot=self.docroot, current_user=current_user,
                            user_role=self.currentUserRole())

    @cherrypy.expose
    def users(self: 'SpiderFootWebUi') -> str:
        """Show user management page (admin only).

        Returns:
            str: User management page HTML or redirect
        """
        # Only admin can access user management
        self.requireAdmin()
        current_user = self.currentUser()

        dbh = SpiderFootDb(self.config)
        users_list = dbh.userList()

        templ = Template(
            filename='spiderfoot/templates/users.tmpl', lookup=self.lookup)
        return templ.render(
            pageid='USERS', docroot=self.docroot, version=__version__,
            users=users_list, current_user=current_user,
            user_role=self.currentUserRole())

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def userlist(self: 'SpiderFootWebUi') -> list:
        """List all users (admin only).

        Returns:
            list: List of user dicts
        """
        # Only admin can list users
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}

        dbh = SpiderFootDb(self.config)
        users = dbh.userList()
        return {'success': True, 'users': users}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def usercreate(self: 'SpiderFootWebUi', username: str, password: str, display_name: str = None, role: str = 'analyst') -> dict:
        """Create a new user (admin only).

        Args:
            username (str): username
            password (str): password
            display_name (str): optional display name
            role (str): user role ('admin' or 'analyst')

        Returns:
            dict: Result
        """
        # Only admin can create users
        current_user = self.currentUser()
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}

        if not username or not password:
            return {'success': False, 'error': 'Username and password are required'}

        # Validate username format
        if not username.isalnum() or len(username) < 3:
            return {'success': False, 'error': 'Username must be at least 3 alphanumeric characters'}

        if len(password) < 8:
            return {'success': False, 'error': 'Password must be at least 8 characters'}

        if role not in ('admin', 'analyst'):
            role = 'analyst'

        dbh = SpiderFootDb(self.config)

        # Check if username already exists
        if dbh.userGet(username):
            return {'success': False, 'error': 'Username already exists'}

        if dbh.userCreate(username, password, display_name, role=role):
            dbh.auditLog(current_user, 'USER_CREATE', detail=f'Created user: {username} (role: {role})', ip_address=self.clientIP())
            return {'success': True, 'message': f'User {username} created successfully'}
        else:
            return {'success': False, 'error': 'Failed to create user'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def userupdate(self: 'SpiderFootWebUi', username: str, display_name: str = None, active: str = None) -> dict:
        """Update a user (admin only).

        Args:
            username (str): username
            display_name (str): optional new display name
            active (str): optional new active status ('true' or 'false')

        Returns:
            dict: Result
        """
        # Only admin can update users
        current_user = self.currentUser()
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}

        dbh = SpiderFootDb(self.config)

        # Check if user exists
        if not dbh.userGet(username):
            return {'success': False, 'error': 'User not found'}

        # Convert active string to bool if provided
        active_bool = None
        if active is not None:
            active_bool = active.lower() == 'true'

        if dbh.userUpdate(username, display_name=display_name, active=active_bool):
            dbh.auditLog(current_user, 'USER_UPDATE', detail=f'Updated user: {username}', ip_address=self.clientIP())
            return {'success': True, 'message': f'User {username} updated successfully'}
        else:
            return {'success': False, 'error': 'Failed to update user'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def userdelete(self: 'SpiderFootWebUi', username: str) -> dict:
        """Delete a user (admin only).

        Args:
            username (str): username to delete

        Returns:
            dict: Result
        """
        # Only admin can delete users
        current_user = self.currentUser()
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}

        if username == 'admin':
            return {'success': False, 'error': 'Cannot delete admin user'}

        dbh = SpiderFootDb(self.config)

        # Check if user exists
        if not dbh.userGet(username):
            return {'success': False, 'error': 'User not found'}

        if dbh.userDelete(username):
            dbh.auditLog(current_user, 'USER_DELETE', detail=f'Deleted user: {username}', ip_address=self.clientIP())
            return {'success': True, 'message': f'User {username} deleted successfully'}
        else:
            return {'success': False, 'error': 'Failed to delete user'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def userchangepassword(self: 'SpiderFootWebUi', username: str, new_password: str) -> dict:
        """Change a user's password (admin only).

        Args:
            username (str): username
            new_password (str): new password

        Returns:
            dict: Result
        """
        # Only admin can change other users' passwords
        current_user = self.currentUser()
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}

        if len(new_password) < 8:
            return {'success': False, 'error': 'Password must be at least 8 characters'}

        dbh = SpiderFootDb(self.config)

        # Check if user exists
        if not dbh.userGet(username):
            return {'success': False, 'error': 'User not found'}

        if dbh.userChangePassword(username, new_password):
            dbh.auditLog(current_user, 'USER_PASSWORD_CHANGE', detail=f'Changed password for user: {username}', ip_address=self.clientIP())
            return {'success': True, 'message': f'Password for {username} changed successfully'}
        else:
            return {'success': False, 'error': 'Failed to change password'}

    @cherrypy.expose
    def auditlog(self: 'SpiderFootWebUi', action: str = None, username: str = None) -> str:
        """Show the activity/audit log page.

        Args:
            action (str): optional action type filter
            username (str): optional username filter

        Returns:
            str: audit log page HTML
        """
        dbh = SpiderFootDb(self.config)
        logs = dbh.auditLogGet(limit=500, username=username, action=action)

        # Format timestamps for display
        for entry in logs:
            entry['time_str'] = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(entry['created'] / 1000))

        # Get unique usernames for filter dropdown
        users = sorted(set(entry['username'] for entry in logs))

        templ = Template(
            filename='spiderfoot/templates/auditlog.tmpl', lookup=self.lookup)
        return templ.render(
            pageid='AUDITLOG', docroot=self.docroot, version=__version__,
            logs=logs, users=users, user_role=self.currentUserRole())

    @cherrypy.expose
    def workspaces(self: 'SpiderFootWebUi') -> str:
        """Show workspace management page.

        Returns:
            str: Workspace management page HTML
        """
        templ = Template(
            filename='spiderfoot/templates/workspaces.tmpl', lookup=self.lookup)
        return templ.render(pageid='WORKSPACES', docroot=self.docroot, version=__version__,
                            user_role=self.currentUserRole())

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def getlaunchcode(self: 'SpiderFootWebUi') -> dict:
        """Get the current scan launch code (admin only).

        Returns:
            dict: launch code info
        """
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}
        dbh = SpiderFootDb(self.config)
        code = dbh.launchCodeGet()
        return {'success': True, 'launch_code': code}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def setlaunchcode(self: 'SpiderFootWebUi', code: str = '') -> dict:
        """Set the scan launch code (admin only).

        Args:
            code (str): launch code string, or empty to disable

        Returns:
            dict: result
        """
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}
        dbh = SpiderFootDb(self.config)
        if dbh.launchCodeSet(code):
            dbh.auditLog(self.currentUser(), 'LAUNCH_CODE_SET',
                         detail=f'Launch code {"set" if code else "cleared"}',
                         ip_address=self.clientIP())
            return {'success': True, 'message': 'Launch code updated'}
        return {'success': False, 'error': 'Failed to update launch code'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def getresourcetier(self: 'SpiderFootWebUi') -> dict:
        """Get the current resource tuning tier (admin only).

        Returns:
            dict: current tier info
        """
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}
        return {
            'success': True,
            'tier': self.config.get('_resource_tier', 'medium'),
        }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def tuneresources(self: 'SpiderFootWebUi', tier: str = '') -> dict:
        """Set the resource tuning tier (admin only).

        Args:
            tier (str): 'light', 'medium', or 'heavy'

        Returns:
            dict: result with restart_required flag
        """
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}
        if tier not in ('light', 'medium', 'heavy'):
            return {'success': False, 'error': 'Invalid tier. Must be light, medium, or heavy.'}

        from spiderfoot.resource_tiers import get_tier_config
        tier_config = get_tier_config(tier)
        old_tier_name = self.config.get('_resource_tier', 'medium')
        old_tier_config = get_tier_config(old_tier_name)

        dbh = SpiderFootDb(self.config)
        dbh.configSet({
            '_resource_tier': tier,
            '_maxthreads': str(tier_config['maxthreads']),
        })

        # Update in-memory config so new DB connections use the new tier
        self.config['_resource_tier'] = tier
        self.config['_maxthreads'] = tier_config['maxthreads']

        dbh.auditLog(
            self.currentUser(), 'RESOURCE_TIER_CHANGE',
            detail=f'Resource tier changed from {old_tier_name} to {tier}',
            ip_address=self.clientIP(),
        )

        restart_required = (
            tier_config['cherrypy_thread_pool'] != old_tier_config['cherrypy_thread_pool']
        )

        return {
            'success': True,
            'tier': tier,
            'restart_required': restart_required,
        }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def checkupdate(self: 'SpiderFootWebUi') -> dict:
        """Check for available updates from GitHub Releases (admin only).

        Returns:
            dict: update availability info
        """
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}

        import urllib.request
        import ssl

        dbh = SpiderFootDb(self.config)
        dbh.auditLog(self.currentUser(), 'UPDATE_CHECK',
                     detail='Checked for updates',
                     ip_address=self.clientIP())

        try:
            api_url = "https://api.github.com/repos/0x31i/asm-ng/releases/latest"
            req = urllib.request.Request(api_url, headers={
                'Accept': 'application/vnd.github.v3+json',
                'User-Agent': f'ASM-NG/{__version__}'
            })

            ctx = ssl.create_default_context()
            with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
                release_data = json.loads(resp.read().decode('utf-8'))

            latest_tag = release_data.get('tag_name', '').lstrip('v')
            current_version = __version__

            # Version comparison via tuple
            try:
                latest_parts = tuple(int(x) for x in latest_tag.split('.'))
                current_parts = tuple(int(x) for x in current_version.split('.'))
                update_available = latest_parts > current_parts
            except (ValueError, AttributeError):
                update_available = False

            return {
                'success': True,
                'current_version': current_version,
                'latest_version': latest_tag,
                'update_available': update_available,
                'release_name': release_data.get('name', ''),
                'release_notes': release_data.get('body', ''),
                'release_url': release_data.get('html_url', ''),
                'published_at': release_data.get('published_at', '')
            }
        except Exception as e:
            return {'success': False, 'error': f'Failed to check for updates: {str(e)}'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def applyupdate(self: 'SpiderFootWebUi', version: str) -> dict:
        """Apply an update by checking out a git tag (admin only).

        Args:
            version (str): version tag to checkout

        Returns:
            dict: update result
        """
        if self.currentUserRole() != 'admin':
            return {'success': False, 'error': 'Unauthorized'}

        import subprocess

        dbh = SpiderFootDb(self.config)
        tag = f"v{version}" if not version.startswith('v') else version
        steps_completed = []

        try:
            # Step 1: Backup the database
            db_path = self.config.get('__database', '')
            if db_path:
                backup_dir = os.path.dirname(db_path) or '.'
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                backup_path = os.path.join(backup_dir, f"spiderfoot_pre_update_{timestamp}.db")
                dbh.backupDB(backup_path)
                steps_completed.append(f"Database backed up to {backup_path}")

            # Step 2: Git fetch
            repo_dir = os.path.dirname(os.path.abspath(__file__))
            result = subprocess.run(
                ['git', 'fetch', 'origin', '--tags'],
                cwd=repo_dir, capture_output=True, text=True, timeout=60
            )
            if result.returncode != 0:
                return {'success': False, 'error': f'git fetch failed: {result.stderr}',
                        'steps_completed': steps_completed}
            steps_completed.append('Fetched latest from origin')

            # Step 3: Check if tag exists
            result = subprocess.run(
                ['git', 'tag', '-l', tag],
                cwd=repo_dir, capture_output=True, text=True, timeout=10
            )
            if tag not in result.stdout.strip().split('\n'):
                return {'success': False, 'error': f'Tag {tag} not found',
                        'steps_completed': steps_completed}

            # Step 4: Git checkout tag
            result = subprocess.run(
                ['git', 'checkout', tag],
                cwd=repo_dir, capture_output=True, text=True, timeout=30
            )
            if result.returncode != 0:
                return {'success': False, 'error': f'git checkout failed: {result.stderr}',
                        'steps_completed': steps_completed}
            steps_completed.append(f'Checked out {tag}')

            # Step 5: Check if requirements changed and pip install
            result = subprocess.run(
                ['git', 'diff', 'HEAD~1..HEAD', '--name-only'],
                cwd=repo_dir, capture_output=True, text=True, timeout=10
            )
            if 'requirements.txt' in result.stdout:
                pip_result = subprocess.run(
                    ['pip', 'install', '-r', 'requirements.txt'],
                    cwd=repo_dir, capture_output=True, text=True, timeout=120
                )
                if pip_result.returncode != 0:
                    steps_completed.append(f'pip install warning: {pip_result.stderr[:200]}')
                else:
                    steps_completed.append('Updated dependencies from requirements.txt')

            # Audit log
            dbh.auditLog(self.currentUser(), 'UPDATE_APPLIED',
                         detail=f'Updated to {tag}',
                         ip_address=self.clientIP())

            return {
                'success': True,
                'message': f'Updated to {tag}. Please restart the application for changes to take effect.',
                'steps_completed': steps_completed,
                'restart_required': True
            }
        except subprocess.TimeoutExpired:
            return {'success': False, 'error': 'Update operation timed out',
                    'steps_completed': steps_completed}
        except Exception as e:
            return {'success': False, 'error': str(e),
                    'steps_completed': steps_completed}

    @cherrypy.expose
    def optsexport(self: 'SpiderFootWebUi', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        self.requireAdmin()
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="SpiderFoot.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'SpiderFootWebUi') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        self.requireAdmin()
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." +
                            mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'SpiderFootWebUi', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        # Audit log: scan deleted
        dbh.auditLog(
            self.currentUser() or 'unknown', 'SCAN_DELETE',
            detail=f"Deleted scan(s): {id}",
            ip_address=self.clientIP()
        )

        return ""

    @cherrypy.expose
    def savesettings(self: 'SpiderFootWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        self.requireAdmin()
        if str(token) != str(self.token):
            return self.error(f"Invalid token ({token})")

        # configFile seems to get set even if a file isn't uploaded
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                tmp = dict()
                for line in contents.split("\n"):
                    if "=" not in line:
                        continue

                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array[1] = ""

                    tmp[opt_array[0]] = '='.join(opt_array[1:])

                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f"Failed to parse input file. Was it generated from SpiderFoot? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                dbh_audit = SpiderFootDb(self.config)
                dbh_audit.auditLog(
                    self.currentUser() or 'unknown', 'SETTINGS_RESET',
                    detail='Settings reset to factory default',
                    ip_address=self.clientIP()
                )
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("Failed to reset settings")

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                value = useropts[opt]
                if not isinstance(value, str):
                    value = str(value)
                cleaned = self.cleanUserInput([value])
                cleanopts[opt] = cleaned[0] if cleaned and len(cleaned) > 0 else ""

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            import logging
            logging.exception("Error processing user input in savesettings")
            return self.error(f"Processing one or more of your inputs failed: {e}")

        # Audit log: settings saved
        dbh.auditLog(
            self.currentUser() or 'unknown', 'SETTINGS_SAVE',
            detail='Settings saved via web UI',
            ip_address=self.clientIP()
        )

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'SpiderFootWebUi', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        self.requireAdmin()
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"Invalid token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                dbh_audit = SpiderFootDb(self.config)
                dbh_audit.auditLog(
                    self.currentUser() or 'unknown', 'SETTINGS_RESET',
                    detail='Settings reset to factory default (raw)',
                    ip_address=self.clientIP()
                )
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Failed to reset settings"]).encode('utf-8')

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"Processing one or more of your inputs failed: {e}"]).encode('utf-8')

        # Audit log: settings saved
        dbh.auditLog(
            self.currentUser() or 'unknown', 'SETTINGS_SAVE',
            detail='Settings saved via raw API',
            ip_address=self.clientIP()
        )

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'SpiderFootWebUi') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str, force: str = "0") -> str:
        """Set a bunch of results (hashes) as false positive or validated.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 (unvalidated), 1 (false positive), or 2 (validated)
            force (str): 0 or 1 - bypass parent element check when unsetting

        Returns:
            str: set status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1", "2"]:
            return json.dumps(["ERROR", "Invalid status flag. Use 0 (unvalidated), 1 (false positive), or 2 (validated)."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP (unless force is set).
        if fp == "0" and force != "1":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive. Use force option to override."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        # When force unsetting, only unset the selected items, not children
        if force == "1" and fp == "0":
            allIds = ids
        else:
            childs = dbh.scanElementChildrenAll(id, ids)
            allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, int(fp))
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    def resultsetfppersist(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str, persist: str = "0", force: str = "0") -> str:
        """Set results as false positive or validated with optional target-level persistence.

        This extends resultsetfp to optionally persist false positives or validated status
        at the target level, so they will be recognized in future scans of the same target.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 (unvalidated), 1 (false positive), or 2 (validated)
            persist (str): 0 or 1 - whether to persist at target level
            force (str): 0 or 1 - bypass parent element check when unsetting

        Returns:
            str: set status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1", "2"]:
            return json.dumps(["ERROR", "Invalid status flag. Use 0 (unvalidated), 1 (false positive), or 2 (validated)."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        target = status[1]  # seed_target

        # Auto-enable persistence if there are multiple scans for this target
        # This ensures FP changes always sync across all scans of the same target
        if persist != "1":
            scanCount = dbh.scanCountForTarget(target)
            if scanCount > 1:
                persist = "1"

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP (unless force is set).
        if fp == "0" and force != "1":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive. Use force option to override."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        # When force unsetting, only unset the selected items, not children
        if force == "1" and fp == "0":
            allIds = ids
        else:
            childs = dbh.scanElementChildrenAll(id, ids)
            allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, int(fp))

        # Handle target-level persistence and cross-scan sync
        if ret and persist == "1":
            # Get the event details for each ID to persist at target level
            events = dbh.scanResultEvent(id)
            eventMap = {row[8]: row for row in events}  # hash -> event data

            for resultId in allIds:  # Persist all marked items including children
                if resultId in eventMap:
                    eventData = eventMap[resultId]
                    eventType = eventData[4]  # type
                    data = eventData[1]  # data
                    sourceData = eventData[2]  # source data for granular matching

                    if fp == "1":
                        # Mark as false positive - add to FP table, remove from validated
                        dbh.targetFalsePositiveAdd(target, eventType, data, sourceData)
                        dbh.targetValidatedRemove(target, eventType, data, sourceData)
                    elif fp == "2":
                        # Mark as validated - add to validated table, remove from FP
                        dbh.targetValidatedAdd(target, eventType, data, sourceData)
                        dbh.targetFalsePositiveRemove(target, eventType, data, sourceData)
                        # Two-way sync: also add to known assets as ANALYST_CONFIRMED
                        try:
                            ka_type = 'domain'
                            if eventType in ('IP_ADDRESS', 'IPV6_ADDRESS', 'AFFILIATE_IPADDR'):
                                ka_type = 'ip'
                            elif eventType in ('HUMAN_NAME', 'USERNAME', 'EMAILADDR',
                                               'AFFILIATE_EMAILADDR', 'SOCIAL_MEDIA'):
                                ka_type = 'employee'
                            current_user = cherrypy.session.get('user', 'anonymous')
                            dbh.knownAssetAdd(target, ka_type, data,
                                              source='ANALYST_CONFIRMED', addedBy=current_user)
                        except Exception:
                            pass  # Non-critical - asset table may not exist on older DBs
                    else:
                        # Clear status - remove from both tables
                        dbh.targetFalsePositiveRemove(target, eventType, data, sourceData)
                        dbh.targetValidatedRemove(target, eventType, data, sourceData)

                    # Sync the scan-level FP flag across all scans of the same target
                    # This ensures that if you change FP status on one scan, all other scans
                    # with the same entry (same type, data, source_data) are also updated
                    dbh.syncFalsePositiveAcrossScans(target, eventType, data, sourceData, int(fp))

        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def targetfplist(self: 'SpiderFootWebUi', target: str = None) -> list:
        """List target-level false positives.

        Args:
            target (str): optional target to filter by

        Returns:
            list: list of target-level false positives
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        fps = dbh.targetFalsePositiveList(target)

        ret = []
        for fp in fps:
            ret.append({
                'id': fp[0],
                'target': fp[1],
                'event_type': fp[2],
                'event_data': fp[3],
                'date_added': fp[4],
                'notes': fp[5]
            })

        return ret

    @cherrypy.expose
    def targetfpadd(self: 'SpiderFootWebUi', target: str, event_type: str, event_data: str, notes: str = None) -> str:
        """Add a target-level false positive.

        Args:
            target (str): target value
            event_type (str): event type
            event_data (str): event data
            notes (str): optional notes

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if not target or not event_type or not event_data:
            return json.dumps(["ERROR", "Missing required parameters."]).encode('utf-8')

        try:
            ret = dbh.targetFalsePositiveAdd(target, event_type, event_data, notes)
            if ret:
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    def targetfpremove(self: 'SpiderFootWebUi', id: str = None, target: str = None, event_type: str = None, event_data: str = None) -> str:
        """Remove a target-level false positive.

        Can be removed by ID or by target/event_type/event_data combination.

        Args:
            id (str): false positive entry ID
            target (str): target value
            event_type (str): event type
            event_data (str): event data

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        try:
            if id:
                ret = dbh.targetFalsePositiveRemoveById(int(id))
            elif target and event_type and event_data:
                ret = dbh.targetFalsePositiveRemove(target, event_type, event_data)
            else:
                return json.dumps(["ERROR", "Must provide either ID or target/event_type/event_data."]).encode('utf-8')

            if ret:
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def targetvalidatedlist(self: 'SpiderFootWebUi', target: str = None) -> list:
        """List target-level validated entries.

        Args:
            target (str): optional target to filter by

        Returns:
            list: list of target-level validated entries
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        vals = dbh.targetValidatedList(target)

        ret = []
        for val in vals:
            ret.append({
                'id': val[0],
                'target': val[1],
                'event_type': val[2],
                'event_data': val[3],
                'date_added': val[4],
                'notes': val[5]
            })

        return ret

    @cherrypy.expose
    def targetvalidatedadd(self: 'SpiderFootWebUi', target: str, event_type: str, event_data: str, notes: str = None) -> str:
        """Add a target-level validated entry.

        Args:
            target (str): target value
            event_type (str): event type
            event_data (str): event data
            notes (str): optional notes

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if not target or not event_type or not event_data:
            return json.dumps(["ERROR", "Missing required parameters."]).encode('utf-8')

        try:
            ret = dbh.targetValidatedAdd(target, event_type, event_data, notes)
            if ret:
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    def targetvalidatedremove(self: 'SpiderFootWebUi', id: str = None, target: str = None, event_type: str = None, event_data: str = None) -> str:
        """Remove a target-level validated entry.

        Can be removed by ID or by target/event_type/event_data combination.

        Args:
            id (str): validated entry ID
            target (str): target value
            event_type (str): event type
            event_data (str): event data

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        try:
            if id:
                ret = dbh.targetValidatedRemoveById(int(id))
            elif target and event_type and event_data:
                ret = dbh.targetValidatedRemove(target, event_type, event_data)
            else:
                return json.dumps(["ERROR", "Must provide either ID or target/event_type/event_data."]).encode('utf-8')

            if ret:
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    # -------------------------------------------------------------------
    # Known Assets endpoints
    # -------------------------------------------------------------------

    @cherrypy.expose
    def knownassetlist(self: 'SpiderFootWebUi', target: str = None, asset_type: str = None) -> str:
        """List known assets for a target.

        Args:
            target: scan target
            asset_type: optional filter by type (ip, domain, employee)

        Returns:
            str: JSON list of assets
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not target:
            return json.dumps([]).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        try:
            rows = dbh.knownAssetList(target, asset_type)
            result = []
            for r in rows:
                result.append({
                    'id': r[0],
                    'target': r[1],
                    'asset_type': r[2],
                    'asset_value': r[3],
                    'source': r[4],
                    'import_batch': r[5],
                    'date_added': r[6],
                    'added_by': r[7],
                    'notes': r[8]
                })
            return json.dumps(result).encode('utf-8')
        except Exception as e:
            return json.dumps([]).encode('utf-8')

    @cherrypy.expose
    def knownassetadd(self: 'SpiderFootWebUi', target: str = None, asset_type: str = None,
                      asset_value: str = None, source: str = 'CLIENT_PROVIDED',
                      notes: str = None) -> str:
        """Add a single known asset.

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not target or not asset_type or not asset_value:
            return json.dumps(["ERROR", "target, asset_type, and asset_value are required."]).encode('utf-8')

        if asset_type not in ('ip', 'domain', 'employee'):
            return json.dumps(["ERROR", "asset_type must be ip, domain, or employee."]).encode('utf-8')

        if source not in ('CLIENT_PROVIDED', 'ANALYST_CONFIRMED'):
            source = 'CLIENT_PROVIDED'

        current_user = cherrypy.session.get('user', 'anonymous')
        dbh = SpiderFootDb(self.config)
        try:
            dbh.knownAssetAdd(target, asset_type, asset_value.strip(),
                              source=source, addedBy=current_user, notes=notes)
            return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

    @cherrypy.expose
    def knownassetremove(self: 'SpiderFootWebUi', id: str = None, ids: str = None) -> str:
        """Remove known asset(s) by ID.

        Args:
            id: single asset ID
            ids: JSON array of IDs for bulk removal

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)
        try:
            if ids:
                id_list = json.loads(ids)
                count = dbh.knownAssetRemoveBulk(id_list)
                return json.dumps(["SUCCESS", f"Removed {count} assets."]).encode('utf-8')
            elif id:
                dbh.knownAssetRemove(assetId=int(id))
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            else:
                return json.dumps(["ERROR", "id or ids required."]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

    @cherrypy.expose
    def knownassetupdate(self: 'SpiderFootWebUi', id: str = None,
                         notes: str = None, source: str = None) -> str:
        """Update a known asset's notes or source.

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not id:
            return json.dumps(["ERROR", "id is required."]).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        try:
            dbh.knownAssetUpdate(int(id), notes=notes, source=source)
            return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

    @cherrypy.expose
    def knownassetimport(self: 'SpiderFootWebUi', target: str = None,
                         asset_type: str = None, importfile: object = None) -> str:
        """Import known assets from a file (.txt, .csv, .xlsx).

        Args:
            target: scan target
            asset_type: 'ip', 'domain', or 'employee'
            importfile: uploaded file object

        Returns:
            str: JSON status with count
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not target or not asset_type or not importfile:
            return json.dumps(["ERROR", "target, asset_type, and importfile are required."]).encode('utf-8')

        if asset_type not in ('ip', 'domain', 'employee'):
            return json.dumps(["ERROR", "asset_type must be ip, domain, or employee."]).encode('utf-8')

        current_user = cherrypy.session.get('user', 'anonymous')
        file_name = getattr(importfile, 'filename', 'unknown')
        file_ext = file_name.lower().rsplit('.', 1)[-1] if '.' in file_name else ''

        try:
            raw = importfile.file.read()
        except Exception as e:
            return json.dumps(["ERROR", f"Failed to read file: {str(e)}"]).encode('utf-8')

        assets = []
        import_batch = str(uuid.uuid4())[:8]

        try:
            if file_ext == 'txt':
                # One value per line
                text = raw.decode('utf-8', errors='ignore')
                for line in text.splitlines():
                    val = line.strip()
                    if val and not val.startswith('#'):
                        assets.append(val)

            elif file_ext == 'csv':
                text = raw.decode('utf-8', errors='ignore')
                reader = csv.reader(StringIO(text))
                # Try to detect if there's a header
                first_row = None
                for row in reader:
                    if not row:
                        continue
                    if first_row is None:
                        first_row = row
                        # If first cell looks like a header, skip it
                        cell = row[0].strip().lower()
                        if cell in ('ip', 'ip address', 'ip_address', 'domain', 'hostname',
                                    'subdomain', 'name', 'employee', 'email', 'value',
                                    'asset', 'host', 'address', 'fqdn'):
                            continue
                    # Take the first column
                    val = row[0].strip()
                    if val:
                        assets.append(val)

            elif file_ext in ('xlsx', 'xls'):
                wb = openpyxl.load_workbook(BytesIO(raw), read_only=True)
                ws = wb.active
                first_row = True
                for row in ws.iter_rows(values_only=True):
                    if not row or not row[0]:
                        continue
                    val = str(row[0]).strip()
                    if first_row:
                        first_row = False
                        cell_lower = val.lower()
                        if cell_lower in ('ip', 'ip address', 'ip_address', 'domain', 'hostname',
                                          'subdomain', 'name', 'employee', 'email', 'value',
                                          'asset', 'host', 'address', 'fqdn'):
                            continue
                    if val:
                        assets.append(val)
                wb.close()
            else:
                return json.dumps(["ERROR", f"Unsupported file type: .{file_ext}. Use .txt, .csv, or .xlsx"]).encode('utf-8')

        except Exception as e:
            return json.dumps(["ERROR", f"Failed to parse file: {str(e)}"]).encode('utf-8')

        if not assets:
            return json.dumps(["ERROR", "No valid entries found in file."]).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        try:
            count = dbh.knownAssetAddBulk(target, asset_type, assets,
                                          source='CLIENT_PROVIDED',
                                          importBatch=import_batch,
                                          addedBy=current_user)
            dbh.assetImportHistoryAdd(target, asset_type, file_name, count, current_user)
            return json.dumps(["SUCCESS", f"Imported {count} new assets ({len(assets)} total in file)."]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

    @cherrypy.expose
    def knownassetmatches(self: 'SpiderFootWebUi', id: str = None) -> str:
        """Find scan results that match known assets (Potential Matches).

        Args:
            id: scan instance ID

        Returns:
            str: JSON list of matches
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not id:
            return json.dumps([]).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        try:
            # Get scan target
            scan_info = dbh.scanInstanceGet(id)
            if not scan_info:
                return json.dumps([]).encode('utf-8')
            target = scan_info[1]  # seed_target

            # Get target-level FP/validated status
            targetFps = dbh.targetFalsePositivesForTarget(target)
            targetValidated = dbh.targetValidatedForTarget(target)

            matches = dbh.knownAssetMatchScanResults(id, target)

            # Enrich matches with target-level status
            for m in matches:
                m['isTargetFp'] = 0
                m['isTargetValidated'] = 0
                # Check target-level status
                for fp_tuple in targetFps:
                    if fp_tuple[0] == m['type'] and fp_tuple[1] == m['data']:
                        m['isTargetFp'] = 1
                        break
                for val_tuple in targetValidated:
                    if val_tuple[0] == m['type'] and val_tuple[1] == m['data']:
                        m['isTargetValidated'] = 1
                        break

            return json.dumps(matches).encode('utf-8')
        except Exception as e:
            return json.dumps([]).encode('utf-8')

    @cherrypy.expose
    def knownassetcount(self: 'SpiderFootWebUi', target: str = None) -> str:
        """Get known asset counts for a target.

        Returns:
            str: JSON with count breakdown
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not target:
            return json.dumps({'total': 0}).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        try:
            counts = dbh.knownAssetCount(target)
            return json.dumps(counts).encode('utf-8')
        except Exception as e:
            return json.dumps({'total': 0}).encode('utf-8')

    @cherrypy.expose
    def knownassetexport(self: 'SpiderFootWebUi', target: str = None, format: str = 'csv') -> bytes:
        """Export known assets as CSV or XLSX.

        Returns:
            bytes: file data
        """
        if not target:
            return b''

        dbh = SpiderFootDb(self.config)
        rows = dbh.knownAssetList(target)

        if format == 'excel':
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            fname = self._export_filename(target, '', 'ASSETS', 'xlsx')
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Known Assets"
            ws.append(["Type", "Value", "Source", "Date Added", "Added By", "Notes"])
            for r in rows:
                ws.append([r[2], r[3], r[4], r[6], r[7], r[8]])
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
        else:
            cherrypy.response.headers['Content-Type'] = "text/csv; charset=utf-8"
            fname = self._export_filename(target, '', 'ASSETS', 'csv')
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(["Type", "Value", "Source", "Date Added", "Added By", "Notes"])
            for r in rows:
                writer.writerow([r[2], r[3], r[4], r[6], r[7], r[8]])
            return output.getvalue().encode('utf-8')

    @cherrypy.expose
    def knownassetexportzip(self: 'SpiderFootWebUi', target: str = None) -> bytes:
        """Export known assets as a ZIP containing separate CSVs per asset type.

        Creates three CSV files bundled in a zip:
        - IPs.csv
        - DOMAINS.csv
        - EMPLOYEES.csv

        Args:
            target (str): the scan target

        Returns:
            bytes: ZIP file data
        """
        if not target:
            return b''

        import zipfile

        dbh = SpiderFootDb(self.config)
        rows = dbh.knownAssetList(target)

        # Separate rows by asset type
        asset_buckets = {
            'ip': [],
            'domain': [],
            'employee': [],
        }
        for r in rows:
            asset_type = r[2]
            if asset_type in asset_buckets:
                asset_buckets[asset_type].append(r)

        headers = ["Type", "Value", "Source", "Date Added", "Added By", "Notes"]

        type_filenames = {
            'ip': 'IPs.csv',
            'domain': 'DOMAINS.csv',
            'employee': 'EMPLOYEES.csv',
        }

        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for atype, arows in asset_buckets.items():
                buf = StringIO()
                writer = csv.writer(buf, dialect='excel')
                writer.writerow(headers)
                for r in arows:
                    writer.writerow([r[2], r[3], r[4], r[6], r[7], r[8]])
                zf.writestr(type_filenames[atype], buf.getvalue())

        zip_buf.seek(0)

        fname = self._export_filename(target, '', 'ASSETS', 'zip')
        cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
        cherrypy.response.headers['Content-Type'] = "application/zip"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return zip_buf.read()

    @cherrypy.expose
    def knownassetimporthistory(self: 'SpiderFootWebUi', target: str = None) -> str:
        """Get asset import history for a target.

        Returns:
            str: JSON list of imports
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not target:
            return json.dumps([]).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        try:
            rows = dbh.assetImportHistoryList(target)
            result = []
            for r in rows:
                result.append({
                    'id': r[0],
                    'target': r[1],
                    'asset_type': r[2],
                    'file_name': r[3],
                    'item_count': r[4],
                    'imported_by': r[5],
                    'date_imported': r[6]
                })
            return json.dumps(result).encode('utf-8')
        except Exception:
            return json.dumps([]).encode('utf-8')

    @cherrypy.expose
    def knownassetsyncverified(self: 'SpiderFootWebUi', id: str = None) -> str:
        """Sync existing validated (FP=2) scan results into known assets.

        Scans tbl_scan_results for all entries with false_positive=2 and
        adds them to tbl_known_assets as ANALYST_CONFIRMED. This back-fills
        known assets from verified rows that existed before the assets feature.

        Args:
            id: scan instance ID

        Returns:
            str: JSON status with count
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not id:
            return json.dumps(["ERROR", "scan id is required."]).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        current_user = cherrypy.session.get('user', 'anonymous')

        try:
            scan_info = dbh.scanInstanceGet(id)
            if not scan_info:
                return json.dumps(["ERROR", "Scan not found."]).encode('utf-8')
            target = scan_info[1]

            # Get all validated entries from this scan
            events = dbh.scanResultEvent(id)
            count = 0
            ip_types = {'IP_ADDRESS', 'IPV6_ADDRESS', 'AFFILIATE_IPADDR'}
            domain_types = {'DOMAIN_NAME', 'INTERNET_NAME', 'AFFILIATE_INTERNET_NAME',
                            'CO_HOSTED_SITE', 'SIMILARDOMAIN', 'INTERNET_NAME_UNRESOLVED'}
            employee_types = {'HUMAN_NAME', 'USERNAME', 'EMAILADDR',
                              'AFFILIATE_EMAILADDR', 'SOCIAL_MEDIA'}

            for ev in events:
                fp_flag = ev[13]  # false_positive column
                if fp_flag != 2:
                    continue
                event_type = ev[4]
                event_data = ev[1]
                if not event_data:
                    continue

                # Determine asset type
                asset_type = None
                if event_type in ip_types:
                    asset_type = 'ip'
                elif event_type in domain_types:
                    asset_type = 'domain'
                elif event_type in employee_types:
                    asset_type = 'employee'

                if asset_type:
                    try:
                        dbh.knownAssetAdd(target, asset_type, event_data,
                                          source='ANALYST_CONFIRMED', addedBy=current_user)
                        count += 1
                    except Exception:
                        pass  # Duplicate - already exists

            # Also sync from target-level validated entries
            try:
                targetValidated = dbh.targetValidatedForTarget(target)
                for (evt, evd, esd) in targetValidated:
                    if not evd:
                        continue
                    asset_type = None
                    if evt in ip_types:
                        asset_type = 'ip'
                    elif evt in domain_types:
                        asset_type = 'domain'
                    elif evt in employee_types:
                        asset_type = 'employee'
                    if asset_type:
                        try:
                            dbh.knownAssetAdd(target, asset_type, evd,
                                              source='ANALYST_CONFIRMED', addedBy=current_user)
                            count += 1
                        except Exception:
                            pass
            except Exception:
                pass

            return json.dumps(["SUCCESS", f"Synced {count} verified entries to known assets."]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

    @cherrypy.expose
    def knownassetverifymatch(self: 'SpiderFootWebUi', id: str = None,
                              result_hash: str = None, result_hashes: str = None,
                              action: str = 'verify') -> str:
        """Verify or FP a potential match from the assets page.

        This handles the two-way sync: when verifying, marks the scan result as
        validated (FP=2), persists at target level, and adds to known assets
        as ANALYST_CONFIRMED.

        Args:
            id: scan instance ID
            result_hash: single result hash
            result_hashes: JSON array of hashes
            action: 'verify', 'fp', 'reset', or 'dismiss'

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if not id:
            return json.dumps(["ERROR", "scan id is required."]).encode('utf-8')

        hashes = []
        if result_hashes:
            hashes = json.loads(result_hashes)
        elif result_hash:
            hashes = [result_hash]
        else:
            return json.dumps(["ERROR", "result_hash or result_hashes required."]).encode('utf-8')

        dbh = SpiderFootDb(self.config)
        current_user = cherrypy.session.get('user', 'anonymous')

        try:
            scan_info = dbh.scanInstanceGet(id)
            if not scan_info:
                return json.dumps(["ERROR", "Scan not found."]).encode('utf-8')
            target = scan_info[1]

            # Build event map from all scan results (same pattern as resultsetfppersist)
            events = dbh.scanResultEvent(id)
            eventMap = {row[8]: row for row in events}  # hash -> event row

            if action == 'verify':
                # Mark as validated (FP=2) in scan with children
                childs = dbh.scanElementChildrenAll(id, hashes)
                allIds = hashes + childs
                dbh.scanResultsUpdateFP(id, allIds, 2)

                # Persist at target level and add to known assets
                for h in allIds:
                    if h in eventMap:
                        ev = eventMap[h]
                        event_type = ev[4]
                        event_data = ev[1]
                        source_data = ev[2]
                        # Persist as target validated
                        dbh.targetValidatedAdd(target, event_type, event_data, source_data)
                        dbh.targetFalsePositiveRemove(target, event_type, event_data, source_data)
                        # Sync across scans
                        dbh.syncFalsePositiveAcrossScans(target, event_type, event_data, source_data, 2)

                # Add only the directly matched items (not children) to known assets
                for h in hashes:
                    if h in eventMap:
                        ev = eventMap[h]
                        event_type = ev[4]
                        event_data = ev[1]
                        asset_type = 'domain'
                        if event_type in ('IP_ADDRESS', 'IPV6_ADDRESS', 'AFFILIATE_IPADDR'):
                            asset_type = 'ip'
                        elif event_type in ('HUMAN_NAME', 'USERNAME', 'EMAILADDR',
                                            'AFFILIATE_EMAILADDR', 'SOCIAL_MEDIA'):
                            asset_type = 'employee'
                        dbh.knownAssetAdd(target, asset_type, event_data,
                                          source='ANALYST_CONFIRMED', addedBy=current_user)

                return json.dumps(["SUCCESS", f"Verified {len(hashes)} items."]).encode('utf-8')

            elif action == 'fp':
                # Mark as false positive (FP=1) in scan with children
                childs = dbh.scanElementChildrenAll(id, hashes)
                allIds = hashes + childs
                dbh.scanResultsUpdateFP(id, allIds, 1)

                for h in allIds:
                    if h in eventMap:
                        ev = eventMap[h]
                        dbh.targetFalsePositiveAdd(target, ev[4], ev[1], ev[2])
                        dbh.targetValidatedRemove(target, ev[4], ev[1], ev[2])
                        dbh.syncFalsePositiveAcrossScans(target, ev[4], ev[1], ev[2], 1)

                return json.dumps(["SUCCESS", f"Marked {len(hashes)} as false positive."]).encode('utf-8')

            elif action == 'reset':
                # Reset to pending (FP=0) in scan with children
                childs = dbh.scanElementChildrenAll(id, hashes)
                allIds = hashes + childs
                dbh.scanResultsUpdateFP(id, allIds, 0)

                for h in allIds:
                    if h in eventMap:
                        ev = eventMap[h]
                        event_type = ev[4]
                        event_data = ev[1]
                        source_data = ev[2]
                        # Remove from both target-level tables
                        dbh.targetValidatedRemove(target, event_type, event_data, source_data)
                        dbh.targetFalsePositiveRemove(target, event_type, event_data, source_data)
                        # Sync FP=0 across scans
                        dbh.syncFalsePositiveAcrossScans(target, event_type, event_data, source_data, 0)

                # Remove ANALYST_CONFIRMED known assets for directly matched items
                for h in hashes:
                    if h in eventMap:
                        ev = eventMap[h]
                        event_type = ev[4]
                        event_data = ev[1]
                        asset_type = 'domain'
                        if event_type in ('IP_ADDRESS', 'IPV6_ADDRESS', 'AFFILIATE_IPADDR'):
                            asset_type = 'ip'
                        elif event_type in ('HUMAN_NAME', 'USERNAME', 'EMAILADDR',
                                            'AFFILIATE_EMAILADDR', 'SOCIAL_MEDIA'):
                            asset_type = 'employee'
                        dbh.knownAssetRemove(target=target, assetType=asset_type, assetValue=event_data)

                return json.dumps(["SUCCESS", f"Reset {len(hashes)} items to pending."]).encode('utf-8')

            elif action == 'dismiss':
                return json.dumps(["SUCCESS", "Dismissed."]).encode('utf-8')

            else:
                return json.dumps(["ERROR", f"Unknown action: {action}"]).encode('utf-8')

        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'SpiderFootWebUi') -> list:
        """List all event types.

        Returns:
            list: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'SpiderFootWebUi', **kwargs) -> list:
        """List all available modules.

        Returns:
            list: list of available modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        
        modlist = list()
        for mod in self.config['__modules__']:
            if "__" in mod:
                continue
            
            modinfo = self.config['__modules__'][mod]
            modlist.append({
                'name': mod,
                'label': modinfo.get('name', mod),
                'descr': modinfo.get('descr', ''),
                'summary': modinfo.get('meta', {}).get('summary', ''),
                'provides': modinfo.get('provides', []),
                'consumes': modinfo.get('consumes', []),
                'cats': modinfo.get('cats', []),
                'group': modinfo.get('group', [])
            })
            
        return sorted(modlist, key=lambda x: x['name'])

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'SpiderFootWebUi') -> list:
        """List all available correlation rules.

        Returns:
            list: list of available correlation rules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        
        rules = list()
        for rule in self.config.get('__correlationrules__', []):
            rules.append({
                'id': rule.get('id', ''),
                'name': rule.get('name', ''),
                'risk': rule.get('risk', 'UNKNOWN'),
                'description': rule.get('description', '')
            })
            
        return sorted(rules, key=lambda x: x['name'])

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'SpiderFootWebUi') -> list:
        """Ping endpoint for health checks.

        Returns:
            list: status response
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'SpiderFootWebUi', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = SpiderFootDb(self.config)

        if not query:
            return self.jsonify_error('400', "Invalid query.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'SpiderFootWebUi', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str, launch_code: str = None) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)
            launch_code (str): launch code for non-admin users

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        # Launch code check for non-admin users
        if self.currentUserRole() != 'admin':
            dbh_lc = SpiderFootDb(self.config)
            stored_code = dbh_lc.launchCodeGet()
            if not stored_code:
                error_msg = "Scan launching is disabled. An administrator must set a launch code first."
                if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                    cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                    return json.dumps(["ERROR", error_msg]).encode('utf-8')
                return self.error(error_msg)
            if launch_code != stored_code:
                error_msg = "Invalid launch code."
                if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                    cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                    return json.dumps(["ERROR", error_msg]).encode('utf-8')
                return self.error(error_msg)

        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]

        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan name was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan name was not specified.")

        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan target was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan target was not specified.")

        if not typelist and not modulelist and not usecase:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

        # Swap the globalscantable for the database handler
        dbh = SpiderFootDb(self.config)

        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)

        modlist = list()

        # User selected modules
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        # User selected types
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')

            # 1. Find all modules that produce the requested types
            modlist = sf.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            # 2. For each type those modules consume, get modules producing
            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = sf.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        # User selected a use case
        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or ('group' in self.config['__modules__'][mod] and
                                        usecase in self.config['__modules__'][mod]['group']):
                    modlist.append(mod)

        # If we somehow got all the way through to here and still don't have any modules selected
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        # Add our mandatory storage module
        if "sfp__stor_db" not in modlist:
            modlist.append("sfp__stor_db")
        modlist.sort()        # Delete the stdout module in case it crept in
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        # Start running a new scan
        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        
        try:
            p = _spawn_ctx.Process(target=startSpiderFootScanner, args=(
                self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}", exc_info=True)
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Track the Process object for kill capability
        with self._scan_processes_lock:
            self._scan_processes[scanId] = p

        # Wait until the scan has initialized
        # Check the database for the scan status results
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        # Audit log: scan started
        dbh.auditLog(
            self.currentUser() or 'unknown', 'SCAN_START',
            detail=f"Scan '{scanname}' on target '{scantarget}' (ID: {scanId})",
            ip_address=self.clientIP()
        )

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(["SUCCESS", scanId]).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'SpiderFootWebUi', id: str) -> str:
        """Stop a scan.

        Supports graceful stop (ABORT-REQUESTED) and force-kill for stuck
        scans.  If the scan is already in ABORT-REQUESTED state, the
        process is killed directly and the status set to ABORTED.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')
        active_statuses = ("RUNNING", "STARTING", "STARTED", "INITIALIZING")
        force_kill_ids = []

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"Scan {scan_id} has already finished.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"Scan {scan_id} has already aborted.")

            if scan_status == "ERROR-FAILED":
                return self.jsonify_error('400', f"Scan {scan_id} has already failed.")

            if scan_status == "ABORT-REQUESTED":
                # Already requested once — escalate to force kill
                force_kill_ids.append(scan_id)
            elif scan_status in active_statuses:
                pass  # Will be handled below
            else:
                return self.jsonify_error(
                    '400',
                    f"The running scan is currently in the state '{scan_status}', "
                    f"please try again later or restart SpiderFoot."
                )

        for scan_id in ids:
            if scan_id in force_kill_ids:
                # Force-kill: the scan didn't respond to ABORT-REQUESTED
                self.log.warning(f"Force-killing stuck scan {scan_id}")
                self._kill_scan_process(scan_id)
                # Set ABORTED using direct connection (bypasses locked DB)
                try:
                    dbh.scanInstanceSet(scan_id, status="ABORTED", ended=time.time() * 1000)
                except Exception:
                    self._force_scan_status(scan_id, "ABORTED")
            else:
                # Graceful: request abort, kill process as backup
                try:
                    dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")
                except Exception:
                    # DB is locked — kill the process and force status
                    self.log.warning(
                        f"Database locked while stopping scan {scan_id}, "
                        f"killing process and forcing status"
                    )
                    self._kill_scan_process(scan_id)
                    self._force_scan_status(scan_id, "ABORTED")

        # Audit log: scan stopped
        try:
            dbh.auditLog(
                self.currentUser() or 'unknown', 'SCAN_STOP',
                detail=f"Stopped scan(s): {id}",
                ip_address=self.clientIP()
            )
        except Exception:
            pass  # Don't fail the stop operation over audit logging

        return ""

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatusoverride(self: 'SpiderFootWebUi', id: str, status: str) -> dict:
        """Override the status of a scan (admin only).

        Allows an administrator to manually set a scan's status, for example
        to mark a stuck scan as FINISHED so its results are preserved.
        If the scan is in an active state, the scan process is killed first.
        Falls back to a direct SQLite connection if the database is locked.

        Args:
            id (str): scan instance ID
            status (str): new status to set

        Returns:
            dict: JSON response with success or error
        """
        self.requireAuth()
        self.requireAdmin()

        valid_statuses = ("FINISHED", "ABORTED", "ERROR-FAILED")
        if status not in valid_statuses:
            return self.jsonify_error(
                '400',
                f"Invalid status '{status}'. Must be one of: {', '.join(valid_statuses)}"
            )

        if not id:
            return self.jsonify_error('400', "No scan specified")

        dbh = SpiderFootDb(self.config)

        try:
            res = dbh.scanInstanceGet(id)
        except Exception:
            res = None

        if not res:
            return self.jsonify_error('404', f"Scan {id} does not exist")

        old_status = res[5]

        # Kill the scan process if it's in an active state
        active_statuses = ("RUNNING", "STARTING", "STARTED", "INITIALIZING", "ABORT-REQUESTED")
        if old_status in active_statuses:
            self._kill_scan_process(id)

        # Try normal DB update first, fall back to direct connection
        try:
            dbh.scanInstanceSet(id, status=status, ended=time.time() * 1000)
        except Exception as e:
            self.log.warning(
                f"scanstatusoverride: normal DB update failed for {id}: {e}, "
                f"using direct connection"
            )
            if not self._force_scan_status(id, status):
                return self.jsonify_error(
                    '500',
                    f"Failed to override scan status: database is locked. "
                    f"The scan process may still be running."
                )

        try:
            dbh.auditLog(
                self.currentUser() or 'unknown', 'SCAN_STATUS_OVERRIDE',
                detail=f"Overrode scan {id} status from '{old_status}' to '{status}'",
                ip_address=self.clientIP()
            )
        except Exception:
            pass  # Don't fail the override over audit logging

        return {"success": True, "old_status": old_status, "new_status": status}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        """Vacuum the database."""
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Vacuuming the database failed"]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", f"Vacuuming the database failed: {e}"]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def dbhealth(self):
        """Get database health status.

        Returns:
            dict: database health information including file size,
                  page counts, integrity status, and WAL info.
        """
        dbh = SpiderFootDb(self.config)
        try:
            health = dbh.dbHealth()

            # Add human-readable file sizes
            if health.get('file_size'):
                size_mb = health['file_size'] / (1024 * 1024)
                health['file_size_human'] = f"{size_mb:.2f} MB"
            if health.get('wal_file_size'):
                wal_mb = health['wal_file_size'] / (1024 * 1024)
                health['wal_file_size_human'] = f"{wal_mb:.2f} MB"

            return health
        except Exception as e:
            return {'status': 'error', 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def dbbackup(self):
        """Create a backup of the database.

        Creates a hot backup using SQLite's built-in backup API.
        The backup is stored in the same directory as the database
        with a timestamped filename.

        Returns:
            dict: backup result with file path and size.
        """
        dbh = SpiderFootDb(self.config)
        try:
            # Determine backup path
            db_path = self.config.get('__database', '')
            if not db_path:
                return {'status': 'error', 'error': 'Database path not configured'}

            backup_dir = os.path.dirname(db_path)
            if not backup_dir:
                backup_dir = '.'
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            backup_filename = f"spiderfoot_backup_{timestamp}.db"
            backup_path = os.path.join(backup_dir, backup_filename)

            result_path = dbh.backupDB(backup_path)
            backup_size = os.path.getsize(result_path)
            size_mb = backup_size / (1024 * 1024)

            return {
                'status': 'success',
                'backup_path': result_path,
                'backup_size': backup_size,
                'backup_size_human': f"{size_mb:.2f} MB",
                'timestamp': timestamp
            }
        except Exception as e:
            return {'status': 'error', 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def dbintegritycheck(self):
        """Run a full integrity check on the database.

        Returns:
            dict: integrity check results with 'ok' status and details.
        """
        dbh = SpiderFootDb(self.config)
        try:
            return dbh.integrityCheck()
        except Exception as e:
            return {'ok': False, 'integrity_check': [str(e)], 'foreign_key_check': []}

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'SpiderFootWebUi', id: str, limit: str = None, rowId: str = None, reverse: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): TBD
            rowId (str): TBD
            reverse (str): TBD

        Returns:
            list: scan log
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2],
                           html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'SpiderFootWebUi', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'SpiderFootWebUi') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        try:
            dbh = SpiderFootDb(self.config)
            data = dbh.scanInstanceList()
        except Exception:
            return []
        retdata = []

        for row in data:
            created = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {
                "CRITICAL": 0,
                "HIGH": 0,
                "MEDIUM": 0,
                "LOW": 0,
                "INFO": 0
            }
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    if c[0] in riskmatrix:
                        riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created,
                           started, finished, row[6], row[7], riskmatrix])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'SpiderFootWebUi', id: str) -> list:
        """Show basic information about a scan, including status and number of
        each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        try:
            dbh = SpiderFootDb(self.config)
            data = dbh.scanInstanceGet(id)
        except Exception:
            return []

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {
            "CRITICAL": 0,
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        try:
            correlations = dbh.scanCorrelationSummary(id, by="risk")
            if correlations:
                for c in correlations:
                    if c[0] in riskmatrix:
                        riskmatrix[c[0]] = c[1]
        except Exception:
            pass

        findingsmatrix = {
            "CRITICAL": 0,
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        try:
            findings = dbh.scanFindingsList(id)
            if findings:
                for f in findings:
                    if f[1] in findingsmatrix:
                        findingsmatrix[f[1]] += 1
        except Exception:
            pass

        return [data[0], data[1], created, started, ended, data[5], riskmatrix, findingsmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanprogress(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return estimated scan progress as a percentage.

        Progress is estimated by comparing how many modules have produced
        results versus the total number of modules enabled for the scan.

        Args:
            id (str): scan ID

        Returns:
            dict: progress info (modulesTotal, modulesWithResults,
                  progressPercent, status)
        """
        dbh = SpiderFootDb(self.config)
        try:
            return dbh.scanProgress(id)
        except Exception:
            return {
                'status': 'UNKNOWN',
                'modulesTotal': 0,
                'modulesWithResults': 0,
                'modulesRunning': 0,
                'eventsQueued': 0,
                'totalEvents': 0,
                'eventsPerSecond': 0.0,
                'progressPercent': 0,
            }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'SpiderFootWebUi', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        if not statusdata:
            return retdata

        config_overrides = load_grade_config_overrides(self.config)
        overrides = config_overrides.get('event_overrides')

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            grading = get_event_grading(row[0], overrides)
            category = grading.get('category', 'Information / Reference')
            rank = grading.get('rank', 5)
            cat_meta = DEFAULT_GRADE_CATEGORIES.get(category, {})
            color = cat_meta.get('color', '#6b7280')
            weight = cat_meta.get('weight', 0.0)
            retdata.append([row[0], row[1], lastseen,
                           row[3], row[4], statusdata[5],
                           category, color, rank, weight])

        return retdata

    def _calculateScanGrade(self, dbh, scan_id: str) -> dict:
        """Calculate the overall grade for a scan based on its event type results.

        Only counts UNVERIFIED and OPEN items (false_positive = 0).
        Items marked as false positive (1) or validated (2) are excluded.

        Args:
            dbh: SpiderFootDb instance
            scan_id: scan instance ID

        Returns:
            dict: Full grade data with categories, overall score, and grade letter.
        """
        # Check if grading is enabled
        if not self.config.get('_grade_enabled', True):
            return {'enabled': False, 'overall_grade': '-', 'overall_score': 0, 'categories': {}}

        try:
            # Query scan results directly (no JOIN with tbl_event_types to avoid
            # silently dropping types not in that table).
            # Count by FP status: 0=unvalidated, 1=false positive, 2=validated.
            # Only unvalidated (0) items count toward grade penalties.
            # Existence of ANY row means the type was found (important for
            # zero_entries_fail logic: all-validated should NOT trigger fail penalty).
            qry = """SELECT r.type,
                count(CASE WHEN CAST(r.false_positive AS INTEGER) = 0 THEN 1 END) AS unval_total,
                count(DISTINCT CASE WHEN CAST(r.false_positive AS INTEGER) = 0 THEN r.data END) AS unval_unique,
                count(*) AS all_total,
                count(DISTINCT r.data) AS all_unique
                FROM tbl_scan_results r
                WHERE r.scan_instance_id = ?
                GROUP BY r.type ORDER BY r.type"""

            with dbh.dbhLock:
                dbh.dbh.execute(qry, [scan_id])
                scan_data = dbh.dbh.fetchall()
        except Exception as e:
            self.log.error(f"Grade calculation query failed: {e}")
            return {'enabled': True, 'overall_grade': '-', 'overall_score': 0, 'categories': {},
                    'error': 'Failed to retrieve scan results'}

        # Build event type count dict: {type_code: {total, unique, existed}}
        # 'total'/'unique' = unvalidated counts only (used for grading penalties)
        # 'existed' = True if ANY results exist (unvalidated, FP, or validated)
        event_type_counts = {}
        for row in scan_data:
            event_type = row[0]
            if event_type == 'ROOT':
                continue
            all_total = row[3]
            event_type_counts[event_type] = {
                'total': row[1],       # unvalidated total
                'unique': row[2],      # unvalidated unique
                'all_total': all_total,  # total including validated/FP
                'all_unique': row[4],  # unique including validated/FP
                'existed': all_total > 0,  # True if any results exist at all
            }

        # EXTERNAL_VULNERABILITIES: Read from tbl_scan_nessus_results (Nessus imports)
        # Only OPEN items (tracking = 0) count toward penalty.
        # But if any results exist at all, set existed=True so closing all
        # vulns removes the penalty rather than being treated as "never found".
        # Row format: [0]=id, [1]=severity, ..., [18]=tracking
        try:
            nessus_results = dbh.scanNessusList(scan_id)
            if nessus_results:
                ne_crit = sum(1 for r in nessus_results if r[1] == 'Critical' and (not r[18] or r[18] == 0))
                ne_high = sum(1 for r in nessus_results if r[1] == 'High' and (not r[18] or r[18] == 0))
                ne_med = sum(1 for r in nessus_results if r[1] == 'Medium' and (not r[18] or r[18] == 0))
                open_total = ne_crit + ne_high + ne_med
                event_type_counts['EXTERNAL_VULNERABILITIES'] = {
                    'total': open_total,
                    'unique': open_total,
                    'crit': ne_crit,
                    'high': ne_high,
                    'med': ne_med,
                    'existed': True,
                }
        except Exception as e:
            self.log.warning(f"Grade: failed to read Nessus results: {e}")

        # WEBAPP_VULNERABILITIES: Read from tbl_scan_burp_results (Burp imports)
        # Same logic: only OPEN items count, but existence is always tracked.
        # Row format: [0]=id, [1]=severity, ..., [18]=tracking
        try:
            burp_results = dbh.scanBurpList(scan_id)
            if burp_results:
                wa_high = sum(1 for r in burp_results if r[1] == 'High' and (not r[18] or r[18] == 0))
                wa_med = sum(1 for r in burp_results if r[1] == 'Medium' and (not r[18] or r[18] == 0))
                wa_low = sum(1 for r in burp_results if r[1] == 'Low' and (not r[18] or r[18] == 0))
                open_total = wa_high + wa_med + wa_low
                event_type_counts['WEBAPP_VULNERABILITIES'] = {
                    'total': open_total,
                    'unique': open_total,
                    'crit': 0,
                    'high': wa_high,
                    'med': wa_med,
                    'existed': True,
                }
        except Exception as e:
            self.log.warning(f"Grade: failed to read Burp results: {e}")

        # Log key event type counts for debugging grade issues
        for etype in ['DNS_SPF', 'DNS_TEXT']:
            if etype in event_type_counts:
                ec = event_type_counts[etype]
                self.log.debug(f"Grade: {etype} IN scan -> unique={ec['unique']} total={ec['total']} all_total={ec['all_total']} => no fail penalty")
            else:
                self.log.debug(f"Grade: {etype} NOT in scan results -> will get fail penalty")

        # Load config overrides from settings
        config_overrides = load_grade_config_overrides(self.config)

        # Calculate grade
        result = calculate_full_grade(event_type_counts, config_overrides)
        return result

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scangrade(self: 'SpiderFootWebUi', id: str) -> dict:
        """Calculate and return the overall grade for a scan.

        Returns a structured grade report with per-category breakdowns,
        individual event type scores, and an overall letter grade.

        Grade results are cached for 30 seconds to avoid expensive
        recalculation on every poll cycle during active scans.

        Args:
            id (str): scan ID

        Returns:
            dict: grade data with categories, overall_score, overall_grade
        """
        # Check cache (keyed by scan ID)
        now = time.time()
        cache = getattr(self, '_grade_cache', {})
        if id in cache:
            cached_time, cached_result = cache[id]
            if now - cached_time < 30:
                return cached_result

        try:
            dbh = SpiderFootDb(self.config)

            # Verify scan exists
            data = dbh.scanInstanceGet(id)
            if not data:
                return {'enabled': False, 'error': 'Scan not found'}

            result = self._calculateScanGrade(dbh, id)
        except Exception:
            return {'enabled': True, 'overall_grade': '-', 'overall_score': 0, 'categories': {},
                    'error': 'Temporarily unavailable'}

        # Store in cache
        if not hasattr(self, '_grade_cache'):
            self._grade_cache = {}
        self._grade_cache[id] = (now, result)

        return result

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanislatest(self: 'SpiderFootWebUi', id: str) -> dict:
        """Check if a scan is the latest scan for its target.

        Also triggers auto-import of entries from older scans if this is the
        latest scan and hasn't been imported yet.

        Args:
            id (str): scan ID

        Returns:
            dict: {isLatest: bool, scanCount: int, importedCount: int, importStatus: str}
        """
        dbh = SpiderFootDb(self.config)

        try:
            result = dbh.isLatestScan(id)

            # Auto-import: if this is the latest scan and there are multiple scans
            # and no entries have been imported yet, trigger import.
            # BUT skip if the scan is still running — the import holds a write lock
            # that blocks the scan subprocess from updating its status.
            scanInfo = dbh.scanInstanceGet(id)
            scanStatus = scanInfo[5] if scanInfo else None
            scanRunning = scanStatus not in (None, "FINISHED", "ABORTED", "ERROR-FAILED")

            if result['isLatest'] and result['scanCount'] > 1 and result['importedCount'] == 0 and not scanRunning:
                importResult = dbh.importEntriesFromOlderScans(id)
                result['importedCount'] = importResult['imported']
                result['importStatus'] = f"Imported {importResult['imported']} entries from previous scans"
                # Deduplicate after auto-import
                if importResult['imported'] > 0:
                    try:
                        dbh.deduplicateScanResults(id)
                    except Exception:
                        pass
            else:
                result['importStatus'] = 'already_imported' if result['importedCount'] > 0 else 'not_applicable'

            return result
        except Exception as e:
            return {'isLatest': False, 'scanCount': 0, 'importedCount': 0, 'importStatus': f'error: {str(e)}'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def triggerscanmimport(self: 'SpiderFootWebUi', id: str) -> dict:
        """Manually trigger import of entries from older scans.

        Args:
            id (str): scan ID

        Returns:
            dict: {success: bool, imported: int, skipped: int, message: str}
        """
        dbh = SpiderFootDb(self.config)

        try:
            # Check if this is a valid scan
            scanInfo = dbh.scanInstanceGet(id)
            if not scanInfo:
                return {'success': False, 'imported': 0, 'skipped': 0, 'message': 'Scan not found'}

            # Perform the import
            result = dbh.importEntriesFromOlderScans(id)

            # Deduplicate after import to remove any cross-scan duplicates
            dedup = dbh.deduplicateScanResults(id)

            return {
                'success': True,
                'imported': result['imported'],
                'skipped': result['skipped'],
                'dedup_removed': dedup['removed'],
                'message': (
                    f"Imported {result['imported']} entries, skipped {result['skipped']} duplicates. "
                    f"Deduplication removed {dedup['removed']} additional duplicate(s)."
                )
            }
        except Exception as e:
            return {'success': False, 'imported': 0, 'skipped': 0, 'message': f'Error: {str(e)}'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def resyncscanentries(self: 'SpiderFootWebUi', id: str) -> dict:
        """Delete imported entries and re-import fresh from older scans.

        Args:
            id (str): scan ID

        Returns:
            dict: {success: bool, deleted: int, imported: int, skipped: int, message: str}
        """
        dbh = SpiderFootDb(self.config)

        try:
            # Check if this is a valid scan
            scanInfo = dbh.scanInstanceGet(id)
            if not scanInfo:
                return {'success': False, 'deleted': 0, 'imported': 0, 'skipped': 0, 'message': 'Scan not found'}

            # Delete existing imported entries
            deleted = dbh.deleteImportedEntries(id)

            # Re-import from older scans
            result = dbh.importEntriesFromOlderScans(id)

            # Deduplicate after re-import
            dedup = dbh.deduplicateScanResults(id)

            return {
                'success': True,
                'deleted': deleted,
                'imported': result['imported'],
                'skipped': result['skipped'],
                'dedup_removed': dedup['removed'],
                'message': (
                    f"Deleted {deleted} old imports, imported {result['imported']} entries fresh. "
                    f"Deduplication removed {dedup['removed']} additional duplicate(s)."
                )
            }
        except Exception as e:
            return {'success': False, 'deleted': 0, 'imported': 0, 'skipped': 0, 'message': f'Error: {str(e)}'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def deduplicatescan(self: 'SpiderFootWebUi', id: str) -> dict:
        """Deduplicate scan results by removing entries with identical
        (Data Element, Source Data Element, Source Module) tuples.

        Keeps the oldest entry, preserves false_positive status.

        Args:
            id (str): scan ID

        Returns:
            dict: {success: bool, removed: int, fp_preserved: int, message: str}
        """
        dbh = SpiderFootDb(self.config)

        try:
            scanInfo = dbh.scanInstanceGet(id)
            if not scanInfo:
                return {'success': False, 'removed': 0, 'fp_preserved': 0,
                        'message': 'Scan not found'}

            result = dbh.deduplicateScanResults(id)

            return {
                'success': True,
                'removed': result['removed'],
                'fp_preserved': result['fp_preserved'],
                'message': (
                    f"Removed {result['removed']} duplicate(s). "
                    f"Preserved FP status on {result['fp_preserved']} kept row(s)."
                )
            }
        except Exception as e:
            self.log.error(f"Error deduplicating scan: {e}", exc_info=True)
            return {'success': False, 'removed': 0, 'fp_preserved': 0,
                    'message': f'Error: {str(e)}'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'SpiderFootWebUi', id: str) -> list:
        """Correlation results from a scan (synchronous, kept for backward compat).

        Args:
            id (str): scan ID

        Returns:
            list: correlation result list or error message
        """
        retdata = []
        dbh = SpiderFootDb(self.config)

        try:
            self.log.debug(f"Fetching correlations for scan {id}")
            corrdata = dbh.scanCorrelationList(id)
            self.log.debug(f"Found {len(corrdata)} correlations")

            if not corrdata:
                self.log.debug(f"No correlations found for scan {id}")
                return retdata

            for row in corrdata:
                # Check if we have a valid row of data
                if len(row) < 6:  # Need at least 6 elements to extract all required fields
                    self.log.error(
                        f"Correlation data format error: missing required fields, got {len(row)} fields")
                    continue

                # scanCorrelationList returns:
                #   0: c.id, 1: c.title, 2: c.rule_id, 3: c.rule_risk,
                #   4: c.rule_name, 5: c.rule_descr, 6: c.rule_logic,
                #   7: event_count, 8: event_types
                correlation_id = row[0]
                correlation = row[1]
                rule_id = row[2]
                rule_risk = row[3]
                rule_name = row[4]
                rule_description = row[5]
                event_count = row[7] if len(row) > 7 else 0
                event_types = row[8] if len(row) > 8 else ""

                retdata.append([correlation_id, correlation, rule_name, rule_risk,
                               rule_id, rule_description, event_count, "", event_types])

        except Exception as e:
            self.log.error(
                f"Error fetching correlations for scan {id}: {e}", exc_info=True)
            # Return empty list on error

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    @cherrypy.config(**{'tools.sessions.on': False, 'tools.auth_check.on': False})
    def scancorrelationsAsync(self: 'SpiderFootWebUi', id: str) -> dict:
        """Start loading correlation results in a background thread.

        Returns a job_id immediately. Poll correlationLoadStatus for progress.
        Sessions disabled to avoid FileSession lock contention with concurrent polls.

        Args:
            id (str): scan ID

        Returns:
            dict: {success, job_id}
        """
        self._cleanupOldCorrelationJobs()

        job_id = str(uuid.uuid4())
        with self._correlation_jobs_lock:
            self._correlation_jobs[job_id] = {
                'status': 'running',
                'progress': 0,
                'step': 'Warming up the correlation engine...',
                'result': None,
                'completed_at': 0,
            }

        worker = threading.Thread(
            target=self._loadCorrelationsWorker,
            args=(job_id, id),
            daemon=True
        )
        worker.start()

        return {'success': True, 'job_id': job_id}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    @cherrypy.config(**{'tools.sessions.on': False, 'tools.auth_check.on': False})
    def correlationLoadStatus(self: 'SpiderFootWebUi', job_id: str) -> dict:
        """Check status of an async correlation load job.

        Sessions disabled to avoid FileSession lock contention during polling.

        Args:
            job_id (str): job ID from scancorrelationsAsync

        Returns:
            dict: {success, status, progress, step, data}
        """
        with self._correlation_jobs_lock:
            job = self._correlation_jobs.get(job_id)
        if not job:
            return {'success': False, 'error': 'Job not found'}
        resp = {
            'success': True,
            'status': job['status'],
            'progress': job['progress'],
            'step': job['step'],
        }
        if job['status'] in ('complete', 'error'):
            resp['data'] = job.get('result')
        return resp

    def _loadCorrelationsWorker(self, job_id: str, scan_id: str) -> None:
        """Background worker that loads correlation data from the database."""
        try:
            self._updateCorrelationJob(job_id, progress=10, step='Querying the database...')

            dbh = SpiderFootDb(self.config)
            corrdata = dbh.scanCorrelationList(scan_id)

            self._updateCorrelationJob(job_id, progress=60, step='Processing results...')

            retdata = []
            if corrdata:
                for row in corrdata:
                    if len(row) < 6:
                        continue
                    # scanCorrelationList returns:
                    #   0: c.id, 1: c.title, 2: c.rule_id, 3: c.rule_risk,
                    #   4: c.rule_name, 5: c.rule_descr, 6: c.rule_logic,
                    #   7: event_count, 8: event_types
                    correlation_id = row[0]
                    correlation = row[1]
                    rule_id = row[2]
                    rule_risk = row[3]
                    rule_name = row[4]
                    rule_description = row[5]
                    event_count = row[7] if len(row) > 7 else 0
                    event_types = row[8] if len(row) > 8 else ""
                    retdata.append([correlation_id, correlation, rule_name, rule_risk,
                                   rule_id, rule_description, event_count, "", event_types])

            self._updateCorrelationJob(job_id, progress=90, step='Preparing display...')

            self._updateCorrelationJob(
                job_id,
                status='complete',
                progress=100,
                step='Complete',
                result=retdata,
                completed_at=time.time()
            )

        except Exception as e:
            self.log.error(f"Correlation load worker error: {e}", exc_info=True)
            self._updateCorrelationJob(
                job_id,
                status='error',
                progress=100,
                step='Error',
                result=[],
                completed_at=time.time()
            )

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanfindings(self: 'SpiderFootWebUi', id: str) -> list:
        """Get findings for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: findings data
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanFindingsList(id)
            return [list(row) for row in data]
        except Exception as e:
            self.log.error(f"Error fetching findings for scan {id}: {e}", exc_info=True)
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanfindingscount(self: 'SpiderFootWebUi', id: str) -> dict:
        """Get count of findings for a scan.

        Args:
            id (str): scan ID

        Returns:
            dict: count of findings
        """
        dbh = SpiderFootDb(self.config)

        try:
            count = dbh.scanFindingsCount(id)
            return {'success': True, 'count': count}
        except Exception as e:
            self.log.error(f"Error counting findings for scan {id}: {e}", exc_info=True)
            return {'success': False, 'count': 0}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def importfindings(self: 'SpiderFootWebUi', id: str = None, importfile=None) -> dict:
        """Import findings from an Excel file (.xlsx).

        Args:
            id (str): scan instance ID
            importfile: uploaded .xlsx file

        Returns:
            dict: import results
        """
        if not id:
            return {'success': False, 'message': 'No scan ID provided.'}

        if importfile is None:
            return {'success': False, 'message': 'No file was uploaded.'}

        try:
            raw = importfile.file.read()
        except Exception as e:
            return {'success': False, 'message': f'Failed to read uploaded file: {e}'}

        if not raw:
            return {'success': False, 'message': 'Uploaded file is empty.'}

        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active

            # Read headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value or '').strip().lower())

            # Map expected columns
            expected = ['priority', 'category', 'tab', 'item', 'description', 'recommendation']
            col_map = {}
            for col_name in expected:
                if col_name in headers:
                    col_map[col_name] = headers.index(col_name)

            if 'priority' not in col_map:
                return {'success': False, 'message': f'Required "Priority" column not found. Found columns: {", ".join(headers)}'}

            findings = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                # Skip completely empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                finding = {}
                for col_name, col_idx in col_map.items():
                    if col_idx < len(row):
                        finding[col_name] = str(row[col_idx] or '').strip()
                    else:
                        finding[col_name] = ''

                if finding.get('priority'):
                    findings.append(finding)

            wb.close()

            if not findings:
                return {'success': False, 'message': 'No valid findings found in the Excel file.'}

            dbh = SpiderFootDb(self.config)
            count = dbh.scanFindingsStore(id, findings)

            return {
                'success': True,
                'count': count,
                'message': f'Successfully imported {count} findings.'
            }

        except Exception as e:
            self.log.error(f"Error importing findings: {e}", exc_info=True)
            return {'success': False, 'message': f'Error processing Excel file: {e}'}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scannessuscount(self: 'SpiderFootWebUi', id: str) -> dict:
        """Get count of Nessus results for a scan.

        Args:
            id (str): scan ID

        Returns:
            dict: count of Nessus results
        """
        dbh = SpiderFootDb(self.config)

        try:
            count = dbh.scanNessusCount(id)
            return {'success': True, 'count': count}
        except Exception as e:
            self.log.error(f"Error counting Nessus results for scan {id}: {e}", exc_info=True)
            return {'success': False, 'count': 0}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scannessuslist(self: 'SpiderFootWebUi', id: str) -> dict:
        """Get list of Nessus results for a scan.

        Args:
            id (str): scan ID

        Returns:
            dict: Nessus results
        """
        dbh = SpiderFootDb(self.config)

        try:
            rows = dbh.scanNessusList(id)
            results = []
            for row in rows:
                results.append({
                    'id': row[0],
                    'severity': row[1],
                    'severity_number': row[2],
                    'plugin_name': row[3],
                    'plugin_id': row[4],
                    'host_ip': row[5],
                    'host_name': row[6],
                    'operating_system': row[7],
                    'description': row[8],
                    'synopsis': row[9],
                    'solution': row[10],
                    'see_also': row[11],
                    'service_name': row[12],
                    'port': row[13],
                    'protocol': row[14],
                    'request': row[15],
                    'plugin_output': row[16],
                    'cvss3_base_score': row[17],
                    'tracking': row[18],
                })
            return {'success': True, 'results': results}
        except Exception as e:
            self.log.error(f"Error listing Nessus results for scan {id}: {e}", exc_info=True)
            return {'success': False, 'results': []}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def importnessus(self: 'SpiderFootWebUi', id: str = None, importfile=None) -> dict:
        """Import Nessus results from a .nessus file into an existing scan.

        Args:
            id (str): scan instance ID
            importfile: uploaded .nessus file

        Returns:
            dict: import results
        """
        if not id:
            return {'success': False, 'message': 'No scan ID provided.'}

        if importfile is None:
            return {'success': False, 'message': 'No file was uploaded.'}

        try:
            raw = importfile.file.read()
            content = raw.decode('utf-8', errors='replace')
        except Exception as e:
            return {'success': False, 'message': f'Failed to read uploaded file: {e}'}

        if not content.strip():
            return {'success': False, 'message': 'Uploaded file is empty.'}

        return self._processNessusImport(content, None, None, importfile,
                                          is_dry_run=False, existing_scan_id=id)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanburpcount(self: 'SpiderFootWebUi', id: str) -> dict:
        """Get count of Burp results for a scan.

        Args:
            id (str): scan ID

        Returns:
            dict: count of Burp results
        """
        dbh = SpiderFootDb(self.config)

        try:
            count = dbh.scanBurpCount(id)
            return {'success': True, 'count': count}
        except Exception as e:
            self.log.error(f"Error counting Burp results for scan {id}: {e}", exc_info=True)
            return {'success': False, 'count': 0}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanburpenhanced(self: 'SpiderFootWebUi', id: str) -> dict:
        """Check if Burp results have been enhanced with HTML report data.

        Args:
            id (str): scan ID

        Returns:
            dict: whether HTML enhancement has been applied
        """
        dbh = SpiderFootDb(self.config)

        try:
            enhanced = dbh.scanBurpEnhanced(id)
            return {'success': True, 'enhanced': enhanced}
        except Exception as e:
            self.log.error(f"Error checking Burp enhanced state for scan {id}: {e}", exc_info=True)
            return {'success': False, 'enhanced': False}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanburplist(self: 'SpiderFootWebUi', id: str) -> dict:
        """Get list of Burp results for a scan.

        Args:
            id (str): scan ID

        Returns:
            dict: Burp results
        """
        dbh = SpiderFootDb(self.config)

        try:
            rows = dbh.scanBurpList(id)
            results = []
            for row in rows:
                results.append({
                    'id': row[0],
                    'severity': row[1],
                    'severity_number': row[2],
                    'host_ip': row[3],
                    'host_name': row[4],
                    'plugin_name': row[5],
                    'issue_type': row[6],
                    'path': row[7],
                    'location': row[8],
                    'confidence': row[9],
                    'issue_background': row[10],
                    'issue_detail': row[11],
                    'solutions': row[12],
                    'see_also': row[13],
                    'references': row[14],
                    'vulnerability_classifications': row[15],
                    'request': row[16],
                    'response': row[17],
                    'tracking': row[18],
                })
            return {'success': True, 'results': results}
        except Exception as e:
            self.log.error(f"Error listing Burp results for scan {id}: {e}", exc_info=True)
            return {'success': False, 'results': []}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def importburp(self: 'SpiderFootWebUi', id: str = None, importfile=None) -> dict:
        """Import Burp results from an XML file into an existing scan.

        Args:
            id (str): scan instance ID
            importfile: uploaded .xml file

        Returns:
            dict: import results
        """
        if not id:
            return {'success': False, 'message': 'No scan ID provided.'}

        if importfile is None:
            return {'success': False, 'message': 'No file was uploaded.'}

        try:
            raw = importfile.file.read()
            content = raw.decode('utf-8', errors='replace')
        except Exception as e:
            return {'success': False, 'message': f'Failed to read uploaded file: {e}'}

        if not content.strip():
            return {'success': False, 'message': 'Uploaded file is empty.'}

        return self._processBurpImport(content, None, None, importfile,
                                        is_dry_run=False, existing_scan_id=id)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def importburphtml(self: 'SpiderFootWebUi', id: str = None, importfile=None) -> dict:
        """Enhance existing Burp results with data from an HTML report.

        Parses the HTML report and merges additional details (issue_detail,
        issue_background, solutions, references, etc.) into existing Burp
        XML results matched by plugin_name.

        Args:
            id (str): scan instance ID (must already have Burp XML data)
            importfile: uploaded .html file

        Returns:
            dict: enhance results
        """
        if not id:
            return {'success': False, 'message': 'No scan ID provided.'}

        if importfile is None:
            return {'success': False, 'message': 'No file was uploaded.'}

        try:
            raw = importfile.file.read()
            content = raw.decode('utf-8', errors='replace')
        except Exception as e:
            return {'success': False, 'message': f'Failed to read uploaded file: {e}'}

        if not content.strip():
            return {'success': False, 'message': 'Uploaded file is empty.'}

        return self._processBurpHtmlEnhance(content, None, None, importfile,
                                            is_dry_run=False, existing_scan_id=id)

    @cherrypy.expose
    def scanvulnsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "xlsx") -> bytes:
        """Export vulnerability scan results (Nessus and Burp).

        For CSV: returns a zip archive containing EXT-VULNS.csv and WEBAPP-VULNS.csv.
        For XLSX/Excel: returns a single VULNS.xlsx with two tabs (EXT-VULNS, WEBAPP-VULNS).

        Args:
            id (str): scan instance ID
            filetype (str): export format (xlsx, excel, csv)

        Returns:
            bytes: exported file data
        """
        dbh = SpiderFootDb(self.config)
        scan = dbh.scanInstanceGet(id)
        _scan_name = scan[0] if scan else ''

        # --- Nessus data ---
        nessus_headers = [
            "Severity", "Severity Number", "Plugin Name", "Plugin ID",
            "Host IP", "Host Name", "Operating System", "Description",
            "Synopsis", "Solution", "See Also", "Service Name", "Port",
            "Protocol", "Request", "Plugin Output", "CVSS3 Base Score", "Tracking"
        ]
        nessus_rows = []
        try:
            rows = dbh.scanNessusList(id)
            tracking_labels = {0: 'OPEN', 1: 'CLOSED', 2: 'TICKETED'}
            for row in rows:
                nessus_rows.append([
                    str(row[1] or ''),   # severity
                    str(row[2] or ''),   # severity_number
                    str(row[3] or ''),   # plugin_name
                    str(row[4] or ''),   # plugin_id
                    str(row[5] or ''),   # host_ip
                    str(row[6] or ''),   # host_name
                    str(row[7] or ''),   # operating_system
                    str(row[8] or ''),   # description
                    str(row[9] or ''),   # synopsis
                    str(row[10] or ''),  # solution
                    str(row[11] or ''),  # see_also
                    str(row[12] or ''),  # service_name
                    str(row[13] or ''),  # port
                    str(row[14] or ''),  # protocol
                    str(row[15] or ''),  # request
                    str(row[16] or ''),  # plugin_output
                    str(row[17] or ''),  # cvss3_base_score
                    tracking_labels.get(int(row[18] or 0), 'OPEN'),  # tracking
                ])
        except Exception:
            pass

        # --- Burp data ---
        burp_headers = [
            "Severity", "Severity Number", "Host IP", "Host Name",
            "Plugin Name", "Issue Type", "Path", "Location", "Confidence",
            "Issue Background", "Issue Detail", "Solutions", "See Also",
            "References", "Vulnerability Classifications",
            "Request", "Response", "Tracking"
        ]
        burp_rows = []
        try:
            rows = dbh.scanBurpList(id)
            tracking_labels = {0: 'OPEN', 1: 'CLOSED', 2: 'TICKETED'}
            for row in rows:
                burp_rows.append([
                    str(row[1] or ''),   # severity
                    str(row[2] or ''),   # severity_number
                    str(row[3] or ''),   # host_ip
                    str(row[4] or ''),   # host_name
                    str(row[5] or ''),   # plugin_name
                    str(row[6] or ''),   # issue_type
                    str(row[7] or ''),   # path
                    str(row[8] or ''),   # location
                    str(row[9] or ''),   # confidence
                    str(row[10] or ''),  # issue_background
                    str(row[11] or ''),  # issue_detail
                    str(row[12] or ''),  # solutions
                    str(row[13] or ''),  # see_also
                    str(row[14] or ''),  # reference_links
                    str(row[15] or ''),  # vulnerability_classifications
                    str(row[16] or ''),  # request
                    str(row[17] or ''),  # response
                    tracking_labels.get(int(row[18] or 0), 'OPEN'),  # tracking
                ])
        except Exception:
            pass

        if filetype.lower() in ["xlsx", "excel"]:
            fname = self._export_filename(_scan_name, id, 'VULNS', 'xlsx')
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"

            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning)
                wb = openpyxl.Workbook()

                # EXT-VULNS sheet (Nessus) — styled with severity colors
                ws_nessus = wb.active
                ws_nessus.title = "EXT-VULNS"
                build_nessus_sheet(ws_nessus, nessus_rows)

                # WEBAPP-VULNS sheet (Burp) — styled with severity colors
                ws_burp = wb.create_sheet("WEBAPP-VULNS")
                build_burp_sheet(ws_burp, burp_rows)

            with BytesIO() as f:
                wb.save(f)
                f.seek(0)
                return f.read()

        if filetype.lower() == 'csv':
            import csv
            import zipfile
            from io import StringIO, BytesIO

            # Build EXT-VULNS.csv
            nessus_buf = StringIO()
            nessus_writer = csv.writer(nessus_buf, dialect='excel')
            nessus_writer.writerow(nessus_headers)
            for row in nessus_rows:
                nessus_writer.writerow(row)
            nessus_csv = nessus_buf.getvalue()

            # Build WEBAPP-VULNS.csv
            burp_buf = StringIO()
            burp_writer = csv.writer(burp_buf, dialect='excel')
            burp_writer.writerow(burp_headers)
            for row in burp_rows:
                burp_writer.writerow(row)
            burp_csv = burp_buf.getvalue()

            # Package into zip
            zip_buf = BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                zf.writestr('EXT-VULNS.csv', nessus_csv)
                zf.writestr('WEBAPP-VULNS.csv', burp_csv)
            zip_buf.seek(0)

            fname = self._export_filename(_scan_name, id, 'VULNS', 'zip')
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename={fname}'
            cherrypy.response.headers['Content-Type'] = "application/zip"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return zip_buf.read()

        return self.error("Invalid export file type.")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scannessussettracking(self: 'SpiderFootWebUi', id: str, resultId: str, tracking: str) -> dict:
        """Update tracking status for a Nessus result.

        Args:
            id (str): scan instance ID
            resultId (str): result row ID
            tracking (str): 0=OPEN, 1=CLOSED, 2=TICKETED

        Returns:
            dict: success status
        """
        dbh = SpiderFootDb(self.config)

        try:
            dbh.scanNessusUpdateTracking(id, int(resultId), int(tracking))
            return {'success': True}
        except Exception as e:
            self.log.error(f"Error updating Nessus tracking for scan {id}, result {resultId}: {e}", exc_info=True)
            return {'success': False, 'message': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanburpsettracking(self: 'SpiderFootWebUi', id: str, resultId: str, tracking: str) -> dict:
        """Update tracking status for a Burp result.

        Args:
            id (str): scan instance ID
            resultId (str): result row ID
            tracking (str): 0=OPEN, 1=CLOSED, 2=TICKETED

        Returns:
            dict: success status
        """
        dbh = SpiderFootDb(self.config)

        try:
            dbh.scanBurpUpdateTracking(id, int(resultId), int(tracking))
            return {'success': True}
        except Exception as e:
            self.log.error(f"Error updating Burp tracking for scan {id}, result {resultId}: {e}", exc_info=True)
            return {'success': False, 'message': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventsettracking(self: 'SpiderFootWebUi', id: str, resultHash: str, tracking: str) -> dict:
        """Update tracking status for a scan result event.

        Args:
            id (str): scan instance ID
            resultHash (str): event hash
            tracking (str): 0=OPEN, 1=CLOSED, 2=TICKETED

        Returns:
            dict: success status
        """
        dbh = SpiderFootDb(self.config)

        try:
            dbh.scanEventUpdateTracking(id, resultHash, int(tracking))

            # Get event details for cross-scan sync
            eventDetails = dbh.scanEventResultByHash(id, resultHash)
            if eventDetails:
                scanInfo = dbh.scanInstanceGet(id)
                target = scanInfo[1] if scanInfo else None
                if target:
                    dbh.syncTrackingAcrossScans(
                        target, eventDetails[0], eventDetails[1],
                        eventDetails[2], int(tracking)
                    )

            return {'success': True}
        except Exception as e:
            self.log.error(f"Error updating event tracking for scan {id}, hash {resultHash}: {e}", exc_info=True)
            return {'success': False, 'message': str(e)}

    @cherrypy.expose
    def scanfindingsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "xlsx", report: str = "basic") -> str:
        """Export findings and correlations from a scan.

        Args:
            id (str): scan ID
            filetype (str): export format (xlsx, csv)
            report (str): 'basic' for simple two-sheet export,
                          'full' for styled report with Executive Summary + category tabs

        Returns:
            str: exported data
        """
        dbh = SpiderFootDb(self.config)
        _scan = dbh.scanInstanceGet(id)
        _scan_name = _scan[0] if _scan else ''

        # Get findings
        findings_rows = []
        try:
            findings_data = dbh.scanFindingsList(id)
            for row in findings_data:
                findings_rows.append([
                    str(row[1]),  # Priority
                    str(row[2]),  # Category
                    str(row[3]),  # Tab
                    str(row[4]),  # Item
                    str(row[5]),  # Description
                    str(row[6]),  # Recommendation
                ])
        except Exception:
            pass

        # Get correlations
        correlation_rows = []
        try:
            corr_data = dbh.scanCorrelationList(id)
            for corr_row in corr_data:
                correlation_rows.append([
                    str(corr_row[1]),   # Title
                    str(corr_row[4]),   # Rule Name
                    str(corr_row[3]),   # Risk
                    str(corr_row[5]),   # Description
                    str(corr_row[6]),   # Rule Logic
                    str(corr_row[7]),   # Event Count
                    str(corr_row[8] or ''),  # Event Types
                ])
        except Exception:
            pass

        # Sort correlations by severity: CRITICAL → HIGH → MEDIUM → LOW → INFO
        _risk_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'INFO': 4}
        correlation_rows.sort(key=lambda r: _risk_order.get(str(r[2]).upper().strip(), 5))

        if filetype.lower() in ["xlsx", "excel"]:
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"

            wb = None

            # Full report: styled with Executive Summary + category tabs
            if report == "full":
                cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={self._export_filename(_scan_name, id, 'REPORT', 'xlsx')}"
                with warnings.catch_warnings():
                    warnings.filterwarnings('ignore', category=UserWarning)

                    wb = openpyxl.Workbook()

                    # Gather grade data (isolated error handling -- grade failure
                    # should not prevent the entire report from generating)
                    self.log.info(f"Full report: calculating grade for scan {id}")
                    try:
                        grade_data = self._calculateScanGrade(dbh, id)
                    except Exception as e:
                        self.log.error(f"Full report: grade calculation failed: {e}", exc_info=True)
                        grade_data = {
                            'enabled': False, 'overall_grade': '-', 'overall_score': 0,
                            'overall_grade_color': '#6b7280', 'overall_grade_bg': '#f3f4f6',
                            'categories': {},
                        }

                    # Gather scan metadata
                    try:
                        scan_instance = dbh.scanInstanceGet(id)
                    except Exception as e:
                        self.log.error(f"Full report: scan instance lookup failed: {e}", exc_info=True)
                        scan_instance = None

                    scan_info = {
                        'name': scan_instance[0] if scan_instance else 'Unknown',
                        'target': scan_instance[1] if scan_instance else 'Unknown',
                        'date': time.strftime(
                            "%Y-%m-%d %H:%M:%S",
                            time.localtime(scan_instance[2]),
                        ) if scan_instance and scan_instance[2] else '',
                    }

                    # Sheet 1: Executive Summary (reuse the default active sheet)
                    self.log.info("Full report: building Executive Summary")
                    ws_summary = wb.active
                    ws_summary.title = "Executive Summary"
                    build_executive_summary(ws_summary, grade_data, scan_info,
                                            findings_rows=findings_rows,
                                            correlation_rows=correlation_rows)

                    # Sheet 2: Findings (black tab, severity colors)
                    self.log.info(f"Full report: building Findings sheet ({len(findings_rows)} rows)")
                    ws_findings = wb.create_sheet("Findings")
                    build_findings_sheet(ws_findings, findings_rows)

                    # Sheet 3: Correlations (dark gray tab, risk colors)
                    self.log.info(f"Full report: building Correlations sheet ({len(correlation_rows)} rows)")
                    ws_corr = wb.create_sheet("Correlations")
                    build_correlations_sheet(ws_corr, correlation_rows)

                    # Build category weight ordering for event-type tab sorting
                    cat_results = grade_data.get('categories', {})
                    _cat_weight_order = {
                        cat_name: (-cat_data.get('weight', 0), cat_name)
                        for cat_name, cat_data in cat_results.items()
                    }

                    used_names = {'Executive Summary', 'Findings', 'Correlations'}

                    # Sheet 4: EXT-VULNS (Nessus) - red tab
                    try:
                        self.log.info(f"Full report: fetching Nessus data for EXT-VULNS tab")
                        nessus_data = dbh.scanNessusList(id)
                        nessus_rows = []
                        tracking_labels = {0: 'OPEN', 1: 'CLOSED', 2: 'TICKETED'}
                        for row in nessus_data:
                            nessus_rows.append([
                                str(row[1] or ''),   # severity
                                str(row[2] or ''),   # severity_number
                                str(row[3] or ''),   # plugin_name
                                str(row[4] or ''),   # plugin_id
                                str(row[5] or ''),   # host_ip
                                str(row[6] or ''),   # host_name
                                str(row[7] or ''),   # operating_system
                                str(row[8] or ''),   # description
                                str(row[9] or ''),   # synopsis
                                str(row[10] or ''),  # solution
                                str(row[11] or ''),  # see_also
                                str(row[12] or ''),  # service_name
                                str(row[13] or ''),  # port
                                str(row[14] or ''),  # protocol
                                str(row[15] or ''),  # request
                                str(row[16] or ''),  # plugin_output
                                str(row[17] or ''),  # cvss3_base_score
                                tracking_labels.get(int(row[18] or 0), 'OPEN'),  # tracking
                            ])
                        self.log.info(f"Full report: building EXT-VULNS sheet ({len(nessus_rows)} rows)")
                        ws_nessus = wb.create_sheet("EXT-VULNS")
                        used_names.add("EXT-VULNS")
                        build_nessus_sheet(ws_nessus, nessus_rows)
                    except Exception as e:
                        self.log.error(f"Full report: EXT-VULNS tab failed: {e}", exc_info=True)

                    # Pre-fetch WEBAPP-VULNS (Burp) data -- sheet created later
                    # in the event-type loop to position it at the Web App Security boundary
                    burp_rows = []
                    try:
                        self.log.info(f"Full report: fetching Burp data for WEBAPP-VULNS tab")
                        burp_data = dbh.scanBurpList(id)
                        tracking_labels = {0: 'OPEN', 1: 'CLOSED', 2: 'TICKETED'}
                        for row in burp_data:
                            burp_rows.append([
                                str(row[1] or ''),   # severity
                                str(row[2] or ''),   # severity_number
                                str(row[3] or ''),   # host_ip
                                str(row[4] or ''),   # host_name
                                str(row[5] or ''),   # plugin_name
                                str(row[6] or ''),   # issue_type
                                str(row[7] or ''),   # path
                                str(row[8] or ''),   # location
                                str(row[9] or ''),   # confidence
                                str(row[10] or ''),  # issue_background
                                str(row[11] or ''),  # issue_detail
                                str(row[12] or ''),  # solutions
                                str(row[13] or ''),  # see_also
                                str(row[14] or ''),  # reference_links
                                str(row[15] or ''),  # vulnerability_classifications
                                str(row[16] or ''),  # request
                                str(row[17] or ''),  # response
                                tracking_labels.get(int(row[18] or 0), 'OPEN'),  # tracking
                            ])
                        self.log.info(f"Full report: WEBAPP-VULNS data ready ({len(burp_rows)} rows)")
                    except Exception as e:
                        self.log.error(f"Full report: WEBAPP-VULNS data fetch failed: {e}", exc_info=True)

                    # Event-type data tabs (one sheet per event type, grouped by category weight order)
                    try:
                        self.log.info(f"Full report: fetching scan result events for data tabs")
                        scan_data = dbh.scanResultEvent(id, 'ALL')

                        # Get target-level FPs
                        target = scan_instance[1] if scan_instance else None
                        target_fps = set()
                        if target:
                            try:
                                target_fps = dbh.targetFalsePositivesForTarget(target)
                            except Exception:
                                pass

                        # Group rows by event type
                        event_type_groups = {}
                        for row in scan_data:
                            if row[4] == "ROOT":
                                continue
                            event_type_code = str(row[4])
                            fp_flag = self._compute_fp_flag(row[13], row[4], row[1], row[2], target_fps)
                            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                            datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                            display_type = translate_event_type(event_type_code)
                            if display_type not in event_type_groups:
                                event_type_groups[display_type] = {
                                    'rows': [],
                                    'event_code': event_type_code,
                                }
                            event_type_groups[display_type]['rows'].append(
                                [lastseen, str(row[3]), str(row[2]), fp_flag, datafield])

                        # Sort event types by category weight (descending), then alphabetically within category
                        def _evt_sort_key(display_type):
                            grp = event_type_groups[display_type]
                            evt_grading = get_event_grading(grp['event_code'])
                            evt_category = evt_grading.get('category', 'Information / Reference')
                            # Use category weight order, falling back to (0, category_name)
                            cat_order = _cat_weight_order.get(evt_category, (0, evt_category))
                            return (cat_order[0], cat_order[1], display_type)

                        sorted_event_types = sorted(event_type_groups.keys(), key=_evt_sort_key)

                        self.log.info(f"Full report: building {len(sorted_event_types)} event-type data tabs")
                        _webapp_vulns_inserted = False
                        for display_type in sorted_event_types:
                            grp = event_type_groups[display_type]
                            evt_code = grp['event_code']
                            evt_rows = grp['rows']

                            # Look up category color via grade config
                            evt_grading = get_event_grading(evt_code)
                            evt_category = evt_grading.get('category', 'Information / Reference')
                            tab_color = CATEGORY_TAB_COLORS.get(evt_category, '#6b7280')

                            # Insert WEBAPP-VULNS right before the first Web App Security event type
                            if not _webapp_vulns_inserted and evt_category == 'Web App Security':
                                self.log.info(f"Full report: inserting WEBAPP-VULNS sheet ({len(burp_rows)} rows)")
                                ws_burp = wb.create_sheet("WEBAPP-VULNS")
                                used_names.add("WEBAPP-VULNS")
                                build_burp_sheet(ws_burp, burp_rows)
                                _webapp_vulns_inserted = True

                            safe_name = sanitize_sheet_name(display_type)
                            original = safe_name
                            suffix = 2
                            while safe_name in used_names:
                                safe_name = f"{original[:28]}({suffix})"
                                suffix += 1
                            used_names.add(safe_name)

                            ws_evt = wb.create_sheet(safe_name)
                            build_event_type_sheet(ws_evt, display_type, evt_rows, tab_color=tab_color)

                        # Fallback: if no Web App Security events, append WEBAPP-VULNS at end
                        if not _webapp_vulns_inserted:
                            self.log.info(f"Full report: appending WEBAPP-VULNS sheet ({len(burp_rows)} rows)")
                            ws_burp = wb.create_sheet("WEBAPP-VULNS")
                            used_names.add("WEBAPP-VULNS")
                            build_burp_sheet(ws_burp, burp_rows)

                    except Exception as e:
                        self.log.error(f"Full report: event-type data tabs failed: {e}", exc_info=True)

                    self.log.info("Full report: workbook built successfully")

            # Basic export (default) or fallback if full report failed
            if wb is None:
                if report != "full":
                    cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={self._export_filename(_scan_name, id, 'FINDINGS', 'xlsx')}"
                with warnings.catch_warnings():
                    warnings.filterwarnings('ignore', category=UserWarning)
                    wb = openpyxl.Workbook()

                    ws_findings = wb.active
                    ws_findings.title = "Findings"
                    build_findings_sheet(ws_findings, findings_rows)

                    ws_corr = wb.create_sheet("Correlations")
                    build_correlations_sheet(ws_corr, correlation_rows)

            try:
                self.log.info("Full report: saving workbook...")
                with BytesIO() as f:
                    wb.save(f)
                    f.seek(0)
                    data = f.read()
                self.log.info(f"Full report: saved successfully ({len(data)} bytes)")
                return data
            except Exception as e:
                self.log.error(f"Full report: wb.save() failed: {e}", exc_info=True)
                # Fall back to basic export so the user gets something
                wb_fallback = openpyxl.Workbook()
                ws_err = wb_fallback.active
                ws_err.title = "Export Error"
                ws_err['A1'].value = "Full report save failed during Excel serialization."
                ws_err['A2'].value = f"Error: {str(e)}"
                ws_err['A3'].value = "The styled workbook was built but could not be saved. Basic data follows."

                ws_fb_findings = wb_fallback.create_sheet("Findings")
                fb_headers = ["Priority", "Category", "Tab", "Item", "Description", "Recommendation"]
                for col_num, header in enumerate(fb_headers, 1):
                    ws_fb_findings.cell(row=1, column=col_num, value=header)
                for row_num, row_data in enumerate(findings_rows, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws_fb_findings.cell(row=row_num, column=col_num, value=_safe_str(cell_value))

                ws_fb_corr = wb_fallback.create_sheet("Correlations")
                fb_corr_headers = ["Correlation", "Rule Name", "Risk", "Description", "Rule Logic", "Event Count", "Event Types"]
                for col_num, header in enumerate(fb_corr_headers, 1):
                    ws_fb_corr.cell(row=1, column=col_num, value=header)
                for row_num, row_data in enumerate(correlation_rows, 2):
                    for col_num, cell_value in enumerate(row_data, 1):
                        ws_fb_corr.cell(row=row_num, column=col_num, value=_safe_str(cell_value))

                with BytesIO() as f:
                    wb_fallback.save(f)
                    f.seek(0)
                    return f.read()

        if filetype.lower() == 'csv':
            import csv
            import zipfile

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={self._export_filename(_scan_name, id, 'FINDINGS', 'zip')}"
            cherrypy.response.headers['Content-Type'] = "application/zip"
            cherrypy.response.headers['Pragma'] = "no-cache"

            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                # FINDINGS.csv
                findings_io = StringIO()
                findings_writer = csv.writer(findings_io, dialect='excel')
                findings_writer.writerow(["Priority", "Category", "Tab", "Item", "Description", "Recommendation"])
                for row in findings_rows:
                    findings_writer.writerow(row)
                zf.writestr("FINDINGS.csv", findings_io.getvalue())

                # CORRELATIONS.csv
                corr_io = StringIO()
                corr_writer = csv.writer(corr_io, dialect='excel')
                corr_writer.writerow(["Correlation", "Rule Name", "Risk", "Description", "Rule Logic", "Event Count", "Event Types"])
                for row in correlation_rows:
                    corr_writer.writerow(row)
                zf.writestr("CORRELATIONS.csv", corr_io.getvalue())

            zip_buffer.seek(0)
            return zip_buffer.read()

        return self.error("Invalid export file type.")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def checkAiCorrelationModes(self: 'SpiderFootWebUi', id: str) -> dict:
        """Check which correlation modes are available for a scan.

        Returns whether the scan has imported data (enabling cross-scan mode).

        Args:
            id (str): scan ID

        Returns:
            dict: available modes and event counts
        """
        dbh = SpiderFootDb(self.config)

        scanInfo = dbh.scanInstanceGet(id)
        if not scanInfo:
            return {'success': False, 'error': 'Scan not found'}

        try:
            qry = """SELECT
                        COUNT(CASE WHEN imported_from_scan IS NULL THEN 1 END) as native_count,
                        COUNT(CASE WHEN imported_from_scan IS NOT NULL THEN 1 END) as imported_count
                     FROM tbl_scan_results
                     WHERE scan_instance_id = ?
                       AND type != 'ROOT'
                       AND type != 'AI_CROSS_SCAN_CORRELATION'
                       AND type != 'AI_SINGLE_SCAN_CORRELATION'
                       AND false_positive = 0"""

            with dbh.dbhLock:
                dbh.dbh.execute(qry, [id])
                row = dbh.dbh.fetchone()

            native_count = row[0] if row else 0
            imported_count = row[1] if row else 0

            return {
                'success': True,
                'native_count': native_count,
                'imported_count': imported_count,
                'has_imported': imported_count > 0,
                'single_scan_available': native_count > 0,
                'cross_scan_available': imported_count > 0 and native_count > 0
            }
        except Exception as e:
            self.log.error(f"Error checking correlation modes: {e}", exc_info=True)
            return {'success': False, 'error': str(e)}

    def _cleanupOldCorrelationJobs(self) -> None:
        """Remove completed correlation jobs older than 10 minutes."""
        cutoff = time.time() - 600
        with self._correlation_jobs_lock:
            expired = [jid for jid, job in self._correlation_jobs.items()
                       if job.get('completed_at', 0) and job['completed_at'] < cutoff]
            for jid in expired:
                del self._correlation_jobs[jid]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    @cherrypy.config(**{'tools.sessions.on': False, 'tools.auth_check.on': False})
    def correlationStatus(self: 'SpiderFootWebUi', job_id: str) -> dict:
        """Check the status of a running correlation job.

        Sessions disabled to avoid FileSession lock contention during polling.

        Args:
            job_id (str): correlation job ID returned by runAiCorrelation

        Returns:
            dict: job status with progress percentage and current step
        """
        with self._correlation_jobs_lock:
            job = self._correlation_jobs.get(job_id)
        if not job:
            return {'success': False, 'error': 'Job not found'}
        return {
            'success': True,
            'status': job['status'],
            'progress': job['progress'],
            'step': job['step'],
            'result': job.get('result')
        }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def runAiCorrelation(self: 'SpiderFootWebUi', id: str, mode: str = 'cross') -> dict:
        """Run correlation analysis on a scan (async with progress tracking).

        Launches the correlation work in a background thread and returns
        immediately with a job_id. Use correlationStatus to poll progress.

        Args:
            id (str): scan ID
            mode (str): 'cross' for cross-scan correlation (requires imported data),
                         'single' for single-scan correlation (current scan only)

        Returns:
            dict: {success, job_id, status} for async polling
        """
        if mode not in ('cross', 'single'):
            return {'success': False, 'error': f'Invalid mode: {mode}'}

        dbh = SpiderFootDb(self.config)

        # Get scan info
        scanInfo = dbh.scanInstanceGet(id)
        if not scanInfo:
            return {'success': False, 'error': 'Scan not found'}

        # Clean up old completed jobs
        self._cleanupOldCorrelationJobs()

        # Create a new job
        job_id = str(uuid.uuid4())
        with self._correlation_jobs_lock:
            self._correlation_jobs[job_id] = {
                'status': 'running',
                'progress': 0,
                'step': 'Initializing...',
                'result': None,
                'completed_at': 0,
            }

        target = scanInfo[1]
        self.log.info(f"Running {mode} correlation analysis for scan {id} (target: {target}) [job {job_id}]")

        # Launch background worker thread
        worker = threading.Thread(
            target=self._correlationWorker,
            args=(job_id, id, target, mode),
            daemon=True
        )
        worker.start()

        return {'success': True, 'job_id': job_id, 'status': 'running'}

    def _updateCorrelationJob(self, job_id: str, **kwargs) -> None:
        """Update a correlation job's state."""
        with self._correlation_jobs_lock:
            job = self._correlation_jobs.get(job_id)
            if job:
                job.update(kwargs)

    def _correlationWorker(self, job_id: str, id: str, target: str, mode: str) -> None:
        """Background worker that runs the full correlation pipeline.

        Updates job progress at each stage so the frontend can poll.
        """
        try:
            dbh = SpiderFootDb(self.config)

            # Stage 1: Deduplication (0% -> 15%)
            self._updateCorrelationJob(job_id, progress=5, step='Deduplicating events...')
            try:
                dedup = dbh.deduplicateScanResults(id)
                self.log.info(f"Pre-correlation dedup: removed={dedup.get('removed', 0)}, fp_preserved={dedup.get('fp_preserved', 0)}")
            except Exception as e:
                self.log.warning(f"Pre-correlation dedup failed (continuing anyway): {e}")

            self._updateCorrelationJob(job_id, progress=15, step='Deduplication complete')

            # Stage 2: Clear old correlations (15% -> 20%)
            ruleId = 'ai_single_scan_correlation' if mode == 'single' else 'ai_cross_scan_correlation'
            try:
                dbh.deleteCorrelationsByRule(id, ruleId)
                self.log.info(f"Cleared previous {ruleId} correlations for scan {id}")
            except Exception as e:
                self.log.warning(f"Failed to clear old correlations (continuing anyway): {e}")

            self._updateCorrelationJob(job_id, progress=20, step='Analyzing scan data...')

            # Stage 3: Run AI correlation (20% -> 55%)
            try:
                if mode == 'single':
                    ai_result = self._runSingleScanCorrelation(dbh, id, target)
                else:
                    ai_result = self._runCrossScanCorrelation(dbh, id, target)
            except Exception as e:
                self.log.error(f"Error running correlation: {e}", exc_info=True)
                self._updateCorrelationJob(
                    job_id,
                    status='error',
                    progress=100,
                    step='Error',
                    result={'success': False, 'error': f'Error running correlation: {str(e)}'},
                    completed_at=time.time()
                )
                return

            self._updateCorrelationJob(job_id, progress=55, step='Running correlation rules...')

            # Stage 4: Rule-based correlations (55% -> 95%)
            try:
                self._rerunRuleCorrelations(dbh, id)
            except Exception as e:
                self.log.warning(f"Failed to re-run rule-based correlations: {e}")

            self._updateCorrelationJob(job_id, progress=95, step='Finalizing results...')

            # Done
            self._updateCorrelationJob(
                job_id,
                status='complete',
                progress=100,
                step='Complete',
                result=ai_result,
                completed_at=time.time()
            )
            self.log.info(f"Correlation job {job_id} complete: {ai_result.get('correlations_found', 0)} correlations found")

        except Exception as e:
            self.log.error(f"Correlation worker error: {e}", exc_info=True)
            self._updateCorrelationJob(
                job_id,
                status='error',
                progress=100,
                step='Error',
                result={'success': False, 'error': str(e)},
                completed_at=time.time()
            )

    def _rerunRuleCorrelations(self: 'SpiderFootWebUi', dbh, scanId: str) -> None:
        """Re-run rule-based correlations for a scan.

        This ensures rule-based correlations (e.g. MALICIOUS_SUBNET,
        MALICIOUS_AFFILIATE) survive event deduplication by recreating
        them from the current scan data.
        """
        import os
        from spiderfoot.correlation.rule_loader import RuleLoader
        from spiderfoot.correlation.rule_executor import RuleExecutor

        correlations_dir = os.path.join(os.path.dirname(__file__), 'correlations')
        if not os.path.exists(correlations_dir):
            self.log.warning("Correlations directory not found, skipping rule-based correlations")
            return

        loader = RuleLoader(correlations_dir)
        rules = loader.load_rules()
        if not rules:
            self.log.warning("No correlation rules loaded")
            return

        load_errors = loader.get_errors()
        if load_errors:
            for fname, err in load_errors:
                self.log.warning(f"Failed to load rule {fname}: {err}")

        self.log.info(f"Loaded {len(rules)} correlation rules for re-run")

        # Delete old rule-based correlations for this scan before re-running
        for rule in rules:
            rule_id = rule.get('id', '')
            if rule_id:
                try:
                    deleted = dbh.deleteCorrelationsByRule(scanId, rule_id)
                    if deleted > 0:
                        self.log.debug(f"Deleted {deleted} old correlations for rule {rule_id}")
                except Exception as e:
                    self.log.warning(f"Failed to delete old correlations for rule {rule_id}: {e}")

        # Run rule-based correlation engine
        executor = RuleExecutor(dbh, rules, scan_ids=[scanId])
        results = executor.run()

        # Log details about rule execution results
        rules_matched = sum(1 for r in results.values() if r.get('matched'))
        total_corrs = sum(r.get('correlations_created', 0) for r in results.values())
        self.log.info(
            f"Re-ran rule-based correlations for scan {scanId}: "
            f"{len(results)} rules evaluated, {rules_matched} matched, "
            f"{total_corrs} correlations created"
        )
        for rule_id, result in results.items():
            if result.get('correlations_created', 0) > 0:
                self.log.info(
                    f"  Rule '{rule_id}': {result['correlations_created']} correlation(s), "
                    f"risk={result.get('meta', {}).get('risk', 'unknown')}"
                )

    @staticmethod
    def _classifyEventTypeRisk(event_type_code: str) -> int:
        """Return an intrinsic risk score (0-100) based on event type code.

        Since most SpiderFoot modules store risk=0 on events, this method
        derives a meaningful risk score from the event type itself.
        """
        code = event_type_code.upper()

        # Critical risk (80-100): actively malicious or compromised
        if any(k in code for k in [
            'MALICIOUS_', 'VULNERABILITY_CVE_CRITICAL', 'DEFACED_',
            'PASSWORD_COMPROMISED', 'HASH_COMPROMISED',
        ]):
            return 90
        if any(k in code for k in [
            'BLACKLISTED_', 'DARKNET_', 'LEAKSITE_',
        ]):
            return 85

        # High risk (60-79): compromised accounts, high-severity vulns
        if any(k in code for k in [
            'COMPROMISED', 'VULNERABILITY_CVE_HIGH', 'VULNERABILITY_GENERAL',
            'VULNERABILITY_DISCLOSURE',
        ]):
            return 70
        if any(k in code for k in [
            'CREDIT_CARD_NUMBER', 'IBAN_NUMBER',
        ]):
            return 65

        # Medium risk (40-59): open ports, interesting findings, hijackable
        if any(k in code for k in [
            'VULNERABILITY_CVE_MEDIUM', 'TCP_PORT_OPEN', 'UDP_PORT_OPEN',
            'HIJACKABLE', 'CLOUD_STORAGE_BUCKET_OPEN',
            'SSL_CERTIFICATE_EXPIRED', 'SSL_CERTIFICATE_MISMATCH',
        ]):
            return 50
        if any(k in code for k in [
            'PHONE_NUMBER', 'PHYSICAL_ADDRESS', 'PHYSICAL_COORDINATES',
            'INTERESTING_FILE', 'URL_PASSWORD', 'URL_UPLOAD',
            'PROXY_HOST', 'VPN_HOST', 'TOR_EXIT_NODE',
        ]):
            return 45

        # Low risk (20-39): affiliates, external accounts, similar domains
        if any(k in code for k in [
            'VULNERABILITY_CVE_LOW', 'AFFILIATE_', 'SIMILARDOMAIN',
            'CO_HOSTED_SITE', 'SOCIAL_MEDIA', 'PUBLIC_CODE_REPO',
            'ACCOUNT_EXTERNAL_OWNED', 'SIMILAR_ACCOUNT',
            'WEBSERVER_STRANGEHEADER', 'SSL_CERTIFICATE_EXPIRING',
            'PROVIDER_JAVASCRIPT', 'JUNK_FILE',
        ]):
            return 25

        # Info (0-19): basic entity data, descriptors
        if any(k in code for k in [
            'DOMAIN_NAME', 'IP_ADDRESS', 'IPV6_ADDRESS', 'EMAILADDR',
            'INTERNET_NAME', 'HUMAN_NAME', 'PERSON_NAME', 'USERNAME',
            'COMPANY_NAME', 'COUNTRY_NAME', 'BGP_AS_', 'NETBLOCK_',
            'DNS_', 'HTTP_CODE', 'GEOINFO', 'WEBSERVER_BANNER',
            'WEBSERVER_TECHNOLOGY', 'WEB_ANALYTICS', 'DESCRIPTION_',
            'LINKED_URL_', 'SOFTWARE_USED', 'SSL_CERTIFICATE_ISSUE',
            'HASH', 'RAW_', 'BASE64_', 'TARGET_WEB_', 'BITCOIN_',
            'ETHEREUM_', 'SEARCH_ENGINE_',
        ]):
            return 10

        return 15  # default for unknown types

    @staticmethod
    def _deriveCorrelationRisk(event_types: set, num_modules: int, num_occurrences: int,
                               classify_fn) -> str:
        """Determine the risk level for a correlation based on event types and corroboration.

        Combines the intrinsic risk of the highest-risk event type with
        boosts for multi-module confirmation and high occurrence counts.
        """
        # Base risk from highest-risk event type
        base_risk = max(classify_fn(et) for et in event_types) if event_types else 0

        # Boost for multi-module corroboration
        if num_modules >= 4:
            base_risk += 20
        elif num_modules >= 3:
            base_risk += 15
        elif num_modules >= 2:
            base_risk += 10

        # Boost for high occurrence count
        if num_occurrences >= 10:
            base_risk += 10
        elif num_occurrences >= 5:
            base_risk += 5

        base_risk = min(base_risk, 100)

        if base_risk >= 90:
            return "CRITICAL"
        if base_risk >= 65:
            return "HIGH"
        if base_risk >= 40:
            return "MEDIUM"
        if base_risk >= 20:
            return "LOW"
        return "INFO"

    def _runSingleScanCorrelation(self, dbh, id: str, target: str) -> dict:
        """Analyze a single scan's results to find significant patterns.

        Groups IOCs by data value across modules/event types. IOCs detected
        by multiple modules or appearing in multiple event types are stored
        as correlation results in tbl_scan_correlation_results so they appear
        in the standard correlations table.
        """
        # Build event type description lookup
        type_descr = {}
        for row in dbh.eventTypes():
            type_descr[row[1]] = row[0]  # event_code -> description

        # Join with source event to get the actual IOC value.
        # For MALICIOUS_*/BLACKLISTED_* events, the data field contains
        # a module-specific description (e.g., "Spamhaus [1.2.3.4]") while
        # the source event's data holds the actual IOC (e.g., "1.2.3.4").
        qry = """SELECT r.generated, r.data, r.module, r.type,
                        r.confidence, r.risk, r.hash,
                        COALESCE(sr.data, r.data) as source_data
                 FROM tbl_scan_results r
                 LEFT JOIN tbl_scan_results sr
                     ON sr.scan_instance_id = r.scan_instance_id
                     AND sr.hash = r.source_event_hash
                 WHERE r.scan_instance_id = ?
                   AND r.type != 'ROOT'
                   AND r.type != 'AI_SINGLE_SCAN_CORRELATION'
                   AND r.type != 'AI_CROSS_SCAN_CORRELATION'
                   AND r.imported_from_scan IS NULL
                   AND r.false_positive = 0"""

        with dbh.dbhLock:
            dbh.dbh.execute(qry, [id])
            all_events = dbh.dbh.fetchall()

        if not all_events:
            self.log.info("Single-scan correlation: no events found in scan")
            return {
                'success': True,
                'mode': 'single',
                'message': 'No events found in scan',
                'correlations_found': 0,
                'total_events': 0
            }

        self.log.info(f"Single-scan correlation: {len(all_events)} events to analyze")

        # Count malicious event types for debugging
        malicious_count = sum(1 for e in all_events if 'MALICIOUS' in (e[3] or '').upper())
        blacklisted_count = sum(1 for e in all_events if 'BLACKLISTED' in (e[3] or '').upper())
        self.log.info(f"Single-scan correlation: {malicious_count} MALICIOUS events, {blacklisted_count} BLACKLISTED events")

        # Group events by IOC value.
        # For MALICIOUS_*/BLACKLISTED_* events, use the source event's data
        # (the actual IP/domain/subnet) as the grouping key so that the same
        # IOC flagged by different modules is properly correlated.
        # For all other event types, group by the event's own data.
        ioc_map = {}
        for event in all_events:
            data = event[1]
            event_type = event[3] or ''
            source_data = event[7]  # from JOIN: COALESCE(sr.data, r.data)
            if not data:
                continue

            if event_type.startswith('MALICIOUS_') or event_type.startswith('BLACKLISTED_'):
                group_key = source_data if source_data else data
            else:
                group_key = data

            if group_key not in ioc_map:
                ioc_map[group_key] = []
            ioc_map[group_key].append({
                'timestamp': event[0],
                'data': data,
                'module': event[2],
                'event_type': event_type,
                'confidence': event[4],
                'risk': event[5],
                'hash': event[6],
            })

        self.log.info(f"Single-scan correlation: {len(ioc_map)} unique IOCs")

        correlations_found = 0
        skipped_below_threshold = 0

        for ioc_data, occurrences in ioc_map.items():
            modules = set(o['module'] for o in occurrences)
            event_types = set(o['event_type'] for o in occurrences)
            total = len(occurrences)

            # Only flag if the same data was found by multiple modules
            # OR appeared in multiple event types OR appeared many times
            if len(modules) < 2 and len(event_types) < 2 and total < 3:
                # Log skipped high-risk IOCs for debugging
                has_malicious = any('MALICIOUS' in et.upper() or 'BLACKLIST' in et.upper() for et in event_types)
                if has_malicious:
                    skipped_below_threshold += 1
                    self.log.debug(
                        f"Skipped malicious IOC below threshold: data={ioc_data[:60]}, "
                        f"modules={len(modules)}, types={len(event_types)} ({','.join(event_types)}), total={total}"
                    )
                continue

            event_hashes = [o['hash'] for o in occurrences]

            # Determine risk level from event type classification + corroboration
            risk_level = self._deriveCorrelationRisk(
                event_types, len(modules), total, self._classifyEventTypeRisk
            )

            # Generate a meaningful headline
            primary_type = max(event_types, key=lambda t: sum(1 for o in occurrences if o['event_type'] == t))
            type_label = type_descr.get(primary_type, primary_type)

            ioc_display = ioc_data[:80] + ('...' if len(ioc_data) > 80 else '')

            if len(modules) >= 2:
                headline = f"{type_label} corroborated by {len(modules)} modules: {ioc_display}"
            elif len(event_types) >= 2:
                headline = f"{type_label} found across {len(event_types)} event types: {ioc_display}"
            else:
                headline = f"{type_label} detected {total} times: {ioc_display}"

            # Build description
            module_list = ', '.join(sorted(modules))
            type_list = ', '.join(sorted(type_descr.get(t, t) for t in event_types))
            description = (
                f"This indicator was detected by {len(modules)} module(s) ({module_list}) "
                f"across {len(event_types)} event type(s) ({type_list}) "
                f"with {total} total occurrence(s). "
            )
            if len(modules) >= 3:
                description += "Multiple independent modules confirm this finding, indicating high significance."
            elif len(modules) >= 2:
                description += "Corroborated by multiple modules, increasing confidence in this finding."

            # Store as a standard correlation result
            try:
                dbh.correlationResultCreate(
                    instanceId=id,
                    event_hash=event_hashes[0],
                    ruleId="ai_single_scan_correlation",
                    ruleName="Single Scan Correlation",
                    ruleDescr=description,
                    ruleRisk=risk_level,
                    ruleYaml="",
                    correlationTitle=headline,
                    eventHashes=event_hashes
                )
                correlations_found += 1
                if risk_level in ('CRITICAL', 'HIGH'):
                    self.log.info(
                        f"Created {risk_level} correlation: {headline[:100]} "
                        f"(modules={len(modules)}, types={len(event_types)}, total={total})"
                    )
            except Exception as e:
                self.log.warning(f"Failed to store correlation: {e}")
                continue

        if skipped_below_threshold > 0:
            self.log.info(f"Single-scan correlation: {skipped_below_threshold} malicious IOC(s) skipped (below threshold)")
        self.log.info(f"Single-scan correlation complete. Found {correlations_found} correlated IOCs.")

        return {
            'success': True,
            'mode': 'single',
            'message': 'Single-scan correlation analysis complete',
            'correlations_found': correlations_found,
            'total_events': len(all_events),
            'unique_iocs': len(ioc_map)
        }

    def _runCrossScanCorrelation(self, dbh, id: str, target: str) -> dict:
        """Find IOCs that appear in both native and imported scan data.

        Matches IOCs across native and imported events, then stores results
        in tbl_scan_correlation_results so they appear in the standard
        correlations table with expandable detail rows.
        """
        # Build event type description lookup
        type_descr = {}
        for row in dbh.eventTypes():
            type_descr[row[1]] = row[0]  # event_code -> description

        qry = """SELECT r.generated, r.data, r.module, r.source_event_hash, r.type,
                        r.confidence, r.visibility, r.risk, r.false_positive,
                        r.hash, r.imported_from_scan,
                        COALESCE(si.name, 'Unknown') as source_scan_name,
                        COALESCE(si.started, '') as source_scan_started
                 FROM tbl_scan_results r
                 LEFT JOIN tbl_scan_instance si ON r.imported_from_scan = si.guid
                 WHERE r.scan_instance_id = ?
                   AND r.type != 'ROOT'
                   AND r.type != 'AI_CROSS_SCAN_CORRELATION'
                   AND r.type != 'AI_SINGLE_SCAN_CORRELATION'
                   AND r.false_positive = 0"""

        with dbh.dbhLock:
            dbh.dbh.execute(qry, [id])
            all_events = dbh.dbh.fetchall()

        if not all_events:
            return {
                'success': True,
                'mode': 'cross',
                'message': 'No events found in scan',
                'correlations_found': 0,
                'native_iocs': 0,
                'imported_iocs': 0
            }

        # Separate native vs imported events
        native_iocs = {}
        imported_iocs = {}

        for event in all_events:
            event_data = event[1]
            event_type = event[4]
            event_timestamp = event[0]
            imported_from = event[10]
            source_scan_name = event[11]
            event_hash = event[9]
            event_risk = event[7]

            if not event_data:
                continue

            if imported_from is None:
                if event_data not in native_iocs:
                    native_iocs[event_data] = []
                native_iocs[event_data].append({
                    'event_type': event_type,
                    'timestamp': event_timestamp,
                    'hash': event_hash,
                    'risk': event_risk,
                    'module': event[2],
                })
            else:
                if event_data not in imported_iocs:
                    imported_iocs[event_data] = []
                imported_iocs[event_data].append({
                    'scan_id': imported_from,
                    'scan_name': source_scan_name,
                    'event_type': event_type,
                    'timestamp': event_timestamp,
                    'hash': event_hash,
                    'risk': event_risk,
                    'module': event[2],
                })

        self.log.info(f"Found {len(native_iocs)} unique native IOCs and {len(imported_iocs)} unique imported IOCs")

        if not imported_iocs:
            return {
                'success': True,
                'mode': 'cross',
                'message': 'No imported historical data found.',
                'correlations_found': 0,
                'native_iocs': len(native_iocs),
                'imported_iocs': 0
            }

        correlations_found = 0

        for ioc_data, native_occurrences in native_iocs.items():
            if ioc_data not in imported_iocs:
                continue

            imported_occurrences = imported_iocs[ioc_data]

            seen_scans = set()
            for imp in imported_occurrences:
                seen_scans.add(imp['scan_id'])

            occurrence_count = len(native_occurrences) + len(imported_occurrences)

            # Collect all event hashes for linking
            event_hashes = [o['hash'] for o in native_occurrences] + [o['hash'] for o in imported_occurrences]

            # Determine the primary event type for the headline
            all_event_types = [o['event_type'] for o in native_occurrences] + [o['event_type'] for o in imported_occurrences]
            primary_type = max(set(all_event_types), key=all_event_types.count)
            type_label = type_descr.get(primary_type, primary_type)

            ioc_display = ioc_data[:80] + ('...' if len(ioc_data) > 80 else '')

            # Determine risk level from event type classification + corroboration
            all_event_types_set = set(all_event_types)
            all_modules = set(o['module'] for o in native_occurrences) | set(o['module'] for o in imported_occurrences)
            risk_level = self._deriveCorrelationRisk(
                all_event_types_set, len(all_modules), occurrence_count,
                self._classifyEventTypeRisk
            )

            # Generate meaningful headline
            scan_names = sorted(set(imp['scan_name'] for imp in imported_occurrences))
            headline = f"{type_label} persists across {len(seen_scans)} historical scan(s): {ioc_display}"

            # Build description
            native_modules = sorted(set(o['module'] for o in native_occurrences))
            imported_modules = sorted(set(o['module'] for o in imported_occurrences))
            description = (
                f"This indicator was found in the current scan ({len(native_occurrences)} occurrence(s) "
                f"via {', '.join(native_modules)}) AND in {len(imported_occurrences)} imported record(s) "
                f"from {len(seen_scans)} historical scan(s) ({', '.join(scan_names)}). "
                f"Persistence across scans indicates this is a recurring indicator worth investigation."
            )

            # Store as a standard correlation result
            try:
                dbh.correlationResultCreate(
                    instanceId=id,
                    event_hash=event_hashes[0],
                    ruleId="ai_cross_scan_correlation",
                    ruleName="Cross-Scan Correlation",
                    ruleDescr=description,
                    ruleRisk=risk_level,
                    ruleYaml="",
                    correlationTitle=headline,
                    eventHashes=event_hashes
                )
                correlations_found += 1
            except Exception as e:
                self.log.warning(f"Failed to store cross-scan correlation: {e}")
                continue

        self.log.info(f"Cross-scan correlation complete. Found {correlations_found} correlations.")

        return {
            'success': True,
            'mode': 'cross',
            'message': 'Cross-scan correlation analysis complete',
            'correlations_found': correlations_found,
            'native_iocs': len(native_iocs),
            'imported_iocs': len(imported_iocs)
        }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'SpiderFootWebUi', id: str, eventType: str = None, filterfp: bool = False, correlationId: str = None) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results
            correlationId (str): filter by events associated with a correlation

        Returns:
            list: scan results with target-level FP and validated status
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        if not eventType:
            eventType = 'ALL'

        try:
            data = dbh.scanResultEvent(
                id, eventType, filterFp=filterfp, correlationId=correlationId)
        except Exception as e:
            self.log.warning(
                f"scaneventresults failed for scan={id}, correlationId={correlationId}, "
                f"eventType={eventType}: {e}"
            )
            return retdata

        # Get the target for this scan to check target-level FPs and validated status
        scanInfo = dbh.scanInstanceGet(id)
        target = scanInfo[1] if scanInfo else None

        # Get all target-level false positives and validated entries for fast lookup
        targetFps = set()
        targetValidated = set()
        knownAssets = {'ip': set(), 'domain': set(), 'employee': set()}
        if target:
            try:
                targetFps = dbh.targetFalsePositivesForTarget(target)
            except Exception:
                pass  # Table may not exist in older databases
            try:
                targetValidated = dbh.targetValidatedForTarget(target)
            except Exception:
                pass  # Table may not exist in older databases
            try:
                knownAssets = dbh.knownAssetValues(target)
            except Exception:
                pass  # Table may not exist in older databases

        # Pre-compute known asset matching sets
        ip_match_types = {'IP_ADDRESS', 'IPV6_ADDRESS', 'AFFILIATE_IPADDR'}
        domain_match_types = {'DOMAIN_NAME', 'INTERNET_NAME', 'AFFILIATE_INTERNET_NAME',
                              'CO_HOSTED_SITE', 'SIMILARDOMAIN', 'INTERNET_NAME_UNRESOLVED'}
        employee_match_types = {'HUMAN_NAME', 'USERNAME', 'EMAILADDR', 'AFFILIATE_EMAILADDR', 'SOCIAL_MEDIA'}
        hasKnownAssets = any(knownAssets.values())

        for row in data:
            lastseen = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            eventDataRaw = row[1]
            sourceDataRaw = row[2]
            eventTypeRaw = row[4]

            # Check if this result matches a target-level false positive (including source for granular matching)
            isTargetFp = 1 if (eventTypeRaw, eventDataRaw, sourceDataRaw) in targetFps else 0
            # Check if this result matches a target-level validated entry
            isTargetValidated = 1 if (eventTypeRaw, eventDataRaw, sourceDataRaw) in targetValidated else 0

            # Check if this result matches a known asset
            isKnownAsset = 0  # 0=no match, 1=client_provided match, 2=analyst_confirmed match
            if hasKnownAssets and eventDataRaw:
                dataLower = eventDataRaw.lower().strip()
                if eventTypeRaw in ip_match_types and dataLower in knownAssets['ip']:
                    isKnownAsset = 1
                elif eventTypeRaw in domain_match_types:
                    if dataLower in knownAssets['domain']:
                        isKnownAsset = 1
                    else:
                        for kd in knownAssets['domain']:
                            if dataLower.endswith('.' + kd):
                                isKnownAsset = 1
                                break
                elif eventTypeRaw in employee_match_types:
                    for ke in knownAssets['employee']:
                        if ke in dataLower:
                            isKnownAsset = 1
                            break

            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4],
                isTargetFp,  # Index 11: target-level false positive flag
                isTargetValidated,  # Index 12: target-level validated flag
                row[15],  # Index 13: imported_from_scan (scan ID if imported, None otherwise)
                isKnownAsset,  # Index 14: known asset match (0=no, 1=match)
                row[16] if len(row) > 16 else 0  # Index 15: tracking (0=OPEN, 1=CLOSED, 2=TICKETED)
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'SpiderFootWebUi', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'SpiderFootWebUi', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'SpiderFootWebUi', id: str, eventType: str, filterFp: str = "1") -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterFp (str): filter out false positives (1=yes, 0=no)

        Returns:
            dict
        """
        dbh = SpiderFootDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            # Filter false positives from leaf set if requested
            filter_fp = filterFp == "1"
            leafSet = dbh.scanResultEvent(id, eventType, filterFp=filter_fp)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)

        # Add a synthetic entry for "Discovery Paths" node if multiple roots exist
        # This provides tooltip data for the synthetic parent node
        if retdata['tree'].get('name') == 'Discovery Paths':
            datamap['Discovery Paths'] = [
                0,  # generated timestamp
                'Discovery Paths',  # data (display name)
                '',  # source_data
                'SpiderFoot',  # module
                'ROOT',  # type
                100,  # confidence
                100,  # visibility
                0,  # risk
                'discovery_paths',  # hash
                'ROOT',  # source_event_hash
                'Discovery Paths',  # event_descr
                'ROOT',  # event_type
                '',  # scan_instance_id
                0,  # false_positive
                0   # parent_fp
            ]

        retdata['data'] = datamap
        retdata['scanId'] = id

        return retdata

    @cherrypy.expose
    def active_maintenance_status(self: 'SpiderFootWebUi') -> str:
        """Display the active maintenance status of the project.

        Returns:
            str: Active maintenance status page HTML
        """
        templ = Template(
            filename='spiderfoot/templates/active_maintenance_status.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def footer(self: 'SpiderFootWebUi') -> str:
        """Display the footer with active maintenance status.

        Returns:
            str: Footer HTML
        """
        templ = Template(
            filename='spiderfoot/templates/footer.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__)

    # Workspace Management API Endpoints
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacelist(self: 'SpiderFootWebUi') -> list:
        """List all workspaces.

        Returns:
            list: List of workspace information
        """
        try:
            workspaces = SpiderFootWorkspace.list_workspaces(self.config)
            return workspaces
        except Exception as e:
            self.log.error(f"Failed to list workspaces: {e}")
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacecreate(self: 'SpiderFootWebUi', name: str, description: str = '') -> dict:
        """Create a new workspace.

        Args:
            name (str): workspace name
            description (str): workspace description

        Returns:
            dict: workspace creation result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, name=name)
            workspace.description = description
            workspace.save_workspace()
            
            return {
                'success': True,
                'workspace_id': workspace.workspace_id,
                'name': workspace.name,
                'description': workspace.description,
                'created_time': workspace.created_time
            }
        except Exception as e:
            self.log.error(f"Failed to create workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceget(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Get workspace details.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: workspace information
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            return {
                'success': True,
                'workspace_id': workspace.workspace_id,
                'name': workspace.name,
                'description': workspace.description,
                               'created_time': workspace.created_time,
                'modified_time': workspace.modified_time,
                'targets': workspace.targets,
                'scans': workspace.scans,
                'metadata': workspace.metadata
            }
        except Exception as e:
            self.log.error(f"Failed to get workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceupdate(self: 'SpiderFootWebUi', workspace_id: str, name: str = None, description: str = None) -> dict:
        """Update workspace details.

        Args:
            workspace_id (str): workspace ID
            name (str): new workspace name
            description (str): new workspace description

        Returns:
            dict: update result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            
            if name is not None:
                workspace.name = name
            if description is not None:
                workspace.description = description
                
            workspace.save_workspace()
            
            return {'success': True, 'message': 'Workspace updated successfully'}
        except Exception as e:
            self.log.error(f"Failed to update workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacedelete(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Delete a workspace.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: deletion result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            workspace.delete_workspace()
            
            return {'success': True, 'message': 'Workspace deleted successfully'}
        except Exception as e:
            self.log.error(f"Failed to delete workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacesummary(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Get workspace summary.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: workspace summary
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            summary = workspace.get_workspace_summary()
            
            return {'success': True, 'summary': summary}
        except Exception as e:
            self.log.error(f"Failed to get workspace summary: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceaddtarget(self: 'SpiderFootWebUi', workspace_id: str, target: str, target_type: str = None) -> dict:
        """Add target to workspace.

        Args:
            workspace_id (str): workspace ID
            target (str): target value
            target_type (str): target type

        Returns:
            dict: add target result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            target_id = workspace.add_target(target, target_type)
            
            return {'success': True, 'target_id': target_id, 'message': 'Target added successfully'}
        except Exception as e:
            self.log.error(f"Failed to add target: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceremovetarget(self: 'SpiderFootWebUi', workspace_id: str, target_id: str) -> dict:
        """Remove target from workspace.

        Args:
            workspace_id (str): workspace ID
            target_id (str): target ID

        Returns:
            dict: remove target result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            success = workspace.remove_target(target_id)
            
            if success:
                return {'success': True, 'message': 'Target removed successfully'}
            else:
                return {'success': False, 'error': 'Target not found'}
        except Exception as e:
            self.log.error(f"Failed to remove target: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceimportscans(self: 'SpiderFootWebUi', workspace_id: str, scan_ids: str) -> dict:
        """Import scans into workspace.

        Args:
            workspace_id (str): workspace ID
            scan_ids (str): comma-separated scan IDs

        Returns:
            dict: import result
        """
        try:
            self.log.info(f"[IMPORT] Starting scan import for workspace: {workspace_id}")
            self.log.debug(f"[IMPORT] Raw scan IDs input: {scan_ids}")
            
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            self.log.info(f"[IMPORT] Loaded workspace: {workspace.name}")
            
            # Clean and split scan IDs (handle both comma-separated and line-separated)
            scan_ids_cleaned = scan_ids.replace('\n', ',').replace('\r', '')
            scan_id_list = [sid.strip() for sid in scan_ids_cleaned.split(',') if sid.strip()]
            
            self.log.info(f"[IMPORT] Processed {len(scan_id_list)} scan IDs: {scan_id_list}")
            
            if not scan_id_list:
                return {'success': False, 'error': 'No valid scan IDs provided'}
            
            # Verify scans exist before importing
            dbh = SpiderFootDb(self.config)
            valid_scans = []
            invalid_scans = []
            
            for scan_id in scan_id_list:
                scan_info = dbh.scanInstanceGet(scan_id)
                if scan_info:
                    valid_scans.append(scan_id)
                    self.log.debug(f"[IMPORT] Verified scan {scan_id}: {scan_info[0]}")
                else:
                    invalid_scans.append(scan_id)
                    self.log.warning(f"[IMPORT] Scan {scan_id} not found in database")
            
            if invalid_scans:
                self.log.warning(f"[IMPORT] Invalid scan IDs: {invalid_scans}")
            
            if not valid_scans:
                return {'success': False, 'error': f'No valid scans found. Invalid IDs: {invalid_scans}'}
            
            # Import valid scans
            if len(valid_scans) == 1:
                success = workspace.import_single_scan(valid_scans[0])
                if success:
                    self.log.info(f"[IMPORT] Successfully imported scan {valid_scans[0]}")
                    return {'success': True, 'message': 'Scan imported successfully'}
                else:
                    self.log.error(f"[IMPORT] Failed to import scan {valid_scans[0]}")
                    return {'success': False, 'error': 'Failed to import scan'}
            else:
                results = workspace.bulk_import_scans(valid_scans)
                successful_imports = sum(1 for success in results.values() if success)
                
                self.log.info(f"[IMPORT] Bulk import completed: {successful_imports}/{len(valid_scans)} successful")
                
                message = f'Import completed: {successful_imports} of {len(valid_scans)} scans imported'
                if invalid_scans:
                    message += f'. Invalid scan IDs: {invalid_scans}'
                
                return {
                    'success': True, 
                    'results': results,
                    'message': message,
                    'successful_imports': successful_imports,
                    'total_attempts': len(scan_id_list),
                    'invalid_scans': invalid_scans
                }
        except Exception as e:
            self.log.error(f"[IMPORT] Failed to import scans: {e}")
            import traceback
            self.log.error(f"[IMPORT] Traceback: {traceback.format_exc()}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacemultiscan(self: 'SpiderFootWebUi', workspace_id: str, targets: str, modules: str, scan_name_prefix: str, enable_correlation: str = 'false') -> dict:
        """Start multi-target scan from workspace.

        Args:
            workspace_id (str): workspace ID
            targets (str): JSON string of selected targets
            modules (str): JSON string of selected modules
            scan_name_prefix (str): prefix for scan names
            enable_correlation (str): whether to enable correlation

        Returns:
            dict: multi-target scan result
        """
        self.log.info(f"[MULTISCAN] Starting multi-target scan for workspace: {workspace_id}")
        self.log.debug(f"[MULTISCAN] Input parameters - targets: {targets}, modules: {modules}, prefix: {scan_name_prefix}")
        
        try:
            self.log.debug(f"[MULTISCAN] Importing startSpiderFootScanner...")
            from sfscan import startSpiderFootScanner
            self.log.debug(f"[MULTISCAN] Import successful")
            
            # Try to load existing workspace, or create a new one if it doesn't exist
            self.log.debug(f"[MULTISCAN] Attempting to load workspace: {workspace_id}")
            try:
                workspace = SpiderFootWorkspace(self.config, workspace_id)
                self.log.info(f"[MULTISCAN] Successfully loaded existing workspace: {workspace_id}")
            except (ValueError, Exception) as e:
                # Workspace doesn't exist, create a new one
                self.log.info(f"[MULTISCAN] Workspace {workspace_id} not found ({e}), creating new one")
                try:
                    workspace = SpiderFootWorkspace(self.config, name=f"Workspace_{workspace_id}")
                    workspace.workspace_id = workspace_id  # Override the generated ID                    workspace.save_workspace()
                    self.log.info(f"[MULTISCAN] Successfully created new workspace: {workspace_id}")
                except Exception as create_error:
                    self.log.error(f"[MULTISCAN] Failed to create workspace: {create_error}")
                    raise
            
            # Parse targets and modules
            self.log.debug("[MULTISCAN] Parsing JSON input data...")
            try:
                target_list = json.loads(targets)
                self.log.debug(f"[MULTISCAN] Parsed {len(target_list)} targets: {[t.get('value', 'unknown') for t in target_list]}")
            except Exception as e:
                self.log.error(f"[MULTISCAN] Failed to parse targets JSON: {e}")
                raise ValueError(f"Invalid targets JSON: {e}")
            
            try:
                module_list = json.loads(modules)
                self.log.debug(f"[MULTISCAN] Parsed {len(module_list)} modules: {module_list}")
            except Exception as e:
                self.log.error(f"[MULTISCAN] Failed to parse modules JSON: {e}")
                raise ValueError(f"Invalid modules JSON: {e}")
            
            scan_ids = []
            
            self.log.info(f"[MULTISCAN] Starting scan loop for {len(target_list)} targets")
            
            # Start a scan for each target
            for i, target in enumerate(target_list):
                self.log.debug(f"[MULTISCAN] Processing target {i+1}/{len(target_list)}: {target}")
                
                target_value = target['value']
                target_type = target.get('type', '')
                
                self.log.debug(f"[MULTISCAN] Target value: {target_value}, type: {target_type}")
                
                # If target type is not provided or empty, detect it
                if not target_type:
                    self.log.debug(f"[MULTISCAN] Detecting target type for: {target_value}")
                    target_type = SpiderFootHelpers.targetTypeFromString(target_value)
                    if target_type is None:
                        self.log.error(f"[MULTISCAN] Could not determine target type for {target_value}")
                        continue
                    else:
                        self.log.debug(f"[MULTISCAN] Detected target type: {target_type}")
                
                # Normalize target value like other scan methods
                original_value = target_value
                if target_type in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
                    target_value = target_value.replace("\"", "")
                else:
                    target_value = target_value.lower()
                
                if original_value != target_value:
                    self.log.debug(f"[MULTISCAN] Normalized target value: {original_value} -> {target_value}")
                
                # Generate scan name
                scan_name = f"{scan_name_prefix} - {target_value}"
                self.log.debug(f"[MULTISCAN] Generated scan name: {scan_name}")
                
                # Create module configuration list (like in working examples)
                modlist = module_list.copy()
                self.log.debug(f"[MULTISCAN] Initial module list: {modlist}")
                
                # Add our mandatory storage module
                if "sfp__stor_db" not in modlist:
                    modlist.append("sfp__stor_db")
                    self.log.debug("[MULTISCAN] Added mandatory sfp__stor_db module")
                
                # Delete the stdout module in case it crept in
                if "sfp__stor_stdout" in modlist:
                    modlist.remove("sfp__stor_stdout")
                    self.log.debug("[MULTISCAN] Removed sfp__stor_stdout module")
                
                self.log.debug(f"[MULTISCAN] Final module list: {modlist}")
                
                # Create configuration copy for this scan
                self.log.debug("[MULTISCAN] Creating configuration copy...")
                cfg = deepcopy(self.config)
                
                # Start the scan using the correct signature
                scanId = SpiderFootHelpers.genScanInstanceId()
                self.log.info(f"[MULTISCAN] Generated scan ID {scanId} for target {target_value}")                
                try:
                    self.log.debug(f"[MULTISCAN] Starting process for scan {scanId}")
                    # Use multiprocessing like the working examples
                    # startSpiderFootScanner signature: (loggingQueue, *args)
                    # where args are: (scanName, scanId, targetValue, targetType, moduleList, globalOpts)
                    p = _spawn_ctx.Process(target=startSpiderFootScanner, args=(
                        self.loggingQueue, scan_name, scanId, target_value, target_type, modlist, cfg))
                    p.daemon = True
                    p.start()
                    self.log.info(f"[MULTISCAN] Successfully started process for scan {scanId}")

                    # Track the Process object for kill capability
                    with self._scan_processes_lock:
                        self._scan_processes[scanId] = p

                    scan_ids.append(scanId)
                    
                    # Wait a moment for the scan to initialize in the database
                    import time
                    time.sleep(0.5)
                    
                    # Import the scan into the workspace
                    self.log.debug(f"[MULTISCAN] Importing scan {scanId} into workspace {workspace_id}")
                    workspace.import_single_scan(scanId, {
                        'source': 'multi_target_scan',
                        'scan_name_prefix': scan_name_prefix,
                        'target_id': target.get('target_id', 'unknown'),
                        'imported_time': time.time()
                    })
                    self.log.debug(f"[MULTISCAN] Successfully imported scan {scanId} into workspace")
                    
                except Exception as e:
                    self.log.error(f"[MULTISCAN] Failed to start scan for target {target_value}: {e}")
                    import traceback
                    self.log.error(f"[MULTISCAN] Traceback: {traceback.format_exc()}")
                    continue
            
            self.log.info(f"[MULTISCAN] Scan loop completed. Started {len(scan_ids)} out of {len(target_list)} scans")
            
            if scan_ids:
                message = f"Started {len(scan_ids)} scans successfully"
                if enable_correlation.lower() == 'true':
                    message += ". Correlation analysis will be available once scans complete"
                
                self.log.info(f"[MULTISCAN] Success: {message}")
                return {
                    'success': True,
                    'message': message,
                    'scan_ids': scan_ids,
                    'workspace_id': workspace_id
                }
            else:
                error_msg = 'Failed to start any scans'
                self.log.error(f"[MULTISCAN] {error_msg}")
                return {'success': False, 'error': error_msg}                
        except Exception as e:
            self.log.error(f"[MULTISCAN] Failed to start multi-target scan: {e}")
            import traceback
            self.log.error(f"[MULTISCAN] Traceback: {traceback.format_exc()}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacemcpreport(self: 'SpiderFootWebUi', workspace_id: str, report_type: str, format: str = 'json', 
                          include_correlations: str = 'true', include_threat_intel: str = 'true', 
                          include_recommendations: str = 'true', tlp_level: str = 'amber') -> dict:
        """Generate MCP CTI report for workspace.

        Args:
            workspace_id (str): workspace ID
            report_type (str): type of report (threat_assessment, ioc_analysis, etc.)
            format (str): output format (json, markdown, pdf, html)
            include_correlations (str): include correlation analysis
            include_threat_intel (str): include threat intelligence context
            include_recommendations (str): include security recommendations
            tlp_level (str): Traffic Light Protocol level

        Returns:
            dict: {'success': bool, 'download_url': str, 'error': str}        """
        try:
            # Validate workspace exists
            workspace = SpiderFootWorkspace(self.config, workspace_id)

            # Get workspace scans for report data
            if not workspace.scans:
                return {'success': True, 'correlations': [], 'message': 'Need at least 2 scans for cross-correlation analysis'}
            
            # Import MCP integration
            try:
                from spiderfoot.mcp_integration import SpiderFootMCPClient
                mcp_client = SpiderFootMCPClient(self.config)
            except ImportError:
                return {'success': False, 'error': 'MCP integration not available'}

            # Prepare report configuration
            report_config = {
                'workspace_id': workspace_id,
                'workspace_name': workspace.get('name', 'Unnamed'),
                'report_type': report_type,
                'format': format,
                'options': {
                    'include_correlations': include_correlations.lower() == 'true',
                    'include_threat_intel': include_threat_intel.lower() == 'true',
                    'include_recommendations': include_recommendations.lower() == 'true',
                    'tlp_level': tlp_level
                },
                'scan_ids': [scan['scan_id'] for scan in workspace.scans]
            }

            # Generate report asynchronously (this is a placeholder for actual MCP integration)
            # In a real implementation, this would call the MCP server
            import uuid
            import time
            report_id = str(uuid.uuid4())
            timestamp = int(time.time())
            
            # Create download URL (placeholder - would be actual file in production)
            download_url = f"/workspacereportdownload?report_id={report_id}&workspace_id={workspace_id}&format={format}"
            
            self.log.info(f"Generated MCP report for workspace {workspace_id}: {report_id}")
            
            return {
                'success': True,
                'report_id': report_id,
                'download_url': download_url,
                'message': f'MCP {report_type} report generated successfully'
            }

        except Exception as e:
            self.log.error(f"Failed to generate MCP report: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out() 
    def workspacetiming(self: 'SpiderFootWebUi', workspace_id: str, timezone: str = None, 
                       default_start_time: str = None, retention_period: str = None,
                       auto_scheduling: str = None, business_hours_only: str = None,
                       enable_throttling: str = None, business_start: str = None, 
                       business_end: str = None) -> dict:
        """Get or set workspace timing configuration.

        Args:
            workspace_id (str): workspace ID
            timezone (str): workspace timezone
            default_start_time (str): default scan start time (HH:MM)
            retention_period (str): data retention period in days
            auto_scheduling (str): enable automatic scheduling
            business_hours_only (str): restrict scans to business hours
            enable_throttling (str): enable scan rate throttling
            business_start (str): business hours start time (HH:MM)
            business_end (str): business hours end time (HH:MM)

        Returns:
            dict: timing configuration or success status
        """
        try:
            # Validate workspace exists
            workspace = SpiderFootWorkspace(self.config, workspace_id)

            # If this is a GET request (no parameters provided for setting)
            if timezone is None and default_start_time is None:
                # Return current timing configuration
                timing_config = workspace.metadata.get('timing_config', {})
                return {
                    'success': True,
                    'timezone': timing_config.get('timezone', 'UTC'),
                    'default_start_time': timing_config.get('default_start_time', '09:00'),
                    'retention_period': timing_config.get('retention_period', '90'),
                    'auto_scheduling': timing_config.get('auto_scheduling', False),
                    'business_hours_only': timing_config.get('business_hours_only', False),
                    'enable_throttling': timing_config.get('enable_throttling', True),
                    'business_start': timing_config.get('business_start', '08:00'),
                    'business_end': timing_config.get('business_end', '18:00')
                }

            # This is a POST request - update timing configuration
            timing_config = {
                'timezone': timezone or 'UTC',
                'default_start_time': default_start_time or '09:00', 
                'retention_period': int(retention_period) if retention_period else 90,
                'auto_scheduling': auto_scheduling == 'true' if auto_scheduling else False,
                'business_hours_only': business_hours_only == 'true' if business_hours_only else False,
                'enable_throttling': enable_throttling != 'false',  # Default to True
                'business_start': business_start or '08:00',
                'business_end': business_end or '18:00',
                'updated_time': time.time()
            }            # Update workspace with timing configuration
            import time
            workspace.metadata['timing_config'] = timing_config
            workspace.save_workspace()

            self.log.info(f"Updated timing configuration for workspace {workspace_id}")
            return {'success': True, 'message': 'Timing configuration updated successfully'}

        except Exception as e:
            self.log.error(f"Failed to handle workspace timing: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    def workspacereportdownload(self: 'SpiderFootWebUi', report_id: str, workspace_id: str, format: str = 'json'):
        """Download generated MCP report.

        Args:
            report_id (str): report identifier
            workspace_id (str): workspace ID
            format (str): report format

        Returns:
            File download or error page
        """
        try:
            # Validate workspace access
            workspace = SpiderFootWorkspace(self.config, workspace_id)            # Generate sample report content (placeholder)
            import json
            from datetime import datetime
            
            sample_report = {
                'report_id': report_id,
                'workspace_id': workspace_id,
                'workspace_name': workspace.name,
                'generated_time': datetime.now().isoformat(),
                'report_type': 'MCP CTI Report',
                'format': format,
                'status': 'This is a placeholder MCP report. Integration with actual MCP server required.',
                'summary': {
                    'total_targets': len(workspace.targets),
                    'total_scans': len(workspace.scans),
                    'risk_level': 'Medium',
                    'key_findings': [
                        'Placeholder finding 1',
                        'Placeholder finding 2', 
                        'Placeholder finding 3'
                    ]
                }
            }            # Set appropriate headers for download
            cherrypy.response.headers['Content-Type'] = 'application/octet-stream'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="mcp_report_{report_id}.{format}"'

            if format == 'json':
                return json.dumps(sample_report, indent=2)
            elif format == 'markdown':
                md_content = f"""# MCP CTI Report
                
**Report ID:** {report_id}
**Workspace:** {workspace.name}
**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Summary
- Total Targets: {len(workspace.targets)}
- Total Scans: {len(workspace.scans)}
- Risk Level: Medium

## Status
This is a placeholder MCP report. Integration with actual MCP server required.
"""
                return md_content
            else:
                return json.dumps(sample_report, indent=2)
                
        except Exception as e:
            self.log.error(f"Failed to download report: {e}")
    @cherrypy.expose
    def documentation(self: 'SpiderFootWebUi', doc: str = None, q: str = None) -> str:
        """
        Render documentation from the documentation/ folder as HTML, including subfolders.
        """
        self.log.debug("Documentation endpoint called with doc=%s, q=%s", doc, q)
        import re
        doc_dir = os.path.join(os.path.dirname(__file__), 'documentation')
        doc_dir = os.path.abspath(doc_dir)
        doc_index = []
        selected_file = None
        content = ''
        search_results = []
        search_query = q or ''
        toc_html = ''
        breadcrumbs = []
        last_updated = ''
        author = ''
        version_dirs = []
        current_version = 'latest'
        related = []
        try:
            # Recursively find all .md files
            md_files = []
            for root, dirs, files in os.walk(doc_dir):
                for fname in files:
                    if fname.endswith('.md'):
                        rel_path = os.path.relpath(os.path.join(root, fname), doc_dir)
                        rel_path = rel_path.replace('\\', '/')  # For Windows compatibility
                        md_files.append(rel_path)
            # Use README.md table for sidebar if present
            readme_path = os.path.join(doc_dir, 'README.md')
            sidebar_entries = []
            if os.path.exists(readme_path):
                with open(readme_path, encoding='utf-8') as f:
                    readme_content = f.read()
                # Extract table rows: | Section | File | Icon |
                table_rows = re.findall(r'\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|', readme_content)
                for section, file, icon in table_rows:
                    # Skip header row
                    if section.lower() == 'section':
                        continue
                    # Only add if file exists in md_files
                    if file.strip() in md_files:
                        sidebar_entries.append({
                            'file': file.strip(),
                            'title': section.strip(),
                            'icon': f'fa {icon.strip()}'
                        })
            # Fallback: use all .md files
            if not sidebar_entries:
                for rel_path in sorted(md_files):
                    title = rel_path.replace('.md', '').replace('_', ' ').replace('/', ' / ').title()
                    sidebar_entries.append({
                        'file': rel_path,
                        'title': title,
                        'icon': 'fa fa-file-text-o'
                    })
            doc_index = sidebar_entries
            # Determine which file to show
            if doc and doc.endswith('.md') and doc in md_files:
                selected_file = doc
            elif doc_index:
                selected_file = doc_index[0]['file']
            # Read and render the selected file
            if selected_file:
                file_path = os.path.join(doc_dir, selected_file)
                try:
                    with open(file_path, encoding='utf-8') as f:
                        raw_content = f.read()
                    content = markdown.markdown(
                        raw_content,
                        extensions=['extra', 'toc', 'tables', 'fenced_code']
                    )
                    # --- POST-PROCESS: Rewrite .md links to /documentation?doc=... ---
                    def md_link_rewrite(match):
                        text, url = match.group(1), match.group(2)
                        if url.endswith('.md'):
                            # Remove leading ./ or / if present
                            url = url.lstrip('./')
                            return f'<a href="/documentation?doc={url}">{text}</a>'
                        return match.group(0)
                    content = re.sub(r'<a href=["\']([^"\']+\.md)["\']>(.*?)</a>',
                                     lambda m: f'<a href="/documentation?doc={m.group(1)}">{m.group(2)}</a>',
                                     content)
                    # Also handle Markdown links rendered as <a href="modules/sfp_virustotal.md">...</a>
                    content = re.sub(r'<a href=["\'](modules/[^"\']+\.md)["\']>(.*?)</a>',
                                     lambda m: f'<a href="/documentation?doc={m.group(1)}">{m.group(2)}</a>',
                                     content)
                except Exception as e:
                    self.log.error("Failed to load documentation file %s: %s", file_path, e)
                    content = (
                        '<div class="alert alert-danger">'
                        f'Failed to load documentation: {e}'
                        '</div>'
                    )
            # Search functionality
            if search_query:
                for entry in doc_index:
                    file_path = os.path.join(doc_dir, entry['file'])
                    try:
                        with open(file_path, encoding='utf-8') as f:
                            text = f.read()
                        if (
                            search_query.lower() in text.lower()
                            or search_query.lower() in entry['title'].lower()
                        ):
                            search_results.append(entry)
                    except Exception as e:
                        self.log.warning("Error searching documentation file %s: %s", file_path, e)
                        continue
            # Breadcrumbs (simple: Home > Current)
            breadcrumbs = [
                {
                    'url': self.docroot + '/documentation',
                    'title': 'Documentation'
                }
            ]
            if selected_file:
                breadcrumbs.append({
                    'url': (
                        self.docroot
                        + '/documentation?doc='
                        + selected_file
                    ),
                    'title': (
                        selected_file.replace('.md', '')
                        .replace('_', ' ')
                        .replace('/', ' / ')
                        .title()
                    )
                })
            # Render template
            templ = Template(
                filename='spiderfoot/templates/documentation.tmpl',
                lookup=self.lookup
            )
            # Provide a dummy highlight function if not searching
            def highlight(text, query):
                import re
                if not text or not query:
                    return text
                pattern = re.compile(re.escape(query), re.IGNORECASE)
                return pattern.sub(lambda m: f'<mark>{m.group(0)}</mark>', text)
            return templ.render(
                docroot=self.docroot,
                doc_index=doc_index,
                selected_file=selected_file,
                content=content,
                search_query=search_query,
                search_results=search_results,
                toc_html=toc_html,
                breadcrumbs=breadcrumbs,
                last_updated=last_updated,
                author=author,
                version_dirs=version_dirs,
                current_version=current_version,
                related=related,
                version=__version__,
                pageid="DOCUMENTATION",
                highlight=highlight,
                user_role=self.currentUserRole()
            )
        except Exception as e:
            self.log.error("Error in documentation endpoint: %s", e, exc_info=True)
            return (
                '<div class="alert alert-danger">'
            f'Error loading documentation: {e}'
            '</div>'
        )
    @cherrypy.expose
    def workspacedetails(self: 'SpiderFootWebUi', workspace_id: str) -> str:
        """Enhanced workspace details page.

        Args:
            workspace_id (str): workspace ID

        Returns:
            str: workspace details page HTML
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            
            # Refresh workspace to get latest scan data
            workspace.load_workspace()
            
            # Get workspace summary and scan details
            dbh = SpiderFootDb(self.config)
            scan_details = []
            
            for scan in workspace.scans:
                scan_info = dbh.scanInstanceGet(scan['scan_id'])
                if scan_info:
                    scan_details.append({
                        'scan_id': scan['scan_id'],
                        'name': scan_info[0],
                        'target': scan_info[1],
                        'status': scan_info[5],
                        'created': scan_info[2],
                        'started': scan_info[3],
                        'ended': scan_info[4],
                        'imported_time': scan.get('imported_time', 0)
                    })
            
            templ = Template(filename='spiderfoot/templates/workspace_details.tmpl', lookup=self.lookup)
            return templ.render(
                workspace=workspace,
                scan_details=scan_details,
                docroot=self.docroot,
                version=__version__,
                pageid="WORKSPACE_DETAILS",
                user_role=self.currentUserRole()
            )
            
        except Exception as e:
            self.log.error(f"Error loading workspace details: {e}")
            return self.error(f"Error loading workspace details: {e}")
        
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacescancorrelations(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Get cross-scan correlations for a workspace.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: correlation analysis results
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            
            if not workspace.scans or len(workspace.scans) < 2:
                return {'success': True, 'correlations': [], 'message': 'Need at least 2 scans for cross-correlation analysis'}
            
            dbh = SpiderFootDb(self.config)
            correlations = []
              # Get correlations for each scan
            finished_scans = 0
            for scan in workspace.scans:
                # Check if scan is finished before looking for correlations
                scan_info = dbh.scanInstanceGet(scan['scan_id'])
                if scan_info and scan_info[5] == 'FINISHED':
                    finished_scans += 1
                    scan_correlations = dbh.scanCorrelationList(scan['scan_id'])
                    for corr in scan_correlations:
                        # scanCorrelationList returns:
                        #   0: id, 1: title, 2: rule_id, 3: rule_risk,
                        #   4: rule_name, 5: rule_descr, 6: rule_logic,
                        #   7: event_count, 8: event_types
                        correlations.append({
                            'scan_id': scan['scan_id'],
                            'correlation_id': corr[0],
                            'correlation': corr[1],
                            'rule_name': corr[4],
                            'rule_risk': corr[3],
                            'rule_id': corr[2],
                            'rule_description': corr[5],
                            'created': ''
                        })
            
            # Check if we have enough finished scans for correlation analysis
            if finished_scans < 2:
                return {
                    'success': True, 
                    'correlations': [], 
                    'correlation_groups': {},
                    'total_correlations': 0,
                    'cross_scan_patterns': 0,
                    'finished_scans': finished_scans,
                    'total_scans': len(workspace.scans),
                    'message': f'Need at least 2 finished scans for correlation analysis. Currently have {finished_scans} finished out of {len(workspace.scans)} total scans.'
                }
              # Group correlations by rule type
            correlation_groups = {}
            for corr in correlations:
                rule_name = corr['rule_name']
                if rule_name not in correlation_groups:
                    correlation_groups[rule_name] = []
                correlation_groups[rule_name].append(corr)
            
            return {
                'success': True,
                'correlations': correlations,
                'correlation_groups': correlation_groups,
                'total_correlations': len(correlations),
                'cross_scan_patterns': len(correlation_groups),
                'finished_scans': finished_scans,
                'total_scans': len(workspace.scans)
            }
            
        except Exception as e:
            self.log.error(f"Error getting workspace correlations: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacescanresults(self: 'SpiderFootWebUi', workspace_id: str, scan_id: str = None, event_type: str = None, limit: int = 100) -> dict:
        """Get scan results for workspace scans.

        Args:
            workspace_id (str): workspace ID
            scan_id (str): specific scan ID (optional)
            event_type (str): filter by event type (optional)
            limit (int): maximum results to return

        Returns:
            dict: scan results data
        """
        try:
            # Convert limit to integer if it's passed as string from HTTP request
            if isinstance(limit, str):
                try:
                    limit = int(limit)
                except (ValueError, TypeError):
                    limit = 100  # fallback to default

            # Ensure limit is positive and reasonable
            if not isinstance(limit, int) or limit <= 0:
                limit = 100
            elif limit > 10000:  # Cap at reasonable maximum
                limit = 10000

            workspace = SpiderFootWorkspace(self.config, workspace_id)
            dbh = SpiderFootDb(self.config)

            if scan_id:
                # Get results for specific scan
                scan_ids = [scan_id]
            else:
                # Get results for all workspace scans
                scan_ids = [scan['scan_id'] for scan in workspace.scans]

            all_results = []
            scan_summaries = {}

            for sid in scan_ids:
                # Get scan summary
                summary = dbh.scanResultSummary(sid, 'type')
                scan_summaries[sid] = summary

                # Get recent events
                if event_type:
                    events = dbh.scanResultEvent(sid, event_type, filterFp=False)
                else:
                    events = dbh.scanResultEvent(sid, 'ALL', filterFp=False)

                # Limit results per scan
                events = events[:limit] if events else []

                for event in events:
                    all_results.append({
                        'scan_id': sid,
                        'timestamp': event[0],
                        'event_type': event[1],
                        'event_data': event[2],
                        'source_module': event[3],
                        'source_event': event[4] if len(event) > 4 else '',
                        'false_positive': event[8] if len(event) > 8 else False
                    })

            return {
                'success': True,
                'results': all_results[:limit],  # Apply overall limit
                'scan_summaries': scan_summaries,
                'total_results': len(all_results),
                'workspace_id': workspace_id
            }
            
        except Exception as e:
            self.log.error(f"Error getting workspace scan results: {e}")
            return {'success': False, 'error': str(e)}